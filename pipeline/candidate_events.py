from __future__ import annotations

import hashlib
import json
import re
import sys
import traceback
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PIPELINE_VERSION = "candidate_event_generation_v0.11"
DASH_TRANSLATION = str.maketrans({"—": "-", "–": "-", "－": "-", "﹣": "-"})
FORBIDDEN_CONFIG_KEY_PARTS = (
    "target_pages", "correct_candidate", "expected_event", "known_investor",
    "known_company", "answer_mapping", "fixed_page", "fixed_event", "gold_answer",
)
DATE_PATTERNS = [
    re.compile(r"(?:19|20)\s*\d{2}\s*年(?:\s*\d{1,2}\s*月(?:\s*\d{1,2}\s*日)?)?"),
    re.compile(r"(?:19|20)\d{2}[./-]\d{1,2}(?:[./-]\d{1,2})?"),
]
HEADING_PATTERNS = [
    (1, re.compile(r"^第[一二三四五六七八九十百\d]+节")),
    (2, re.compile(r"^[一二三四五六七八九十百]+、")),
    (3, re.compile(r"^[（(][一二三四五六七八九十百]+[）)]")),
    (4, re.compile(r"^\d+[、.]")),
    (5, re.compile(r"^[（(]\d+[）)]")),
    (4, re.compile(r"^(?:19|20)\d{2}年")),
]


@dataclass
class FrozenRange:
    patch_id: str
    company_id: str
    company_short_name: str
    chapter_type: str
    final_start_pdf_page: int
    final_end_pdf_page: int
    final_start_printed_page_raw: str | None
    final_end_printed_page_raw: str | None
    printed_page_value_type: str
    disclosure_scope: str
    final_status: str
    decision: str
    source_run_id: str | None


@dataclass
class TextUnit:
    company_id: str
    pdf_page: int
    line_index: int
    segment_index: int
    text: str
    normalized_text: str
    global_index: int


@dataclass
class EventSection:
    section_id: str
    company_id: str
    source_patch_id: str
    source_kind: str
    title: str
    title_unit_index: int
    heading_level: int
    event_type_candidate: str
    event_type_candidates: list[str]
    explicit_combined_event: bool
    ordinal_labels: list[str]
    timeline_summary: bool
    units: list[TextUnit]
    date_roles: dict[str, list[str]]
    event_period: str | None
    event_date_primary_role: str | None
    event_date_selection_basis: str
    negative_disclosure: bool
    negative_reason: str | None
    entity_scope_candidate: str
    service_provider_mentions: list[str]
    signals: list[str]
    confidence: float
    review_reasons: list[str]


@dataclass
class CandidateEvidence:
    evidence_id: str
    candidate_event_id: str
    company_id: str
    source_patch_id: str
    evidence_role: str
    source_kind: str
    pdf_page_start: int
    pdf_page_end: int
    printed_page_start: str | None
    printed_page_end: str | None
    printed_page_value_type: str
    source_line_ranges: list[dict[str, int]]
    evidence_text: str
    evidence_sha256: str
    fragment_file: str
    matched_signals: list[str]


@dataclass
class CandidateEvent:
    candidate_event_id: str
    company_id: str
    company_short_name: str
    source_chapter_type: str
    source_patch_id: str
    disclosure_scope: str
    event_type_candidate: str
    event_type_candidates: list[str]
    ordinal_labels: list[str]
    event_period: str | None
    event_date_text: str | None
    event_date_primary_role: str | None
    event_date_selection_basis: str
    event_dates: dict[str, list[str]]
    event_title: str
    pdf_page_start: int
    pdf_page_end: int
    printed_page_start: str | None
    printed_page_end: str | None
    printed_page_value_type: str
    primary_evidence_id: str
    supporting_evidence_ids: list[str]
    entity_scope_candidate: str
    service_provider_mentions: list[str]
    matched_signals: list[str]
    candidate_confidence: float
    candidate_status: str
    review_required: bool
    review_reasons: list[str]


@dataclass
class NegativeDisclosure:
    negative_disclosure_id: str
    company_id: str
    source_patch_id: str
    event_type_candidate: str
    title: str
    pdf_page_start: int
    pdf_page_end: int
    printed_page_start: str | None
    printed_page_end: str | None
    printed_page_value_type: str
    evidence_text: str
    negative_reason: str
    record_status: str


@dataclass
class SummaryDisclosure:
    summary_disclosure_id: str
    company_id: str
    source_patch_id: str
    summary_kind: str
    event_type_candidates: list[str]
    title: str
    pdf_page_start: int
    pdf_page_end: int
    printed_page_start: str | None
    printed_page_end: str | None
    printed_page_value_type: str
    evidence_text: str
    evidence_sha256: str
    summary_reason: str
    record_status: str


@dataclass
class CoverageGap:
    coverage_gap_id: str
    company_id: str
    source_patch_id: str
    gap_type: str
    source_summary_disclosure_id: str
    title: str
    pdf_page_start: int
    pdf_page_end: int
    printed_page_start: str | None
    printed_page_end: str | None
    printed_page_value_type: str
    disclosed_event_counts: dict[str, int]
    represented_candidate_counts: dict[str, int]
    missing_event_counts: dict[str, int]
    evidence_text: str
    gap_reason: str
    record_status: str


@dataclass
class ShareholderEvidence:
    shareholder_evidence_id: str
    company_id: str
    company_short_name: str
    source_patch_id: str
    evidence_category: str
    pdf_page_start: int
    pdf_page_end: int
    printed_page_start: str | None
    printed_page_end: str | None
    printed_page_value_type: str
    evidence_text: str
    evidence_sha256: str
    matched_signals: list[str]
    evidence_status: str


def now_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def make_run_id() -> str:
    return datetime.now().astimezone().strftime("CANDIDATEEVENT_V11_%Y%m%d_%H%M%S")


def normalize_text(text: str) -> str:
    return re.sub(r"\s+", "", text.strip()).translate(DASH_TRANSLATION)


def sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8", errors="replace")).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while chunk := handle.read(1024 * 1024): digest.update(chunk)
    return digest.hexdigest()


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8-sig") as handle:
        for line_number, line in enumerate(handle, start=1):
            if not line.strip(): continue
            try: rows.append(json.loads(line))
            except json.JSONDecodeError as exc:
                raise ValueError(f"JSONL解析失败：{path} 第{line_number}行：{exc}") from exc
    return rows


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def write_jsonl(path: Path, rows: Iterable[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        for row in rows: handle.write(json.dumps(row, ensure_ascii=False) + "\n")


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8-sig"))


def validate_rule_config(payload: Any, location: str = "$") -> None:
    if isinstance(payload, dict):
        for key, value in payload.items():
            if any(part in key.lower() for part in FORBIDDEN_CONFIG_KEY_PARTS):
                raise ValueError(f"候选事件配置包含疑似答案映射字段：{location}.{key}")
            validate_rule_config(value, f"{location}.{key}")
    elif isinstance(payload, list):
        for index, value in enumerate(payload): validate_rule_config(value, f"{location}[{index}]")


def parse_filename(file_name: str) -> tuple[str | None, str | None]:
    match = re.match(r"^(?P<code>\d{6})_(?P<name>.+?)_IPO招股说明书\.pdf$", file_name, flags=re.IGNORECASE)
    return (match.group("code"), match.group("name").strip()) if match else (None, None)


def frozen_range_from_record(record: dict[str, Any]) -> FrozenRange:
    required = ["patch_id", "company_id", "company_short_name", "chapter_type", "final_start_pdf_page", "final_end_pdf_page", "printed_page_value_type", "disclosure_scope", "final_status", "decision"]
    missing = [key for key in required if record.get(key) in (None, "")]
    if missing: raise ValueError(f"章节Patch缺少字段 {missing}：{record.get('patch_id')}")
    start, end = int(record["final_start_pdf_page"]), int(record["final_end_pdf_page"])
    if start <= 0 or end < start: raise ValueError(f"章节Patch页码范围非法：{record.get('patch_id')} {start}-{end}")
    return FrozenRange(
        patch_id=str(record["patch_id"]), company_id=str(record["company_id"]), company_short_name=str(record["company_short_name"]),
        chapter_type=str(record["chapter_type"]), final_start_pdf_page=start, final_end_pdf_page=end,
        final_start_printed_page_raw=str(record["final_start_printed_page_raw"]) if record.get("final_start_printed_page_raw") is not None else None,
        final_end_printed_page_raw=str(record["final_end_printed_page_raw"]) if record.get("final_end_printed_page_raw") is not None else None,
        printed_page_value_type=str(record["printed_page_value_type"]), disclosure_scope=str(record["disclosure_scope"]),
        final_status=str(record["final_status"]), decision=str(record["decision"]),
        source_run_id=str(record["source_run_id"]) if record.get("source_run_id") else None,
    )


def load_frozen_ranges(path: Path) -> list[FrozenRange]:
    ranges = [frozen_range_from_record(row) for row in read_jsonl(path)]
    ranges = [item for item in ranges if item.chapter_type in {"equity_history", "shareholders"}]
    seen: set[tuple[str, str]] = set()
    for item in ranges:
        key = (item.company_id, item.chapter_type)
        if key in seen: raise ValueError(f"同公司同章节存在多个冻结Patch：{key}")
        seen.add(key)
    return ranges


def parse_printed_page(raw: str | None) -> tuple[str | None, int | None]:
    if raw is None: return None, None
    match = re.fullmatch(r"(?:(?P<prefix>\d+(?:-\d+)*)-)?(?P<number>\d+)", normalize_text(raw))
    return (match.group("prefix"), int(match.group("number"))) if match else (None, None)


def format_printed_page(prefix: str | None, number: int) -> str:
    return f"{prefix}-{number}" if prefix else str(number)


class PrintedPageResolver:
    def __init__(self, frozen_range: FrozenRange, pagination_patch: dict[str, Any] | None) -> None:
        self.frozen_range = frozen_range
        self.prefix: str | None = None
        self.offset: int | None = None
        self.value_type = frozen_range.printed_page_value_type
        if pagination_patch:
            self.prefix = str(pagination_patch["printed_page_prefix"]) if pagination_patch.get("printed_page_prefix") else None
            self.offset = int(pagination_patch["offset"])
            self.value_type = str(pagination_patch.get("value_type") or self.value_type)
        else:
            prefix, number = parse_printed_page(frozen_range.final_start_printed_page_raw)
            if number is not None:
                self.prefix, self.offset = prefix, frozen_range.final_start_pdf_page - number

    def resolve(self, pdf_page: int) -> str | None:
        if self.offset is None:
            if pdf_page == self.frozen_range.final_start_pdf_page: return self.frozen_range.final_start_printed_page_raw
            if pdf_page == self.frozen_range.final_end_pdf_page: return self.frozen_range.final_end_printed_page_raw
            return None
        number = pdf_page - self.offset
        return format_printed_page(self.prefix, number) if number > 0 else None


def load_page_text(path: Path) -> dict[int, str]:
    pages: dict[int, str] = {}
    for row in read_jsonl(path):
        page = int(row["pdf_page"])
        if page in pages: raise ValueError(f"逐页文本出现重复PDF页：{path} 页{page}")
        pages[page] = str(row.get("text") or "")
    return pages


def split_dense_line(text: str) -> list[str]:
    matches = list(DATE_PATTERNS[0].finditer(text))
    if len(matches) < 2 or len(text) < 120: return [text]
    segments: list[str] = []
    prefix = text[:matches[0].start()].strip()
    if prefix: segments.append(prefix)
    for index, match in enumerate(matches):
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        segment = text[match.start():end].strip()
        if segment: segments.append(segment)
    return segments or [text]


def build_text_units(company_id: str, pages: dict[int, str], start_page: int, end_page: int) -> list[TextUnit]:
    units: list[TextUnit] = []
    global_index = 0
    for pdf_page in range(start_page, end_page + 1):
        lines = [line.strip() for line in pages.get(pdf_page, "").splitlines() if line.strip()]
        for line_index, line in enumerate(lines):
            for segment_index, segment in enumerate(split_dense_line(line)):
                normalized = normalize_text(segment)
                if normalized:
                    units.append(TextUnit(company_id, pdf_page, line_index, segment_index, segment, normalized, global_index))
                    global_index += 1
    return units


def extract_dates(text: str) -> list[str]:
    values: list[str] = []
    for pattern in DATE_PATTERNS:
        for match in pattern.finditer(text):
            cleaned = re.sub(r"\s+", "", match.group(0))
            if cleaned not in values: values.append(cleaned)
    return values


def normalized_date_key(value: str) -> str:
    normalized = normalize_text(value)
    match = re.match(r"(?P<year>(?:19|20)\d{2})年(?:(?P<month>\d{1,2})月)?(?:(?P<day>\d{1,2})日)?", normalized)
    if match:
        parts = [match.group("year")]
        if match.group("month"): parts.append(match.group("month").zfill(2))
        if match.group("day"): parts.append(match.group("day").zfill(2))
        return "-".join(parts)
    match = re.match(r"(?P<year>(?:19|20)\d{2})[./-](?P<month>\d{1,2})(?:[./-](?P<day>\d{1,2}))?", normalized)
    if match:
        parts = [match.group("year"), match.group("month").zfill(2)]
        if match.group("day"): parts.append(match.group("day").zfill(2))
        return "-".join(parts)
    return normalized


def heading_level(text: str) -> int:
    normalized = normalize_text(text)
    for level, pattern in HEADING_PATTERNS:
        if pattern.search(normalized): return level
    return 9


def is_heading_like(unit: TextUnit, rules: dict[str, Any]) -> bool:
    if len(unit.normalized_text) > int(rules["section_detection"]["heading_max_chars"]): return False
    if heading_level(unit.text) < 9: return True
    return unit.normalized_text.endswith(tuple(rules["section_detection"]["heading_suffixes"]))


def longest_event_type_hits(normalized: str, rules: dict[str, Any]) -> tuple[list[str], list[str], bool]:
    # Trading venue names contain 股份转让 but are not equity-transfer events.
    venue_context = any(normalize_text(term) in normalized for term in rules["context_exclusions"]["equity_transfer_venue_terms"])
    matches: list[tuple[str, str, int]] = []
    for event_type in rules["event_type_priority"]:
        for term in rules["event_types"][event_type]["terms"]:
            nterm = normalize_text(term)
            if nterm and nterm in normalized:
                if event_type == "EQUITY_TRANSFER" and venue_context:
                    continue
                matches.append((event_type, term, len(nterm)))

        # Regex patterns allow issuer-predecessor names to appear between
        # "前身" and "设立", without putting company-specific answers in config.
        for pattern in rules.get("event_type_patterns", {}).get(event_type, []):
            match = re.search(pattern, normalized)
            if match:
                if event_type == "EQUITY_TRANSFER" and venue_context:
                    continue
                matched_text = match.group(0)
                matches.append((
                    event_type,
                    f"regex:{matched_text}",
                    len(matched_text),
                ))

    if not matches:
        return [], [], False
    matches.sort(key=lambda item: (-item[2], rules["event_type_priority"].index(item[0])))
    types: list[str] = []
    terms: list[str] = []
    for event_type, term, _ in matches:
        if event_type not in types:
            types.append(event_type); terms.append(term)

    # Specific types suppress broader substring matches.
    suppressions = rules["type_suppression"]
    selected: list[str] = []
    for event_type in types:
        suppressed = any(controller in types for controller in suppressions.get(event_type, []))
        if not suppressed: selected.append(event_type)

    explicit_combined = False
    if len(selected) > 1:
        combined_markers = rules["section_detection"]["combined_event_markers"]
        explicit_combined = any(
            marker in normalized
            for marker in map(normalize_text, combined_markers)
        ) or bool(
            re.search(
                r"(?:股权转让|股份转让).{0,12}(?:和|及|并).{0,12}增资"
                r"|增资.{0,12}(?:和|及|并).{0,12}(?:股权转让|股份转让)",
                normalized,
            )
        )
        if not explicit_combined:
            selected = [selected[0]]
    selected_terms = [term for etype, term, _ in matches if etype in selected]
    return selected, selected_terms, explicit_combined


def is_generic_summary_heading(normalized: str, rules: dict[str, Any]) -> bool:
    return any(normalize_text(term) in normalized for term in rules["section_detection"]["generic_summary_terms"])


def is_child_background_heading(normalized: str, rules: dict[str, Any]) -> bool:
    return any(normalize_text(term) in normalized for term in rules["section_detection"]["background_heading_terms"])


def service_provider_mentions(text: str, rules: dict[str, Any]) -> list[str]:
    normalized = normalize_text(text)
    mentions: list[str] = []
    for term in rules["entity_roles"]["service_provider_terms"]:
        if normalize_text(term) in normalized and term not in mentions:
            mentions.append(term)
    return mentions


def clean_other_entity_subject(
    name: str,
    rules: dict[str, Any],
) -> str:
    cleaned = normalize_text(name)
    cleaned = re.sub(r"^(?:19|20)\d{2}年", "", cleaned)
    changed = True
    prefixes = [
        normalize_text(value)
        for value in rules["entity_roles"].get(
            "subject_cleanup_prefixes", []
        )
    ]
    while changed and cleaned:
        changed = False
        for prefix in prefixes:
            if cleaned.startswith(prefix):
                cleaned = cleaned[len(prefix):]
                changed = True
                break
    for suffix in rules["entity_roles"].get(
        "subject_cleanup_suffixes", []
    ):
        normalized_suffix = normalize_text(suffix)
        while cleaned.endswith(normalized_suffix):
            cleaned = cleaned[:-len(normalized_suffix)]
    return cleaned


def is_generic_legal_form_subject(
    name: str,
    rules: dict[str, Any],
) -> bool:
    cleaned = clean_other_entity_subject(name, rules)
    if not cleaned:
        return True

    # Headings such as “（一）有限公司设立情况” may be captured as
    # “一有限公司”. They are legal-form labels, not company names.
    without_numbering = re.sub(
        r"^[（(]?[一二三四五六七八九十百\d]+[）)、.]?",
        "",
        cleaned,
    )
    legal_forms = {
        "公司",
        "有限公司",
        "有限责任公司",
        "股份有限公司",
        "股份公司",
        *{
            normalize_text(value)
            for value in rules["entity_roles"].get(
                "generic_issuer_subjects", []
            )
        },
    }
    return cleaned in legal_forms or without_numbering in legal_forms


def issuer_core_tokens(
    company_short_name: str,
    rules: dict[str, Any],
) -> list[str]:
    issuer = normalize_text(company_short_name)
    tokens: list[str] = []
    if issuer:
        tokens.append(issuer)
    reduced = issuer
    suffixes = sorted(
        [
            normalize_text(value)
            for value in rules["entity_roles"].get(
                "company_short_name_business_suffixes", []
            )
        ],
        key=len,
        reverse=True,
    )
    changed = True
    while changed and reduced:
        changed = False
        for suffix in suffixes:
            if reduced.endswith(suffix) and len(reduced) > len(suffix):
                reduced = reduced[:-len(suffix)]
                if reduced and reduced not in tokens:
                    tokens.append(reduced)
                changed = True
                break
    minimum = int(
        rules["entity_roles"].get(
            "minimum_predecessor_core_length", 2
        )
    )
    if len(reduced) >= minimum:
        tail = reduced[-minimum:]
        if tail not in tokens:
            tokens.append(tail)
    elif len(issuer) >= minimum:
        tail = issuer[-minimum:]
        if tail not in tokens:
            tokens.append(tail)
    return [token for token in tokens if len(token) >= minimum]


def relation_predecessor_names(
    text: str,
    rules: dict[str, Any],
) -> set[str]:
    normalized = normalize_text(text)
    names: set[str] = set()
    for pattern in rules["entity_roles"].get(
        "predecessor_relation_patterns", []
    ):
        for match in re.finditer(pattern, normalized):
            name = clean_other_entity_subject(
                match.group("name"), rules
            )
            if name:
                names.add(name)
    return names


def absorption_counterparty_names(
    text: str,
    rules: dict[str, Any],
) -> set[str]:
    normalized = normalize_text(text)
    names: set[str] = set()
    patterns = [
        r"(?:发行人|公司|本公司)(?:拟)?吸收合并(?P<name>[\u4e00-\u9fffA-Za-z0-9（）()]{2,30}?)(?:，|。|；|后|并|的|情况)",
        r"[\u4e00-\u9fffA-Za-z0-9（）()]{2,30}?(?:有限责任公司|有限公司|有限|股份公司)?吸收合并(?P<name>[\u4e00-\u9fffA-Za-z0-9（）()]{2,30}?)(?:，|。|；|注册资本|后|并|的|情况)",
        r"吸收合并对象为(?P<name>[\u4e00-\u9fffA-Za-z0-9（）()]{2,30}?)(?:，|。|；)",
        r"被吸收合并方(?:为)?(?P<name>[\u4e00-\u9fffA-Za-z0-9（）()]{2,30}?)(?:，|。|；)",
    ]
    for pattern in patterns:
        for match in re.finditer(pattern, normalized):
            name = clean_other_entity_subject(
                match.group("name"),
                rules,
            )
            if name:
                names.add(name)
    return names


def issuer_establishment_mentions(
    text: str,
    company_short_name: str,
    rules: dict[str, Any],
) -> set[str]:
    """Find issuer/predecessor establishment wording without substring leakage."""
    normalized = normalize_text(text)
    names: set[str] = set()
    legal_suffix = r"(?:有限责任公司|有限公司|有限)"
    verb = (
        r"(?:的)?(?:"
        r"成立(?!日期|时间)"
        r"|设立(?:情况|程序)?"
        r")"
    )
    prefix = (
        r"(?:^|[，。；、：:]|公司前身|发行人前身|"
        r"本公司前身|前身|"
        r"(?:19|20)\d{2}年(?:\d{1,2}月(?:\d{1,2}日)?)?)"
    )

    for core in issuer_core_tokens(
        company_short_name,
        rules,
    ):
        pattern = (
            prefix
            + r"(?P<name>"
            + re.escape(core)
            + r"[\u4e00-\u9fffA-Za-z0-9（）()]{0,12}?"
            + legal_suffix
            + r")"
            + verb
        )
        for match in re.finditer(pattern, normalized):
            name = clean_other_entity_subject(
                match.group("name"),
                rules,
            )
            if (
                name
                and subject_is_issuer_predecessor(
                    name,
                    company_short_name,
                    normalized,
                    rules,
                )
            ):
                names.add(name)

    return names


def establishment_subject_names(
    text: str,
    rules: dict[str, Any],
) -> set[str]:
    normalized = normalize_text(text)
    names: set[str] = set()
    pattern = (
        r"(?:^|[，。；、])"
        r"(?P<name>[\u4e00-\u9fffA-Za-z0-9（）()]{2,30}?"
        r"(?:有限责任公司|有限公司|有限))"
        r"(?:的)?(?:成立|设立情况|设立)"
    )
    for match in re.finditer(pattern, normalized):
        name = clean_other_entity_subject(
            match.group("name"),
            rules,
        )
        if name:
            names.add(name)
    return names


def establishment_is_absorption_counterparty(
    section_text: str,
    document_text: str,
    rules: dict[str, Any],
) -> bool:
    counterparties = absorption_counterparty_names(
        document_text,
        rules,
    )
    if not counterparties:
        return False
    for subject in establishment_subject_names(
        section_text,
        rules,
    ):
        for counterparty in counterparties:
            if (
                subject == counterparty
                or subject in counterparty
                or counterparty in subject
            ):
                return True
    return False


def establishment_is_issuer_predecessor(
    section_text: str,
    company_short_name: str,
    rules: dict[str, Any],
) -> bool:
    subjects = establishment_subject_names(
        section_text,
        rules,
    )
    return any(
        subject_is_issuer_predecessor(
            subject,
            company_short_name,
            section_text,
            rules,
        )
        for subject in subjects
    )


def subject_is_absorption_counterparty(
    subject: str,
    text: str,
    rules: dict[str, Any],
) -> bool:
    cleaned = clean_other_entity_subject(subject, rules)
    for counterparty in absorption_counterparty_names(
        text,
        rules,
    ):
        if (
            cleaned == counterparty
            or cleaned in counterparty
            or counterparty in cleaned
        ):
            return True
    return False


def subject_is_issuer_predecessor(
    subject: str,
    company_short_name: str,
    text: str,
    rules: dict[str, Any],
) -> bool:
    cleaned = clean_other_entity_subject(subject, rules)
    if not cleaned:
        return False
    relation_names = relation_predecessor_names(text, rules)
    if any(
        cleaned == name
        or cleaned in name
        or name in cleaned
        for name in relation_names
    ):
        return True

    legal_suffixes = [
        normalize_text(value)
        for value in rules["entity_roles"].get(
            "predecessor_legal_suffixes", []
        )
    ]
    if not any(cleaned.endswith(suffix) for suffix in legal_suffixes):
        return False
    return any(
        core in cleaned
        for core in issuer_core_tokens(
            company_short_name, rules
        )
    )


def other_entity_subjects_from_text(
    text: str,
    company_short_name: str,
    rules: dict[str, Any],
) -> list[str]:
    normalized = normalize_text(text)
    issuer = normalize_text(company_short_name)
    generic = {
        normalize_text(value)
        for value in rules["entity_roles"].get(
            "generic_issuer_subjects", []
        )
    }
    service_terms = [
        normalize_text(value)
        for value in rules["entity_roles"]["service_provider_terms"]
    ]
    subjects: list[str] = []

    for pattern in rules["entity_roles"].get(
        "other_entity_subject_patterns", []
    ):
        for match in re.finditer(pattern, normalized):
            name = clean_other_entity_subject(
                match.group("name"), rules
            )
            if (
                not name
                or name in generic
                or is_generic_legal_form_subject(
                    name,
                    rules,
                )
            ):
                continue
            if issuer and (issuer in name or name in issuer):
                continue
            if subject_is_absorption_counterparty(
                name,
                normalized,
                rules,
            ):
                if name not in subjects:
                    subjects.append(name)
                continue
            if subject_is_issuer_predecessor(
                name,
                company_short_name,
                normalized,
                rules,
            ):
                continue
            if any(term in name for term in service_terms):
                continue
            if name not in subjects:
                subjects.append(name)
    return subjects


def service_provider_only_unit(
    unit: TextUnit,
    event_types: list[str],
    rules: dict[str, Any],
) -> bool:
    if not service_provider_mentions(unit.text, rules):
        return False
    normalized = unit.normalized_text
    action_terms = [
        normalize_text(term)
        for term in rules["entity_roles"]["service_provider_action_terms"]
    ]
    transaction_terms = [
        normalize_text(term)
        for term in rules["entity_roles"]["issuer_transaction_terms"]
    ]
    support_terms = [
        normalize_text(term)
        for term in rules["entity_roles"].get(
            "service_provider_support_terms", []
        )
    ]
    has_issuer_transaction = any(
        term in normalized for term in transaction_terms
    )
    if has_issuer_transaction:
        return False
    return (
        any(term in normalized for term in action_terms)
        or any(term in normalized for term in support_terms)
    )


def negative_reason_from_units(units: list[TextUnit], rules: dict[str, Any]) -> str | None:
    probe = "".join(unit.normalized_text for unit in units[:int(rules["negative_disclosure"]["probe_units"])])
    for pattern in rules["negative_disclosure"]["patterns"]:
        if re.search(pattern, probe): return pattern
    return None


def extract_ordinal_labels(text: str) -> list[str]:
    normalized = normalize_text(text)
    labels: list[str] = []
    for match in re.finditer(
        r"第(?P<label>[一二三四五六七八九十百\d]+)次"
        r"(?:股权转让|股份转让|增资|减资|资本公积转增股本|股票发行)?",
        normalized,
    ):
        label = f"第{match.group('label')}次"
        if label not in labels:
            labels.append(label)
    return labels


def contextual_units_after(
    unit: TextUnit,
    units: list[TextUnit],
    count: int,
) -> list[TextUnit]:
    by_index = {
        item.global_index: item for item in units
    }
    selected: list[TextUnit] = []
    for index in range(
        unit.global_index,
        unit.global_index + count + 1,
    ):
        item = by_index.get(index)
        if item is None:
            continue
        if item.pdf_page != unit.pdf_page:
            break
        selected.append(item)
    return selected


def is_service_provider_aggregate_history(
    unit: TextUnit,
    units: list[TextUnit],
    rules: dict[str, Any],
) -> bool:
    probe_units = contextual_units_after(
        unit,
        units,
        int(
            rules["summary_disclosures"].get(
                "service_provider_probe_units",
                6,
            )
        ),
    )
    probe = "".join(
        item.normalized_text for item in probe_units
    )
    if not probe:
        return False

    has_service_provider = bool(
        service_provider_mentions(probe, rules)
    )
    has_report_context = any(
        normalize_text(term) in probe
        for term in rules["summary_disclosures"].get(
            "service_provider_report_terms",
            [],
        )
    )
    has_aggregate_count = bool(
        disclosed_event_counts(probe)
    )
    return (
        has_service_provider
        and has_report_context
        and has_aggregate_count
    )


def is_timeline_summary_row(
    unit: TextUnit,
    rules: dict[str, Any],
) -> bool:
    normalized = unit.normalized_text
    if any(
        re.search(pattern, normalized)
        for pattern in rules["summary_disclosures"][
            "timeline_row_patterns"
        ]
    ):
        return True
    return False


def is_aggregate_summary_statement(
    unit: TextUnit,
    rules: dict[str, Any],
) -> bool:
    normalized = unit.normalized_text
    return any(
        re.search(pattern, normalized)
        for pattern in rules["summary_disclosures"][
            "aggregate_summary_patterns"
        ]
    )


def is_metadata_date_field(
    unit: TextUnit,
    rules: dict[str, Any],
) -> bool:
    normalized = unit.normalized_text
    return any(
        re.search(pattern, normalized)
        for pattern in rules["summary_disclosures"][
            "metadata_date_patterns"
        ]
    )


def is_cross_reference_only(
    units: list[TextUnit],
    rules: dict[str, Any],
) -> bool:
    probe_units = units[: int(
        rules["summary_disclosures"]["cross_reference_probe_units"]
    )]
    probe = "".join(unit.normalized_text for unit in probe_units)
    if not probe:
        return False
    pattern_hit = any(
        re.search(pattern, probe)
        for pattern in rules["summary_disclosures"][
            "cross_reference_patterns"
        ]
    )
    if not pattern_hit:
        return False
    detail_hit = any(
        re.search(pattern, probe)
        for pattern in rules["summary_disclosures"][
            "transaction_detail_patterns"
        ]
    )
    return not detail_hit


def is_classification_wrapper_heading(
    unit: TextUnit,
    rules: dict[str, Any],
) -> bool:
    normalized = unit.normalized_text
    return any(
        normalize_text(term) in normalized
        for term in rules["summary_disclosures"][
            "classification_wrapper_terms"
        ]
    )


def is_procedural_narrative(
    unit: TextUnit,
    event_types: list[str],
    rules: dict[str, Any],
) -> bool:
    normalized = unit.normalized_text
    if not re.match(r"^(?:19|20)\d{2}年", normalized):
        return False
    if extract_ordinal_labels(normalized):
        return False
    if any(
        normalize_text(marker) in normalized
        for marker in rules["section_detection"][
            "standalone_dated_event_markers"
        ]
    ):
        return False
    return any(
        normalize_text(term) in normalized
        for term in rules["section_detection"][
            "procedural_narrative_terms"
        ]
    )


def _first_sorted(values: list[str]) -> str | None:
    return sorted(values)[0] if values else None


def choose_primary_event_date(
    event_type: str,
    title_dates: list[str],
    roles: dict[str, list[str]],
    rules: dict[str, Any],
) -> tuple[str | None, str | None, str]:
    normalized_title_dates = [
        normalized_date_key(value) for value in title_dates
    ]
    if normalized_title_dates:
        return (
            normalized_title_dates[0],
            "event_period",
            "TITLE_DATE",
        )

    type_preferences = rules["event_date_selection"].get(
        event_type,
        rules["date_role_precedence"],
    )
    for role in type_preferences:
        values = roles.get(role, [])
        if not values:
            continue
        value = (
            _first_sorted(values)
            if event_type in {
                "LIMITED_COMPANY_ESTABLISHMENT",
                "JOINT_STOCK_COMPANY_ESTABLISHMENT",
                "OVERALL_CHANGE",
            }
            else values[0]
        )
        return value, role, f"ROLE_PRIORITY:{role}"

    if event_type in set(
        rules["review"]["undated_event_types_allowed"]
    ):
        return None, None, "UNDATED_DISCLOSURE_ALLOWED"
    return None, None, "NO_RELEVANT_DATE"


def classify_date_roles(
    units: list[TextUnit],
    title: str,
    event_type: str,
    rules: dict[str, Any],
) -> tuple[
    dict[str, list[str]],
    str | None,
    str | None,
    str,
]:
    roles: dict[str, list[str]] = defaultdict(list)
    title_dates = extract_dates(title)
    for date in title_dates:
        key = normalized_date_key(date)
        if key not in roles["event_period"]:
            roles["event_period"].append(key)

    for unit in units:
        dates = extract_dates(unit.text)
        if not dates:
            continue
        normalized = unit.normalized_text
        role = "other_date"
        for candidate_role, terms in rules["date_roles"].items():
            if any(
                normalize_text(term) in normalized
                for term in terms
            ):
                role = candidate_role
                break
        for date in dates:
            key = normalized_date_key(date)
            if key not in roles[role]:
                roles[role].append(key)

    period, primary_role, basis = choose_primary_event_date(
        event_type,
        title_dates,
        dict(roles),
        rules,
    )
    return dict(roles), period, primary_role, basis


def entity_scope_from_section(
    units: list[TextUnit],
    company_short_name: str,
    rules: dict[str, Any],
) -> tuple[str, list[str]]:
    text = "".join(unit.normalized_text for unit in units[:40])
    services = service_provider_mentions(text, rules)
    if any(
        normalize_text(term) in text
        for term in rules["entity_roles"]["subsidiary_markers"]
    ):
        return "SUBSIDIARY_OR_OTHER_ENTITY_RISK", services
    if other_entity_subjects_from_text(
        text,
        company_short_name,
        rules,
    ):
        return "SUBSIDIARY_OR_OTHER_ENTITY_RISK", services
    return "ISSUER_OR_UNRESOLVED", services


def is_structural_boundary_heading(
    unit: TextUnit,
    rules: dict[str, Any],
) -> bool:
    if not is_heading_like(unit, rules):
        return False

    normalized = unit.normalized_text
    event_types, _, _ = longest_event_type_hits(
        normalized,
        rules,
    )

    if is_timeline_summary_row(unit, rules):
        return False
    if is_metadata_date_field(unit, rules):
        return False
    if is_aggregate_summary_statement(unit, rules):
        return False
    if event_types and is_procedural_narrative(
        unit,
        event_types,
        rules,
    ):
        return False

    if re.match(r"^(?:19|20)\d{2}年", normalized):
        if event_types and service_provider_only_unit(
            unit,
            event_types,
            rules,
        ):
            return False
        return bool(event_types)

    return True


def structural_heading_indexes(
    units: list[TextUnit],
    rules: dict[str, Any],
) -> list[int]:
    return [
        unit.global_index
        for unit in units
        if is_structural_boundary_heading(unit, rules)
    ]


def detect_root_event_headings(
    units: list[TextUnit],
    rules: dict[str, Any],
) -> list[
    tuple[TextUnit, list[str], list[str], bool]
]:
    roots: list[
        tuple[TextUnit, list[str], list[str], bool]
    ] = []
    recent_root_by_type: dict[str, TextUnit] = {}

    for unit in units:
        if not is_heading_like(unit, rules):
            continue

        event_types, terms, combined = longest_event_type_hits(
            unit.normalized_text,
            rules,
        )
        if not event_types:
            continue
        if is_generic_summary_heading(
            unit.normalized_text,
            rules,
        ):
            continue
        if is_child_background_heading(
            unit.normalized_text,
            rules,
        ):
            continue
        if is_timeline_summary_row(unit, rules):
            continue
        if is_metadata_date_field(unit, rules):
            continue
        if is_aggregate_summary_statement(unit, rules):
            continue
        if is_classification_wrapper_heading(unit, rules):
            continue
        if is_cross_reference_only([unit], rules):
            continue
        if is_procedural_narrative(
            unit,
            event_types,
            rules,
        ):
            continue
        if service_provider_only_unit(
            unit,
            event_types,
            rules,
        ):
            continue

        stripped = unit.normalized_text.lstrip(
            "1234567890、.（）()一二三四五六七八九十"
        )
        if (
            not extract_dates(unit.text)
            and any(
                normalize_text(term) == stripped
                for term in terms
            )
        ):
            prior = recent_root_by_type.get(event_types[0])
            if (
                prior
                and unit.pdf_page - prior.pdf_page
                <= int(
                    rules["section_detection"][
                        "generic_child_attach_page_gap"
                    ]
                )
            ):
                continue

        roots.append((unit, event_types, terms, combined))
        for event_type in event_types:
            recent_root_by_type[event_type] = unit

    return roots


def section_units_for_heading(root: TextUnit, units: list[TextUnit], heading_indexes: list[int], rules: dict[str, Any]) -> list[TextUnit]:
    root_level = heading_level(root.text)
    end_index = units[-1].global_index + 1
    for candidate_index in heading_indexes:
        if candidate_index <= root.global_index: continue
        candidate = units[candidate_index]
        if heading_level(candidate.text) <= root_level:
            end_index = candidate_index; break
    max_units = int(rules["section_detection"]["max_section_units"])
    selected = [unit for unit in units if root.global_index <= unit.global_index < end_index]
    return selected[:max_units]


def make_summary_disclosure(
    frozen_range: FrozenRange,
    unit: TextUnit,
    units: list[TextUnit],
    summary_kind: str,
    event_types: list[str],
    reason: str,
    sequence: int,
    rules: dict[str, Any],
) -> SummaryDisclosure:
    by_index = {item.global_index: item for item in units}
    after = int(
        rules["summary_disclosures"]["context_after_units"]
    )
    selected = [
        by_index[index]
        for index in range(
            unit.global_index,
            min(
                max(by_index),
                unit.global_index + after,
            ) + 1,
        )
        if index in by_index
    ]
    text = "\n".join(item.text for item in selected).strip()
    return SummaryDisclosure(
        summary_disclosure_id=(
            f"SUM-{frozen_range.company_id}-{sequence:04d}"
        ),
        company_id=frozen_range.company_id,
        source_patch_id=frozen_range.patch_id,
        summary_kind=summary_kind,
        event_type_candidates=event_types,
        title=unit.text,
        pdf_page_start=min(item.pdf_page for item in selected),
        pdf_page_end=max(item.pdf_page for item in selected),
        printed_page_start=None,
        printed_page_end=None,
        printed_page_value_type=(
            frozen_range.printed_page_value_type
        ),
        evidence_text=text,
        evidence_sha256=sha256_text(text),
        summary_reason=reason,
        record_status="SUMMARY_DISCLOSURE_NOT_BUSINESS_CANDIDATE",
    )


def collect_summary_disclosures(
    frozen_range: FrozenRange,
    units: list[TextUnit],
    rules: dict[str, Any],
) -> list[SummaryDisclosure]:
    output: list[SummaryDisclosure] = []
    seen: set[tuple[str, int, str]] = set()

    for unit in units:
        event_types, _, _ = longest_event_type_hits(
            unit.normalized_text,
            rules,
        )
        if not event_types:
            continue

        kind: str | None = None
        reason: str | None = None
        if is_service_provider_aggregate_history(
            unit,
            units,
            rules,
        ):
            kind = "SERVICE_PROVIDER_AGGREGATE_HISTORY"
            reason = (
                "中介机构报告汇总确认多个历史事件；"
                "不作为单一业务事件候选"
            )
        elif is_aggregate_summary_statement(unit, rules):
            kind = "AGGREGATE_EVENT_SUMMARY"
            reason = "汇总披露仅说明事件次数或概况，不代表单一可抽取事件"
        elif is_timeline_summary_row(unit, rules):
            kind = "TIMELINE_SUMMARY_ROW"
            reason = "时间线摘要行只作为详细事件的辅助定位证据"
        elif is_metadata_date_field(unit, rules):
            kind = "METADATA_DATE_FIELD"
            reason = "公司基本信息中的日期字段只作为日期辅助证据"
        elif is_cross_reference_only([unit], rules):
            kind = "CROSS_REFERENCE_ONLY"
            reason = "仅引用其他章节或文件，缺少本段交易细节"
        elif service_provider_only_unit(
            unit, event_types, rules
        ):
            kind = "SERVICE_PROVIDER_PROCEDURE"
            reason = rules["summary_disclosures"][
                "service_provider_procedure_reason"
            ]
        elif is_classification_wrapper_heading(unit, rules):
            kind = "CLASSIFICATION_WRAPPER"
            reason = "分类性标题由下级详细事件承载，不单独生成业务候选"

        if kind is None or reason is None:
            continue
        key = (kind, unit.pdf_page, unit.normalized_text)
        if key in seen:
            continue
        seen.add(key)
        output.append(make_summary_disclosure(
            frozen_range,
            unit,
            units,
            kind,
            event_types,
            reason,
            len(output) + 1,
            rules,
        ))
    return output


CHINESE_COUNT_VALUES = {
    "零": 0,
    "一": 1,
    "二": 2,
    "两": 2,
    "三": 3,
    "四": 4,
    "五": 5,
    "六": 6,
    "七": 7,
    "八": 8,
    "九": 9,
    "十": 10,
}


def parse_disclosed_count(value: str) -> int | None:
    normalized = normalize_text(value)
    if normalized.isdigit():
        return int(normalized)
    if normalized in CHINESE_COUNT_VALUES:
        return CHINESE_COUNT_VALUES[normalized]
    if normalized.startswith("十") and len(normalized) == 2:
        return 10 + CHINESE_COUNT_VALUES.get(normalized[1], 0)
    if normalized.endswith("十") and len(normalized) == 2:
        return CHINESE_COUNT_VALUES.get(normalized[0], 0) * 10
    if "十" in normalized and len(normalized) == 3:
        left, right = normalized.split("十", 1)
        return (
            CHINESE_COUNT_VALUES.get(left, 0) * 10
            + CHINESE_COUNT_VALUES.get(right, 0)
        )
    return None


def disclosed_event_counts(text: str) -> dict[str, int]:
    normalized = normalize_text(text)
    patterns = {
        "EQUITY_TRANSFER": [
            r"(?P<count>[一二两三四五六七八九十\d]+)次股权转让",
            r"(?P<count>[一二两三四五六七八九十\d]+)次股份转让",
        ],
        "CAPITAL_INCREASE": [
            r"(?P<count>[一二两三四五六七八九十\d]+)次增资扩股",
            r"(?P<count>[一二两三四五六七八九十\d]+)次增资",
        ],
        "SHARE_CAPITAL_CHANGE": [
            r"(?P<count>[一二两三四五六七八九十\d]+)次资本公积转增股本",
        ],
        "LISTING_OR_DIRECTIONAL_FINANCING": [
            r"共进行过(?P<count>[一二两三四五六七八九十\d]+)次定向发行",
        ],
    }
    output: dict[str, int] = {}
    for event_type, event_patterns in patterns.items():
        values: list[int] = []
        for pattern in event_patterns:
            for match in re.finditer(pattern, normalized):
                count = parse_disclosed_count(match.group("count"))
                if count is not None:
                    values.append(count)
        if values:
            output[event_type] = max(values)
    return output


def build_coverage_gaps(
    frozen_range: FrozenRange,
    summaries: list[SummaryDisclosure],
    candidates: list[CandidateEvent],
) -> list[CoverageGap]:
    represented = {
        event_type: sum(
            event_type in candidate.event_type_candidates
            for candidate in candidates
        )
        for event_type in {
            event_type
            for candidate in candidates
            for event_type in candidate.event_type_candidates
        }
    }

    gaps: list[CoverageGap] = []
    seen_disclosed_signatures: set[
        tuple[tuple[str, int], ...]
    ] = set()
    for summary in summaries:
        if summary.summary_kind not in {
            "AGGREGATE_EVENT_SUMMARY",
            "TIMELINE_SUMMARY_ROW",
            "SERVICE_PROVIDER_AGGREGATE_HISTORY",
        }:
            continue

        disclosed = disclosed_event_counts(
            summary.title + "\n" + summary.evidence_text
        )
        if not disclosed:
            continue

        signature = tuple(sorted(disclosed.items()))
        if signature in seen_disclosed_signatures:
            continue
        seen_disclosed_signatures.add(signature)

        represented_for_summary = {
            event_type: represented.get(event_type, 0)
            for event_type in disclosed
        }
        missing = {
            event_type: count - represented_for_summary[event_type]
            for event_type, count in disclosed.items()
            if count > represented_for_summary[event_type]
        }
        if not missing:
            continue

        reference_limited = bool(
            re.search(
                r"参见|详见|具体见|公开转让说明书|申报文件",
                normalize_text(summary.evidence_text),
            )
        )
        gap_type = (
            "REFERENCE_LIMITED_AGGREGATE_HISTORY"
            if reference_limited
            else "AGGREGATE_MULTI_EVENT_NOT_EXPANDED"
        )
        gaps.append(CoverageGap(
            coverage_gap_id=(
                f"GAP-{frozen_range.company_id}-"
                f"{len(gaps)+1:04d}"
            ),
            company_id=frozen_range.company_id,
            source_patch_id=frozen_range.patch_id,
            gap_type=gap_type,
            source_summary_disclosure_id=(
                summary.summary_disclosure_id
            ),
            title=summary.title,
            pdf_page_start=summary.pdf_page_start,
            pdf_page_end=summary.pdf_page_end,
            printed_page_start=summary.printed_page_start,
            printed_page_end=summary.printed_page_end,
            printed_page_value_type=(
                summary.printed_page_value_type
            ),
            disclosed_event_counts=disclosed,
            represented_candidate_counts=(
                represented_for_summary
            ),
            missing_event_counts=missing,
            evidence_text=summary.evidence_text,
            gap_reason=(
                "原文披露多个历史事件，但当前冻结范围内未逐项展开。"
                "该记录用于防止静默漏项；招股书未披露的单事件信息不得编造。"
            ),
            record_status="DISCLOSURE_COVERAGE_GAP",
        ))
    return gaps


def build_heading_sections(
    frozen_range: FrozenRange,
    units: list[TextUnit],
    rules: dict[str, Any],
) -> tuple[
    list[EventSection],
    list[NegativeDisclosure],
]:
    if not units:
        return [], []

    headings = structural_heading_indexes(units, rules)
    sections: list[EventSection] = []
    negatives: list[NegativeDisclosure] = []
    document_text = "".join(
        unit.normalized_text for unit in units
    )

    # Classification wrapper headings are not business candidates, but their
    # following body may contain an explicit negative disclosure that must be
    # preserved separately.
    for unit in units:
        if not is_classification_wrapper_heading(unit, rules):
            continue
        event_types, _, _ = longest_event_type_hits(
            unit.normalized_text,
            rules,
        )
        if not event_types:
            continue
        wrapper_units = section_units_for_heading(
            unit,
            units,
            headings,
            rules,
        )
        negative_reason = negative_reason_from_units(
            wrapper_units,
            rules,
        )
        if not negative_reason:
            continue
        text = "\n".join(
            item.text for item in wrapper_units
        )
        negatives.append(NegativeDisclosure(
            negative_disclosure_id=(
                f"NEG-{frozen_range.company_id}-"
                f"{len(negatives)+1:04d}"
            ),
            company_id=frozen_range.company_id,
            source_patch_id=frozen_range.patch_id,
            event_type_candidate=event_types[0],
            title=unit.text,
            pdf_page_start=min(
                item.pdf_page for item in wrapper_units
            ),
            pdf_page_end=max(
                item.pdf_page for item in wrapper_units
            ),
            printed_page_start=None,
            printed_page_end=None,
            printed_page_value_type=(
                frozen_range.printed_page_value_type
            ),
            evidence_text=text,
            negative_reason=negative_reason,
            record_status=(
                "NEGATIVE_DISCLOSURE_NOT_CANDIDATE"
            ),
        ))

    for sequence, (
        root,
        event_types,
        terms,
        combined,
    ) in enumerate(
        detect_root_event_headings(units, rules),
        start=1,
    ):
        nested_parent: EventSection | None = None
        for section in sections:
            if (
                section.units
                and section.units[0].global_index
                < root.global_index
                <= section.units[-1].global_index
            ):
                nested_parent = section
                break

        if nested_parent is not None:
            child_type = event_types[0]
            if (
                nested_parent.event_type_candidate
                == "ABSORPTION_MERGER"
                or (
                    nested_parent.explicit_combined_event
                    and child_type
                    in nested_parent.event_type_candidates
                )
                or child_type
                == nested_parent.event_type_candidate
            ):
                continue

        section_units = section_units_for_heading(
            root,
            units,
            headings,
            rules,
        )
        negative_reason = negative_reason_from_units(
            section_units,
            rules,
        )
        (
            date_roles,
            event_period,
            primary_date_role,
            date_basis,
        ) = classify_date_roles(
            section_units,
            root.text,
            event_types[0],
            rules,
        )
        entity_scope, services = entity_scope_from_section(
            section_units,
            frozen_range.company_short_name,
            rules,
        )
        text = "\n".join(
            unit.text for unit in section_units
        )
        if (
            event_types[0]
            == "LIMITED_COMPANY_ESTABLISHMENT"
        ):
            if establishment_is_absorption_counterparty(
                text,
                document_text,
                rules,
            ):
                entity_scope = (
                    "SUBSIDIARY_OR_OTHER_ENTITY_RISK"
                )
            elif establishment_is_issuer_predecessor(
                text,
                frozen_range.company_short_name,
                rules,
            ):
                entity_scope = "ISSUER_OR_UNRESOLVED"

        if negative_reason:
            negatives.append(NegativeDisclosure(
                negative_disclosure_id=(
                    f"NEG-{frozen_range.company_id}-"
                    f"{len(negatives)+1:04d}"
                ),
                company_id=frozen_range.company_id,
                source_patch_id=frozen_range.patch_id,
                event_type_candidate=event_types[0],
                title=root.text,
                pdf_page_start=min(
                    unit.pdf_page for unit in section_units
                ),
                pdf_page_end=max(
                    unit.pdf_page for unit in section_units
                ),
                printed_page_start=None,
                printed_page_end=None,
                printed_page_value_type=(
                    frozen_range.printed_page_value_type
                ),
                evidence_text=text,
                negative_reason=negative_reason,
                record_status=(
                    "NEGATIVE_DISCLOSURE_NOT_CANDIDATE"
                ),
            ))
            continue

        signals = [
            "event_heading",
            *[
                f"matched_term:{term}"
                for term in terms
            ],
        ]
        if combined:
            signals.append("explicit_combined_event")
        if services:
            signals.append(
                "service_provider_supporting_evidence"
            )

        confidence = (
            0.76
            + (0.08 if event_period else 0)
            + (0.06 if len(section_units) >= 3 else 0)
        )
        review: list[str] = []
        if (
            not event_period
            and event_types[0]
            not in set(
                rules["review"][
                    "undated_event_types_allowed"
                ]
            )
        ):
            review.append("事件段落未识别到主要事件日期")
        if entity_scope == "SUBSIDIARY_OR_OTHER_ENTITY_RISK":
            review.append("事件发生主体可能为子公司或其他主体")

        sections.append(EventSection(
            section_id=(
                f"SEC-{frozen_range.company_id}-H-"
                f"{sequence:04d}"
            ),
            company_id=frozen_range.company_id,
            source_patch_id=frozen_range.patch_id,
            source_kind="HEADING_SECTION",
            title=root.text,
            title_unit_index=root.global_index,
            heading_level=heading_level(root.text),
            event_type_candidate=event_types[0],
            event_type_candidates=event_types,
            explicit_combined_event=combined,
            ordinal_labels=extract_ordinal_labels(
                root.text
            ),
            timeline_summary=False,
            units=section_units,
            date_roles=date_roles,
            event_period=event_period,
            event_date_primary_role=primary_date_role,
            event_date_selection_basis=date_basis,
            negative_disclosure=False,
            negative_reason=None,
            entity_scope_candidate=entity_scope,
            service_provider_mentions=services,
            signals=signals,
            confidence=round(min(confidence, 0.98), 3),
            review_reasons=review,
        ))
    return sections, negatives


def unit_is_covered(unit: TextUnit, sections: list[EventSection]) -> bool:
    return any(section.units and section.units[0].global_index <= unit.global_index <= section.units[-1].global_index for section in sections)


def timeline_event_is_represented(
    unit: TextUnit,
    event_type: str,
    existing: list[EventSection],
    rules: dict[str, Any],
) -> bool:
    """Suppress a timeline row only when the same event is already represented.

    v0.5 suppressed a timeline row whenever any event of the same type existed
    anywhere in the chapter. That could hide an earlier establishment event.
    """
    unit_periods = [
        normalized_date_key(value)
        for value in extract_dates(unit.text)
    ]
    unit_ordinals = set(extract_ordinal_labels(unit.text))

    for section in existing:
        if not event_types_semantically_equivalent(
            section.event_type_candidate,
            event_type,
            rules,
        ):
            continue

        section_ordinals = set(section.ordinal_labels)
        if unit_ordinals and section_ordinals:
            if unit_ordinals.intersection(section_ordinals):
                return True
            continue

        if unit_periods and section.event_period:
            if any(
                period_compatible(period, section.event_period)
                for period in unit_periods
            ):
                return True
            continue

        if (
            not unit_periods
            and section_similarity(
                EventSection(
                    section_id="TIMELINE-PROBE",
                    company_id=section.company_id,
                    source_patch_id=section.source_patch_id,
                    source_kind="TIMELINE_ONLY",
                    title=unit.text,
                    title_unit_index=unit.global_index,
                    heading_level=9,
                    event_type_candidate=event_type,
                    event_type_candidates=[event_type],
                    explicit_combined_event=False,
                    ordinal_labels=list(unit_ordinals),
                    timeline_summary=True,
                    units=[unit],
                    date_roles={},
                    event_period=None,
                    event_date_primary_role=None,
                    event_date_selection_basis="TIMELINE_PROBE",
                    negative_disclosure=False,
                    negative_reason=None,
                    entity_scope_candidate="ISSUER_OR_UNRESOLVED",
                    service_provider_mentions=[],
                    signals=[],
                    confidence=0.0,
                    review_reasons=[],
                ),
                section,
            )
            >= float(
                rules["summary_disclosures"].get(
                    "timeline_existing_title_similarity",
                    0.72,
                )
            )
        ):
            return True

    return False


def build_fallback_sections(
    frozen_range: FrozenRange,
    units: list[TextUnit],
    existing: list[EventSection],
    rules: dict[str, Any],
) -> tuple[list[EventSection], dict[str, int]]:
    sections: list[EventSection] = []
    diagnostics = {
        "excluded_context_count": 0,
        "procedural_line_suppressed_count": 0,
        "timeline_row_suppressed_count": 0,
        "metadata_date_suppressed_count": 0,
        "cross_reference_suppressed_count": 0,
        "aggregate_summary_suppressed_count": 0,
        "classification_wrapper_suppressed_count": 0,
    }
    by_index = {unit.global_index: unit for unit in units}
    document_text = "".join(
        unit.normalized_text for unit in units
    )

    for unit in units:
        if unit_is_covered(unit, existing):
            continue

        event_types, terms, combined = longest_event_type_hits(
            unit.normalized_text,
            rules,
        )
        if not event_types:
            continue
        if negative_reason_from_units([unit], rules):
            continue

        if is_aggregate_summary_statement(unit, rules):
            diagnostics[
                "aggregate_summary_suppressed_count"
            ] += 1
            continue
        if is_metadata_date_field(unit, rules):
            diagnostics[
                "metadata_date_suppressed_count"
            ] += 1
            continue
        if is_cross_reference_only([unit], rules):
            diagnostics[
                "cross_reference_suppressed_count"
            ] += 1
            continue
        if is_classification_wrapper_heading(unit, rules):
            diagnostics[
                "classification_wrapper_suppressed_count"
            ] += 1
            continue
        if is_service_provider_aggregate_history(
            unit,
            units,
            rules,
        ):
            diagnostics[
                "procedural_line_suppressed_count"
            ] += 1
            continue
        if is_procedural_narrative(
            unit,
            event_types,
            rules,
        ):
            diagnostics[
                "procedural_line_suppressed_count"
            ] += 1
            continue

        timeline_summary = is_timeline_summary_row(
            unit,
            rules,
        )
        if timeline_summary:
            if timeline_event_is_represented(
                unit,
                event_types[0],
                existing,
                rules,
            ):
                diagnostics[
                    "timeline_row_suppressed_count"
                ] += 1
                continue

        dates = extract_dates(unit.text)
        if not dates:
            continue
        if service_provider_only_unit(
            unit,
            event_types,
            rules,
        ):
            diagnostics["excluded_context_count"] += 1
            continue
        if (
            any(
                normalize_text(term) in unit.normalized_text
                for term in rules["context_exclusions"][
                    "equity_transfer_venue_terms"
                ]
            )
            and event_types[0] == "EQUITY_TRANSFER"
        ):
            diagnostics["excluded_context_count"] += 1
            continue

        before = int(rules["fallback"]["before_units"])
        after = int(rules["fallback"]["after_units"])
        fragment = [
            by_index[index]
            for index in range(
                max(min(by_index), unit.global_index - before),
                min(max(by_index), unit.global_index + after)
                + 1,
            )
            if index in by_index
        ]
        (
            date_roles,
            event_period,
            primary_date_role,
            date_basis,
        ) = classify_date_roles(
            fragment,
            unit.text,
            event_types[0],
            rules,
        )
        entity_scope, services = entity_scope_from_section(
            fragment,
            frozen_range.company_short_name,
            rules,
        )
        if (
            event_types[0]
            == "LIMITED_COMPANY_ESTABLISHMENT"
        ):
            fragment_text = "\n".join(
                item.text for item in fragment
            )
            if establishment_is_absorption_counterparty(
                fragment_text,
                document_text,
                rules,
            ):
                entity_scope = (
                    "SUBSIDIARY_OR_OTHER_ENTITY_RISK"
                )
            elif establishment_is_issuer_predecessor(
                fragment_text,
                frozen_range.company_short_name,
                rules,
            ):
                entity_scope = "ISSUER_OR_UNRESOLVED"
        review: list[str] = []
        if timeline_summary:
            review.append(
                "仅有时间线摘要，未发现对应详细事件段落"
            )
        if entity_scope == "SUBSIDIARY_OR_OTHER_ENTITY_RISK":
            review.append("事件发生主体可能为子公司或其他主体")

        sections.append(EventSection(
            section_id=(
                f"SEC-{frozen_range.company_id}-F-"
                f"{len(sections)+1:04d}"
            ),
            company_id=frozen_range.company_id,
            source_patch_id=frozen_range.patch_id,
            source_kind=(
                "TIMELINE_ONLY"
                if timeline_summary
                else "DATED_FALLBACK"
            ),
            title=unit.text,
            title_unit_index=unit.global_index,
            heading_level=9,
            event_type_candidate=event_types[0],
            event_type_candidates=event_types,
            explicit_combined_event=combined,
            ordinal_labels=extract_ordinal_labels(
                unit.text
            ),
            timeline_summary=timeline_summary,
            units=fragment,
            date_roles=date_roles,
            event_period=event_period,
            event_date_primary_role=primary_date_role,
            event_date_selection_basis=date_basis,
            negative_disclosure=False,
            negative_reason=None,
            entity_scope_candidate=entity_scope,
            service_provider_mentions=services,
            signals=[
                (
                    "timeline_only"
                    if timeline_summary
                    else "dated_fallback"
                ),
                *[
                    f"matched_term:{term}"
                    for term in terms
                ],
            ],
            confidence=0.62 if timeline_summary else 0.68,
            review_reasons=review,
        ))
    return sections, diagnostics


def strip_legal_suffix(
    name: str,
    rules: dict[str, Any],
) -> str:
    cleaned = clean_other_entity_subject(name, rules)
    suffixes = sorted(
        [
            normalize_text(value)
            for value in rules["entity_roles"].get(
                "predecessor_legal_suffixes",
                [],
            )
        ],
        key=len,
        reverse=True,
    )
    for suffix in suffixes:
        if cleaned.endswith(suffix):
            return cleaned[:-len(suffix)]
    return cleaned


def subject_has_exact_issuer_core(
    subject: str,
    company_short_name: str,
    rules: dict[str, Any],
) -> bool:
    stem = strip_legal_suffix(subject, rules)
    return any(
        stem == core
        for core in issuer_core_tokens(
            company_short_name,
            rules,
        )
    )


def issuer_legal_entity_mentions(
    text: str,
    company_short_name: str,
    rules: dict[str, Any],
) -> set[str]:
    normalized = normalize_text(text)
    names: set[str] = set()
    suffix = r"(?:有限责任公司|有限公司|有限)"
    prefix = (
        r"(?:^|[，。；、：:]|公司前身|发行人前身|"
        r"本公司前身|前身|"
        r"(?:19|20)\d{2}年(?:\d{1,2}月(?:\d{1,2}日)?)?)"
    )
    for core in issuer_core_tokens(
        company_short_name,
        rules,
    ):
        pattern = (
            prefix
            + r"(?P<name>"
            + re.escape(core)
            + r"[\u4e00-\u9fffA-Za-z0-9（）()]{0,12}?"
            + suffix
            + r")"
        )
        for match in re.finditer(pattern, normalized):
            name = clean_other_entity_subject(
                match.group("name"),
                rules,
            )
            if name:
                names.add(name)
    return names


def issuer_establishment_context_mentions(
    text: str,
    company_short_name: str,
    rules: dict[str, Any],
) -> set[str]:
    normalized = normalize_text(text)
    names: set[str] = set()
    suffix = r"(?:有限责任公司|有限公司|有限)"
    prefix = (
        r"(?:^|[，。；、：:]|公司前身|发行人前身|"
        r"本公司前身|前身|"
        r"(?:19|20)\d{2}年(?:\d{1,2}月(?:\d{1,2}日)?)?)"
    )
    establishment_action = (
        r"(?:"
        r"成立(?!日期|时间)"
        r"|设立(?:情况|程序)?"
        r"|取得[^，。；]{0,16}?营业执照"
        r"|办理[^，。；]{0,16}?工商(?:设立|注册)?登记"
        r")"
    )
    for core in issuer_core_tokens(
        company_short_name,
        rules,
    ):
        pattern = (
            prefix
            + r"(?P<name>"
            + re.escape(core)
            + r"[\u4e00-\u9fffA-Za-z0-9（）()]{0,12}?"
            + suffix
            + r")"
            + r"[^，。；]{0,24}?"
            + establishment_action
        )
        for match in re.finditer(pattern, normalized):
            name = clean_other_entity_subject(
                match.group("name"),
                rules,
            )
            if name:
                names.add(name)
    return names


def section_contains_exact_issuer_establishment(
    section: EventSection,
    company_short_name: str,
    rules: dict[str, Any],
) -> bool:
    """Return True only for the issuer's own detailed establishment section.

    A long section about an absorbed company may mention the issuer later.
    Searching the entire section caused that unrelated section to block the
    issuer-predecessor timeline recovery. v0.11 restricts the decision to the
    title and the opening evidence units.
    """
    opening_units = section.units[:6]
    opening_text = "\n".join(
        [section.title]
        + [unit.text for unit in opening_units]
    )
    normalized_opening = normalize_text(opening_text)

    if not any(
        term in normalized_opening
        for term in (
            "设立情况",
            "公司设立",
            "成立于",
            "取得营业执照",
            "工商设立登记",
            "工商注册登记",
        )
    ):
        return False

    subjects = set(
        issuer_establishment_mentions(
            opening_text,
            company_short_name,
            rules,
        )
    )
    subjects.update(
        issuer_establishment_context_mentions(
            opening_text,
            company_short_name,
            rules,
        )
    )

    return any(
        subject_has_exact_issuer_core(
            subject,
            company_short_name,
            rules,
        )
        and not subject_is_absorption_counterparty(
            subject,
            opening_text,
            rules,
        )
        for subject in subjects
    )


def predecessor_subject_score(
    subject: str,
    company_short_name: str,
    rules: dict[str, Any],
) -> tuple[int, int, int, str]:
    """Lower tuples are better.

    Exact “issuer core + legal suffix” names outrank longer names that contain
    business descriptors, such as “谷捷金属制品有限公司”.
    """
    cleaned = clean_other_entity_subject(subject, rules)
    stem = strip_legal_suffix(cleaned, rules)
    cores = issuer_core_tokens(
        company_short_name,
        rules,
    )
    exact_core = any(stem == core for core in cores)
    shortest_extra = min(
        (
            max(len(stem) - len(core), 0)
            for core in cores
            if core in stem
        ),
        default=len(stem),
    )
    return (
        0 if exact_core else 1,
        shortest_extra,
        len(cleaned),
        cleaned,
    )


def generic_timeline_establishment_units(
    page_units: list[TextUnit],
    rules: dict[str, Any],
) -> list[TextUnit]:
    """Find timeline rows describing an unnamed issuer predecessor setup.

    Typical PDF text:
      注册资本1,000万元，由昆山谷捷出资设立

    The predecessor name may be located in a separate column or omitted from
    the extracted row. The event remains an issuer-history row because the
    frozen range is the issuer's equity-history timeline.
    """
    patterns = [
        r"注册资本[^，。；]{0,40}?(?:，|,)?由[^，。；]{1,30}?出资设立",
        r"由[^，。；]{1,30}?共同出资设立",
        r"由[^，。；]{1,30}?投资设立",
    ]
    output: list[TextUnit] = []
    for unit in page_units:
        normalized = unit.normalized_text
        if any(
            re.search(pattern, normalized)
            for pattern in patterns
        ):
            output.append(unit)
    return output


def recover_missing_issuer_establishment(
    frozen_range: FrozenRange,
    units: list[TextUnit],
    existing: list[EventSection],
    rules: dict[str, Any],
) -> list[EventSection]:
    """Recover an issuer-predecessor establishment from a timeline page."""

    detailed_issuer_establishment = [
        section
        for section in existing
        if (
            section.event_type_candidate
            == "LIMITED_COMPANY_ESTABLISHMENT"
            and section.entity_scope_candidate
            != "SUBSIDIARY_OR_OTHER_ENTITY_RISK"
            and section.source_kind
            != "PAGE_TIMELINE_RECOVERY"
            and section_contains_exact_issuer_establishment(
                section,
                frozen_range.company_short_name,
                rules,
            )
        )
    ]
    if detailed_issuer_establishment:
        return []

    represented_establishment_periods = [
        section.event_period
        for section in existing
        if (
            section.event_type_candidate
            == "LIMITED_COMPANY_ESTABLISHMENT"
            and section.entity_scope_candidate
            != "SUBSIDIARY_OR_OTHER_ENTITY_RISK"
            and section.event_period
        )
    ]

    document_text = "".join(
        unit.normalized_text for unit in units
    )
    by_page: dict[int, list[TextUnit]] = defaultdict(list)
    for unit in units:
        by_page[unit.pdf_page].append(unit)

    recovery_options: list[
        tuple[
            tuple[int, int, int, int, str],
            int,
            str,
            str,
            list[TextUnit],
            str,
        ]
    ] = []

    for pdf_page, page_units in sorted(by_page.items()):
        page_text = "".join(
            unit.normalized_text for unit in page_units
        )

        explicit_subjects = set(
            issuer_establishment_mentions(
                page_text,
                frozen_range.company_short_name,
                rules,
            )
        )
        explicit_subjects.update({
            subject
            for subject in establishment_subject_names(
                page_text,
                rules,
            )
            if subject_is_issuer_predecessor(
                subject,
                frozen_range.company_short_name,
                page_text,
                rules,
            )
        })

        explicit_subjects = {
            subject
            for subject in explicit_subjects
            if (
                "成立日期" not in page_text
                or (
                    normalize_text(subject) + "成立日期"
                    not in page_text
                )
            )
            and not subject_is_absorption_counterparty(
                subject,
                document_text,
                rules,
            )
            and subject_has_exact_issuer_core(
                subject,
                frozen_range.company_short_name,
                rules,
            )
        }

        generic_units = generic_timeline_establishment_units(
            page_units,
            rules,
        )
        if not explicit_subjects and not generic_units:
            continue

        date_unit_pairs: list[tuple[str, TextUnit]] = []
        for unit in page_units:
            for value in extract_dates(unit.text):
                key = normalized_date_key(value)
                if key:
                    date_unit_pairs.append((key, unit))
        if not date_unit_pairs:
            continue

        date_keys = sorted({
            key for key, _ in date_unit_pairs
        })
        event_period = min(date_keys)
        if any(
            period_compatible(
                represented,
                event_period,
            )
            for represented in represented_establishment_periods
        ):
            continue

        timeline_date_count = len(date_keys)
        event_label_count = sum(
            any(
                term in unit.normalized_text
                for term in (
                    "成立",
                    "设立",
                    "股权转让",
                    "增资",
                    "吸收合并",
                    "整体变更",
                )
            )
            for unit in page_units
        )

        if explicit_subjects:
            subject = min(
                explicit_subjects,
                key=lambda value: predecessor_subject_score(
                    value,
                    frozen_range.company_short_name,
                    rules,
                ),
            )
            subject_rank = predecessor_subject_score(
                subject,
                frozen_range.company_short_name,
                rules,
            )
            source_signal = "EXPLICIT_ISSUER_PREDECESSOR"
            target_units = [
                unit
                for unit in page_units
                if (
                    normalize_text(subject)
                    in unit.normalized_text
                    and (
                        "成立" in unit.normalized_text
                        or "设立" in unit.normalized_text
                    )
                )
            ]
        else:
            subject = "发行人前身"
            subject_rank = (1, 99, 99, subject)
            source_signal = "GENERIC_CAPITAL_SETUP_ROW"
            target_units = generic_units

        target_indexes = [
            page_units.index(unit)
            for unit in target_units
            if unit in page_units
        ]
        selected: list[TextUnit] = []
        if target_indexes:
            center = target_indexes[0]
            selected.extend(
                page_units[
                    max(0, center - 3):
                    min(len(page_units), center + 5)
                ]
            )
        else:
            selected.extend(page_units)

        # Add the unit carrying the selected earliest date when the PDF text
        # has split the date and description into different columns.
        earliest_date_units = [
            unit
            for key, unit in date_unit_pairs
            if key == event_period
        ]
        selected.extend(earliest_date_units)
        selected = sorted(
            {
                unit.global_index: unit
                for unit in selected
            }.values(),
            key=lambda unit: unit.global_index,
        )

        option_score = (
            subject_rank[0],
            subject_rank[1],
            -timeline_date_count,
            -event_label_count,
            subject_rank[3],
        )
        recovery_options.append((
            option_score,
            pdf_page,
            subject,
            event_period,
            selected,
            source_signal,
        ))

    if not recovery_options:
        return []

    (
        _,
        pdf_page,
        subject,
        event_period,
        selected,
        source_signal,
    ) = min(
        recovery_options,
        key=lambda item: item[0],
    )

    recovered_action = (
        "成立"
        if (
            source_signal
            == "EXPLICIT_ISSUER_PREDECESSOR"
            and any(
                "成立" in unit.normalized_text
                for unit in selected
            )
        )
        else "设立"
    )
    title = (
        f"{event_period}，{subject}{recovered_action}"
    )
    return [EventSection(
        section_id=(
            f"SEC-{frozen_range.company_id}-R-0001"
        ),
        company_id=frozen_range.company_id,
        source_patch_id=frozen_range.patch_id,
        source_kind="PAGE_TIMELINE_RECOVERY",
        title=title,
        title_unit_index=selected[0].global_index,
        heading_level=9,
        event_type_candidate=(
            "LIMITED_COMPANY_ESTABLISHMENT"
        ),
        event_type_candidates=[
            "LIMITED_COMPANY_ESTABLISHMENT"
        ],
        explicit_combined_event=False,
        ordinal_labels=[],
        timeline_summary=True,
        units=selected,
        date_roles={
            "event_period": [event_period],
        },
        event_period=event_period,
        event_date_primary_role="event_period",
        event_date_selection_basis=(
            "PAGE_TIMELINE_EARLIEST_DATE"
        ),
        negative_disclosure=False,
        negative_reason=None,
        entity_scope_candidate="ISSUER_OR_UNRESOLVED",
        service_provider_mentions=(
            service_provider_mentions(
                "".join(
                    item.normalized_text
                    for item in selected
                ),
                rules,
            )
        ),
        signals=[
            "page_timeline_recovery",
            source_signal.lower(),
            f"recovered_pdf_page:{pdf_page}",
        ],
        confidence=0.82,
        review_reasons=[],
    )]


def section_similarity(left: EventSection, right: EventSection) -> float:
    return SequenceMatcher(None, normalize_text(left.title), normalize_text(right.title)).ratio()


def period_compatible(left: str | None, right: str | None) -> bool:
    if not left or not right: return False
    return left == right or left.startswith(right) or right.startswith(left) or left[:7] == right[:7]


def semantic_equivalence_group(
    left_type: str,
    right_type: str,
    rules: dict[str, Any],
) -> dict[str, Any] | None:
    if left_type == right_type:
        return None
    for group in rules.get(
        "semantic_event_equivalence", {}
    ).get("groups", []):
        types = set(group.get("types", []))
        if left_type in types and right_type in types:
            return group
    return None


def event_types_semantically_equivalent(
    left_type: str,
    right_type: str,
    rules: dict[str, Any],
) -> bool:
    return (
        left_type == right_type
        or semantic_equivalence_group(
            left_type, right_type, rules
        ) is not None
    )


def canonical_cluster_event_type(
    sections: list[EventSection],
    primary_type: str,
    rules: dict[str, Any],
) -> str:
    cluster_types = {
        section.event_type_candidate for section in sections
    }
    for group in rules.get(
        "semantic_event_equivalence", {}
    ).get("groups", []):
        if len(cluster_types.intersection(group.get("types", []))) >= 2:
            return str(group["canonical_type"])

    canonicalization = rules.get(
        "semantic_event_equivalence", {}
    ).get("canonicalization_evidence_terms", {}).get(
        primary_type
    )
    if canonicalization:
        text = "".join(
            unit.normalized_text
            for section in sections
            for unit in section.units
        )
        if any(
            normalize_text(term) in text
            for term in canonicalization.get("terms", [])
        ):
            return str(canonicalization["canonical_type"])
    return primary_type


def section_ranges_overlap(
    left: EventSection,
    right: EventSection,
) -> bool:
    return not (
        left.units[-1].global_index < right.units[0].global_index
        or right.units[-1].global_index < left.units[0].global_index
    )


def should_merge_sections(
    left: EventSection,
    right: EventSection,
    rules: dict[str, Any],
) -> bool:
    equivalent_group = semantic_equivalence_group(
        left.event_type_candidate,
        right.event_type_candidate,
        rules,
    )
    same_type = (
        left.event_type_candidate
        == right.event_type_candidate
    )
    consequence_pair = None
    for item in rules.get("merge", {}).get(
        "consequence_attachment", []
    ):
        pair = {
            str(item.get("parent_type")),
            str(item.get("child_type")),
        }
        if {
            left.event_type_candidate,
            right.event_type_candidate,
        } == pair:
            consequence_pair = item
            break

    if (
        not same_type
        and equivalent_group is None
        and consequence_pair is None
    ):
        return False

    if (
        left.ordinal_labels
        and right.ordinal_labels
        and set(left.ordinal_labels).isdisjoint(
            right.ordinal_labels
        )
    ):
        return False

    page_gap = abs(
        left.units[0].pdf_page
        - right.units[0].pdf_page
    )
    overlap = section_ranges_overlap(left, right)

    if consequence_pair is not None:
        text = normalize_text(left.title + right.title)
        markers = [
            normalize_text(value)
            for value in consequence_pair.get("markers", [])
        ]
        return (
            page_gap
            <= int(consequence_pair.get("max_page_gap", 3))
            and any(marker in text for marker in markers)
        )

    if equivalent_group is not None:
        max_gap = int(
            equivalent_group.get("overlap_or_page_gap", 2)
        )
        return overlap or page_gap <= max_gap

    if (
        left.event_type_candidate == "ABSORPTION_MERGER"
        and page_gap
        <= int(rules["merge"][
            "absorption_merger_page_gap"
        ])
    ):
        return True

    if (
        (left.service_provider_mentions or right.service_provider_mentions)
        and page_gap <= int(
            rules["merge"].get(
                "service_provider_support_page_gap", 3
            )
        )
        and (
            "本次" in normalize_text(left.title)
            or "本次" in normalize_text(right.title)
        )
    ):
        return True

    if left.event_period and right.event_period:
        if not period_compatible(
            left.event_period,
            right.event_period,
        ):
            return False
        return page_gap <= int(
            rules["merge"]["period_match_page_gap"]
        )

    if overlap:
        return True

    if (
        page_gap
        <= int(rules["merge"]["no_period_page_gap"])
        and section_similarity(left, right)
        >= float(rules["merge"]["title_similarity"])
    ):
        return True

    if (
        (left.timeline_summary or right.timeline_summary)
        and page_gap
        <= int(rules["merge"]["timeline_support_page_gap"])
        and (
            not left.event_period
            or not right.event_period
            or period_compatible(
                left.event_period,
                right.event_period,
            )
        )
    ):
        return True

    return False


def merge_event_sections(sections: list[EventSection], rules: dict[str, Any]) -> list[list[EventSection]]:
    clusters: list[list[EventSection]] = []
    for section in sorted(sections, key=lambda item: (item.units[0].pdf_page, item.title_unit_index)):
        target = None
        for cluster in clusters:
            if any(should_merge_sections(section, existing, rules) for existing in cluster): target = cluster; break
        if target is None: clusters.append([section])
        else: target.append(section)
    return clusters


def group_source_line_ranges(units: list[TextUnit]) -> list[dict[str, int]]:
    grouped: dict[int, list[int]] = defaultdict(list)
    for unit in units: grouped[unit.pdf_page].append(unit.line_index)
    return [{"pdf_page": page, "line_start": min(indexes), "line_end": max(indexes)} for page, indexes in sorted(grouped.items())]


def merge_date_roles(sections: list[EventSection]) -> dict[str, list[str]]:
    merged: dict[str, list[str]] = defaultdict(list)
    for section in sections:
        for role, values in section.date_roles.items():
            for value in values:
                if value not in merged[role]: merged[role].append(value)
    return dict(merged)


def choose_event_period(
    sections: list[EventSection],
    date_roles: dict[str, list[str]],
    rules: dict[str, Any],
    canonical_event_type: str | None = None,
    force_role_priority: bool = False,
) -> tuple[str | None, str | None, str]:
    if force_role_priority and canonical_event_type:
        preferences = rules["event_date_selection"].get(
            canonical_event_type,
            rules["date_role_precedence"],
        )
        for role in preferences:
            values = date_roles.get(role, [])
            if values:
                return (
                    sorted(values)[0],
                    role,
                    f"CLUSTER_CANONICAL_ROLE:{role}",
                )

    ordered = sorted(
        sections,
        key=lambda item: (
            item.source_kind
            not in {"HEADING_SECTION", "DATED_FALLBACK"},
            -item.confidence,
            item.units[0].pdf_page,
        ),
    )
    for section in ordered:
        if section.event_period:
            return (
                section.event_period,
                section.event_date_primary_role,
                section.event_date_selection_basis,
            )

    for role in rules["date_role_precedence"]:
        if date_roles.get(role):
            return (
                date_roles[role][0],
                role,
                f"CLUSTER_ROLE_PRIORITY:{role}",
            )

    return None, None, "NO_RELEVANT_DATE"


def clean_title(text: str, max_chars: int = 180) -> str:
    value = re.sub(r"\s+", " ", text.strip())
    return value if len(value) <= max_chars else value[:max_chars-1] + "…"


def build_candidate_records(
    frozen_range: FrozenRange,
    sections: list[EventSection],
    resolver: PrintedPageResolver,
    external_fragment_dir: Path,
    rules: dict[str, Any],
) -> tuple[
    list[CandidateEvent],
    list[CandidateEvidence],
    list[dict[str, Any]],
]:
    clusters = merge_event_sections(sections, rules)
    external_fragment_dir.mkdir(
        parents=True,
        exist_ok=True,
    )
    candidates: list[CandidateEvent] = []
    evidences: list[CandidateEvidence] = []
    reviews: list[dict[str, Any]] = []

    for sequence, cluster in enumerate(
        clusters,
        start=1,
    ):
        candidate_id = (
            f"CE-{frozen_range.company_id}-{sequence:04d}"
        )
        cluster.sort(
            key=lambda item: (
                item.source_kind
                not in {
                    "HEADING_SECTION",
                    "DATED_FALLBACK",
                },
                -item.confidence,
                item.units[0].pdf_page,
            )
        )
        primary = cluster[0]

        for evidence_sequence, section in enumerate(
            cluster,
            start=1,
        ):
            evidence_id = (
                f"EVD-{frozen_range.company_id}-"
                f"{sequence:04d}-{evidence_sequence:02d}"
            )
            text = "\n".join(
                unit.text for unit in section.units
            ).strip()
            file_path = (
                external_fragment_dir / f"{evidence_id}.txt"
            )
            file_path.write_text(text, encoding="utf-8")
            start_page = min(
                unit.pdf_page for unit in section.units
            )
            end_page = max(
                unit.pdf_page for unit in section.units
            )
            evidences.append(CandidateEvidence(
                evidence_id=evidence_id,
                candidate_event_id=candidate_id,
                company_id=frozen_range.company_id,
                source_patch_id=frozen_range.patch_id,
                evidence_role=(
                    "PRIMARY"
                    if section is primary
                    else "SUPPORTING"
                ),
                source_kind=section.source_kind,
                pdf_page_start=start_page,
                pdf_page_end=end_page,
                printed_page_start=resolver.resolve(
                    start_page
                ),
                printed_page_end=resolver.resolve(
                    end_page
                ),
                printed_page_value_type=resolver.value_type,
                source_line_ranges=group_source_line_ranges(
                    section.units
                ),
                evidence_text=text,
                evidence_sha256=sha256_text(text),
                fragment_file=str(file_path),
                matched_signals=section.signals,
            ))

        date_roles = merge_date_roles(cluster)
        canonical_event_type = canonical_cluster_event_type(
            cluster,
            primary.event_type_candidate,
            rules,
        )
        semantic_group_merged = any(
            semantic_equivalence_group(
                left.event_type_candidate,
                right.event_type_candidate,
                rules,
            ) is not None
            for left in cluster
            for right in cluster
            if left is not right
        )
        (
            event_period,
            primary_date_role,
            date_basis,
        ) = choose_event_period(
            cluster,
            date_roles,
            rules,
            canonical_event_type=canonical_event_type,
            force_role_priority=semantic_group_merged,
        )

        event_types: list[str] = []
        signals: list[str] = []
        review_reasons: list[str] = []
        services: list[str] = []
        ordinal_labels: list[str] = []
        entity_scope = "ISSUER_OR_UNRESOLVED"

        for section in cluster:
            for value in section.event_type_candidates:
                if value not in event_types:
                    event_types.append(value)
            for value in section.signals:
                if value not in signals:
                    signals.append(value)
            for value in section.review_reasons:
                if value not in review_reasons:
                    review_reasons.append(value)
            for value in section.service_provider_mentions:
                if value not in services:
                    services.append(value)
            for value in section.ordinal_labels:
                if value not in ordinal_labels:
                    ordinal_labels.append(value)
            if (
                section.entity_scope_candidate
                == "SUBSIDIARY_OR_OTHER_ENTITY_RISK"
            ):
                entity_scope = (
                    "SUBSIDIARY_OR_OTHER_ENTITY_RISK"
                )
        if canonical_event_type not in event_types:
            event_types.insert(0, canonical_event_type)

        primary_evidence = next(
            evidence
            for evidence in evidences
            if (
                evidence.candidate_event_id == candidate_id
                and evidence.evidence_role == "PRIMARY"
            )
        )
        if (
            primary_evidence.printed_page_start is None
            or primary_evidence.printed_page_end is None
        ):
            review_reasons.append(
                "候选正文页码无法完整确认"
            )

        confidence = round(
            min(
                0.99,
                max(
                    section.confidence
                    for section in cluster
                )
                + (0.05 if len(cluster) > 1 else 0),
            ),
            3,
        )
        if confidence < float(
            rules["review"]["low_confidence_threshold"]
        ):
            review_reasons.append(
                "候选置信度低于复核阈值"
            )

        if (
            not event_period
            and primary.event_type_candidate
            not in set(
                rules["review"][
                    "undated_event_types_allowed"
                ]
            )
        ):
            review_reasons.append(
                "候选事件缺少可识别主要日期"
            )

        candidate = CandidateEvent(
            candidate_event_id=candidate_id,
            company_id=frozen_range.company_id,
            company_short_name=(
                frozen_range.company_short_name
            ),
            source_chapter_type="equity_history",
            source_patch_id=frozen_range.patch_id,
            disclosure_scope=(
                frozen_range.disclosure_scope
            ),
            event_type_candidate=canonical_event_type,
            event_type_candidates=event_types,
            ordinal_labels=ordinal_labels,
            event_period=event_period,
            event_date_text=event_period,
            event_date_primary_role=primary_date_role,
            event_date_selection_basis=date_basis,
            event_dates=date_roles,
            event_title=clean_title(primary.title),
            pdf_page_start=min(
                evidence.pdf_page_start
                for evidence in evidences
                if evidence.candidate_event_id
                == candidate_id
            ),
            pdf_page_end=max(
                evidence.pdf_page_end
                for evidence in evidences
                if evidence.candidate_event_id
                == candidate_id
            ),
            printed_page_start=(
                primary_evidence.printed_page_start
            ),
            printed_page_end=(
                primary_evidence.printed_page_end
            ),
            printed_page_value_type=resolver.value_type,
            primary_evidence_id=(
                primary_evidence.evidence_id
            ),
            supporting_evidence_ids=[
                evidence.evidence_id
                for evidence in evidences
                if (
                    evidence.candidate_event_id
                    == candidate_id
                    and evidence.evidence_role
                    == "SUPPORTING"
                )
            ],
            entity_scope_candidate=entity_scope,
            service_provider_mentions=services,
            matched_signals=signals,
            candidate_confidence=confidence,
            candidate_status="AUTO_CANDIDATE",
            review_required=bool(review_reasons),
            review_reasons=review_reasons,
        )
        candidates.append(candidate)

        if candidate.review_required:
            reviews.append({
                "company_id": candidate.company_id,
                "review_type": "CANDIDATE_EVENT_REVIEW",
                "record_id": candidate.candidate_event_id,
                "pdf_page_start": candidate.pdf_page_start,
                "pdf_page_end": candidate.pdf_page_end,
                "event_type_candidate": (
                    candidate.event_type_candidate
                ),
                "event_title": candidate.event_title,
                "auto_value": asdict(candidate),
                "reason": "；".join(
                    candidate.review_reasons
                ),
                "manual_status": "PENDING",
                "manual_decision": None,
                "confirmed_event_type": None,
                "confirmed_pdf_page_start": None,
                "confirmed_pdf_page_end": None,
                "manual_note": None,
            })

    return candidates, evidences, reviews


def shareholder_hits(unit: TextUnit, rules: dict[str, Any]) -> tuple[list[str], list[str]]:
    categories, signals = [], []
    for category, config in rules["shareholder_evidence_categories"].items():
        matched = [term for term in config["terms"] if normalize_text(term) in unit.normalized_text]
        if matched:
            categories.append(category); signals.extend(f"{category}:{term}" for term in matched)
    return categories, signals


def text_similarity(left: str, right: str) -> float:
    return SequenceMatcher(None, normalize_text(left)[:1000], normalize_text(right)[:1000]).ratio()


def build_shareholder_evidence(frozen_range: FrozenRange, units: list[TextUnit], resolver: PrintedPageResolver, rules: dict[str, Any]) -> list[ShareholderEvidence]:
    if not units: return []
    hits_by_category: dict[str, list[tuple[TextUnit, list[str]]]] = defaultdict(list)
    for unit in units:
        categories, signals = shareholder_hits(unit, rules)
        for category in categories: hits_by_category[category].append((unit, [s for s in signals if s.startswith(category + ":")]))
    by_index = {unit.global_index: unit for unit in units}
    output: list[ShareholderEvidence] = []
    max_records = int(rules["shareholder_evidence"]["max_records_per_category"])
    for category, hits in hits_by_category.items():
        clusters: list[list[tuple[TextUnit, list[str]]]] = []
        for hit in hits:
            if not clusters:
                clusters.append([hit]); continue
            last_unit = clusters[-1][-1][0]
            if hit[0].pdf_page - last_unit.pdf_page <= int(rules["shareholder_evidence"]["merge_page_gap"]) and hit[0].global_index - last_unit.global_index <= int(rules["shareholder_evidence"]["merge_unit_gap"]):
                clusters[-1].append(hit)
            else: clusters.append([hit])
        category_records: list[ShareholderEvidence] = []
        for cluster in clusters:
            start_index = max(min(by_index), cluster[0][0].global_index - int(rules["shareholder_evidence"]["context_before_units"]))
            end_index = min(max(by_index), cluster[-1][0].global_index + int(rules["shareholder_evidence"]["context_after_units"]))
            selected = [by_index[index] for index in range(start_index, end_index+1) if index in by_index]
            max_units = int(rules["shareholder_evidence"]["max_block_units"])
            selected = selected[:max_units]
            text = "\n".join(unit.text for unit in selected).strip()
            if not text: continue
            if any(text_similarity(text, existing.evidence_text) >= float(rules["shareholder_evidence"]["duplicate_similarity"]) for existing in category_records): continue
            signals = []
            for _, sigs in cluster:
                for signal in sigs:
                    if signal not in signals: signals.append(signal)
            start_page, end_page = min(u.pdf_page for u in selected), max(u.pdf_page for u in selected)
            category_records.append(ShareholderEvidence(
                shareholder_evidence_id="", company_id=frozen_range.company_id, company_short_name=frozen_range.company_short_name,
                source_patch_id=frozen_range.patch_id, evidence_category=category, pdf_page_start=start_page, pdf_page_end=end_page,
                printed_page_start=resolver.resolve(start_page), printed_page_end=resolver.resolve(end_page),
                printed_page_value_type=resolver.value_type, evidence_text=text, evidence_sha256=sha256_text(text),
                matched_signals=signals, evidence_status="AUTO_AUXILIARY_EVIDENCE",
            ))
            if len(category_records) >= max_records: break
        for record in category_records:
            record.shareholder_evidence_id = f"SHE-{frozen_range.company_id}-{len(output)+1:04d}"
            output.append(record)
    return output


def other_entity_summary_disclosures_from_sections(
    frozen_range: FrozenRange,
    sections: list[EventSection],
    rules: dict[str, Any],
    start_sequence: int,
) -> list[SummaryDisclosure]:
    if not sections:
        return []

    groups: list[list[EventSection]] = []
    for section in sorted(
        sections,
        key=lambda item: (
            item.units[0].pdf_page,
            item.title_unit_index,
        ),
    ):
        target: list[EventSection] | None = None
        for group in groups:
            if any(
                event_types_semantically_equivalent(
                    section.event_type_candidate,
                    existing.event_type_candidate,
                    rules,
                )
                and not (
                    section.units[0].pdf_page
                    > existing.units[-1].pdf_page + 1
                    or existing.units[0].pdf_page
                    > section.units[-1].pdf_page + 1
                )
                for existing in group
            ):
                target = group
                break
        if target is None:
            groups.append([section])
        else:
            target.append(section)

    output: list[SummaryDisclosure] = []
    for index, group in enumerate(groups, start=start_sequence):
        units_by_key: dict[tuple[int, int, int], TextUnit] = {}
        event_types: list[str] = []
        for section in group:
            for unit in section.units:
                units_by_key[(
                    unit.pdf_page,
                    unit.line_index,
                    unit.segment_index,
                )] = unit
            for event_type in section.event_type_candidates:
                if event_type not in event_types:
                    event_types.append(event_type)
        units = sorted(
            units_by_key.values(),
            key=lambda item: item.global_index,
        )
        text = "\n".join(unit.text for unit in units).strip()
        output.append(SummaryDisclosure(
            summary_disclosure_id=(
                f"SUM-{frozen_range.company_id}-{index:04d}"
            ),
            company_id=frozen_range.company_id,
            source_patch_id=frozen_range.patch_id,
            summary_kind="OTHER_ENTITY_EVENT",
            event_type_candidates=event_types,
            title=group[0].title,
            pdf_page_start=min(unit.pdf_page for unit in units),
            pdf_page_end=max(unit.pdf_page for unit in units),
            printed_page_start=None,
            printed_page_end=None,
            printed_page_value_type=(
                frozen_range.printed_page_value_type
            ),
            evidence_text=text,
            evidence_sha256=sha256_text(text),
            summary_reason=rules["summary_disclosures"][
                "other_entity_event_reason"
            ],
            record_status=(
                "SUMMARY_DISCLOSURE_NOT_BUSINESS_CANDIDATE"
            ),
        ))
    return output


def validate_company_outputs(equity_range: FrozenRange, shareholder_range: FrozenRange, candidates: list[CandidateEvent], evidence: list[CandidateEvidence], shareholder_evidence: list[ShareholderEvidence], negatives: list[NegativeDisclosure]) -> dict[str, Any]:
    errors, warnings = [], []
    candidate_ids = [item.candidate_event_id for item in candidates]
    evidence_ids = [item.evidence_id for item in evidence]
    if len(candidate_ids) != len(set(candidate_ids)): errors.append("候选事件ID重复")
    if len(evidence_ids) != len(set(evidence_ids)): errors.append("候选证据ID重复")
    evidence_by_id = {item.evidence_id: item for item in evidence}
    for item in candidates:
        if not (equity_range.final_start_pdf_page <= item.pdf_page_start <= item.pdf_page_end <= equity_range.final_end_pdf_page): errors.append(f"{item.candidate_event_id}超出股本历史冻结范围")
        if item.primary_evidence_id not in evidence_by_id: errors.append(f"{item.candidate_event_id}缺少主证据")
        if item.event_type_candidate in {neg.event_type_candidate for neg in negatives} and any(neg.pdf_page_start <= item.pdf_page_start <= neg.pdf_page_end for neg in negatives): errors.append(f"{item.candidate_event_id}与否定披露范围重叠")
    for item in evidence:
        if not (equity_range.final_start_pdf_page <= item.pdf_page_start <= item.pdf_page_end <= equity_range.final_end_pdf_page): errors.append(f"{item.evidence_id}超出股本历史冻结范围")
        if not item.evidence_text.strip(): errors.append(f"{item.evidence_id}证据文本为空")
    for item in shareholder_evidence:
        if not (shareholder_range.final_start_pdf_page <= item.pdf_page_start <= item.pdf_page_end <= shareholder_range.final_end_pdf_page): errors.append(f"{item.shareholder_evidence_id}超出股东冻结范围")
    if not candidates: warnings.append("冻结股本历史范围内未生成候选事件")
    if "PARTIAL" in equity_range.disclosure_scope: warnings.append("股本历史冻结范围属于部分披露")
    if "REFERENCED" in equity_range.disclosure_scope: warnings.append("股本历史主要引用其他文件")
    return {"company_id": equity_range.company_id, "validation_status": "FAILED" if errors else "PASSED", "errors": errors, "warnings": warnings,
            "candidate_event_count": len(candidates), "candidate_evidence_count": len(evidence), "shareholder_evidence_count": len(shareholder_evidence), "negative_disclosure_count": len(negatives)}


def autosize_sheet(worksheet: Any) -> None:
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max((len("" if cell.value is None else str(cell.value)) for cell in column_cells), default=0)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 45)


def write_review_workbook(
    path: Path,
    review_items: list[dict[str, Any]],
    candidates: list[CandidateEvent],
    shareholder_evidence: list[ShareholderEvidence],
    negatives: list[NegativeDisclosure],
    summary_disclosures: list[SummaryDisclosure],
    coverage_gaps: list[CoverageGap],
    company_summaries: list[dict[str, Any]],
    frozen_ranges: list[FrozenRange],
) -> None:
    workbook = Workbook()
    review_sheet = workbook.active
    review_sheet.title = "人工复核队列"
    candidate_sheet = workbook.create_sheet("候选事件")
    shareholder_sheet = workbook.create_sheet("股东辅助证据")
    negative_sheet = workbook.create_sheet("否定披露")
    disclosure_sheet = workbook.create_sheet("摘要与引用披露")
    gap_sheet = workbook.create_sheet("覆盖缺口")
    summary_sheet = workbook.create_sheet("公司摘要")
    range_sheet = workbook.create_sheet("冻结范围")

    review_sheet.append([
        "公司代码", "复核类型", "记录ID",
        "PDF起始页", "PDF结束页", "候选事件类型",
        "候选标题", "Auto值", "复核原因",
        "人工状态", "人工结论", "确认事件类型",
        "确认起始PDF页", "确认结束PDF页", "人工备注",
    ])
    for item in review_items:
        review_sheet.append([
            item["company_id"],
            item["review_type"],
            item["record_id"],
            item["pdf_page_start"],
            item["pdf_page_end"],
            item["event_type_candidate"],
            item["event_title"],
            json.dumps(
                item["auto_value"],
                ensure_ascii=False,
            ),
            item["reason"],
            item["manual_status"],
            item["manual_decision"],
            item["confirmed_event_type"],
            item["confirmed_pdf_page_start"],
            item["confirmed_pdf_page_end"],
            item["manual_note"],
        ])

    candidate_sheet.append([
        "公司代码", "候选事件ID", "事件类型候选",
        "全部类型候选", "序次标签", "主要事件日期",
        "主要日期角色", "日期选择依据", "日期角色JSON",
        "候选标题", "PDF起始页", "PDF结束页",
        "正文起始页", "正文结束页", "正文页码类型",
        "主体范围候选", "中介机构提及", "披露范围",
        "置信度", "是否复核", "复核原因", "主证据ID",
        "支持证据ID", "信号",
    ])
    for item in candidates:
        candidate_sheet.append([
            item.company_id,
            item.candidate_event_id,
            item.event_type_candidate,
            "；".join(item.event_type_candidates),
            "；".join(item.ordinal_labels),
            item.event_period,
            item.event_date_primary_role,
            item.event_date_selection_basis,
            json.dumps(item.event_dates, ensure_ascii=False),
            item.event_title,
            item.pdf_page_start,
            item.pdf_page_end,
            item.printed_page_start,
            item.printed_page_end,
            item.printed_page_value_type,
            item.entity_scope_candidate,
            "；".join(item.service_provider_mentions),
            item.disclosure_scope,
            item.candidate_confidence,
            "是" if item.review_required else "否",
            "；".join(item.review_reasons),
            item.primary_evidence_id,
            "；".join(item.supporting_evidence_ids),
            "；".join(item.matched_signals),
        ])

    shareholder_sheet.append([
        "公司代码", "证据ID", "证据类别",
        "PDF起始页", "PDF结束页", "正文起始页",
        "正文结束页", "正文页码类型", "证据原文",
        "信号", "状态",
    ])
    for item in shareholder_evidence:
        shareholder_sheet.append([
            item.company_id,
            item.shareholder_evidence_id,
            item.evidence_category,
            item.pdf_page_start,
            item.pdf_page_end,
            item.printed_page_start,
            item.printed_page_end,
            item.printed_page_value_type,
            item.evidence_text,
            "；".join(item.matched_signals),
            item.evidence_status,
        ])

    negative_sheet.append([
        "公司代码", "否定披露ID", "事件类型",
        "标题", "PDF起始页", "PDF结束页",
        "正文起始页", "正文结束页", "否定原因",
        "证据原文", "状态",
    ])
    for item in negatives:
        negative_sheet.append([
            item.company_id,
            item.negative_disclosure_id,
            item.event_type_candidate,
            item.title,
            item.pdf_page_start,
            item.pdf_page_end,
            item.printed_page_start,
            item.printed_page_end,
            item.negative_reason,
            item.evidence_text,
            item.record_status,
        ])

    disclosure_sheet.append([
        "公司代码", "摘要披露ID", "摘要类型",
        "事件类型候选", "标题", "PDF起始页",
        "PDF结束页", "正文起始页", "正文结束页",
        "正文页码类型", "摘要原因", "证据原文",
        "状态",
    ])
    for item in summary_disclosures:
        disclosure_sheet.append([
            item.company_id,
            item.summary_disclosure_id,
            item.summary_kind,
            "；".join(item.event_type_candidates),
            item.title,
            item.pdf_page_start,
            item.pdf_page_end,
            item.printed_page_start,
            item.printed_page_end,
            item.printed_page_value_type,
            item.summary_reason,
            item.evidence_text,
            item.record_status,
        ])

    gap_sheet.append([
        "公司代码", "覆盖缺口ID", "缺口类型",
        "来源摘要ID", "标题", "PDF起始页", "PDF结束页",
        "正文起始页", "正文结束页", "正文页码类型",
        "原文披露事件数JSON", "已生成候选数JSON",
        "缺失事件数JSON", "缺口原因", "证据原文", "状态",
    ])
    for item in coverage_gaps:
        gap_sheet.append([
            item.company_id,
            item.coverage_gap_id,
            item.gap_type,
            item.source_summary_disclosure_id,
            item.title,
            item.pdf_page_start,
            item.pdf_page_end,
            item.printed_page_start,
            item.printed_page_end,
            item.printed_page_value_type,
            json.dumps(
                item.disclosed_event_counts,
                ensure_ascii=False,
            ),
            json.dumps(
                item.represented_candidate_counts,
                ensure_ascii=False,
            ),
            json.dumps(
                item.missing_event_counts,
                ensure_ascii=False,
            ),
            item.gap_reason,
            item.evidence_text,
            item.record_status,
        ])

    summary_sheet.append([
        "公司代码", "公司简称", "候选事件数",
        "候选证据数", "股东辅助证据数", "否定披露数",
        "摘要与引用披露数", "覆盖缺口数", "复核项数",
        "部分披露候选数", "错误主体风险数",
        "正文页码覆盖候选数", "公司状态",
    ])
    for item in company_summaries:
        summary_sheet.append([
            item["company_id"],
            item["company_short_name"],
            item["candidate_event_count"],
            item["candidate_evidence_count"],
            item["shareholder_evidence_count"],
            item["negative_disclosure_count"],
            item["summary_disclosure_count"],
            item["coverage_gap_count"],
            item["review_item_count"],
            item["partial_disclosure_candidate_count"],
            item["wrong_entity_risk_count"],
            item["printed_page_mapped_candidate_count"],
            item["company_status"],
        ])

    range_sheet.append([
        "公司代码", "公司简称", "章节类型", "Patch ID",
        "起始PDF页", "结束PDF页", "起始正文页",
        "结束正文页", "正文页码类型", "披露范围",
        "冻结状态",
    ])
    for item in frozen_ranges:
        range_sheet.append([
            item.company_id,
            item.company_short_name,
            item.chapter_type,
            item.patch_id,
            item.final_start_pdf_page,
            item.final_end_pdf_page,
            item.final_start_printed_page_raw,
            item.final_end_printed_page_raw,
            item.printed_page_value_type,
            item.disclosure_scope,
            item.final_status,
        ])

    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )
        autosize_sheet(worksheet)

    review_sheet.column_dimensions["H"].width = 75
    review_sheet.column_dimensions["I"].width = 55
    candidate_sheet.column_dimensions["J"].width = 55
    candidate_sheet.column_dimensions["U"].width = 50
    shareholder_sheet.column_dimensions["I"].width = 80
    negative_sheet.column_dimensions["J"].width = 80
    disclosure_sheet.column_dimensions["L"].width = 80
    gap_sheet.column_dimensions["O"].width = 80
    range_sheet.column_dimensions["J"].width = 42

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def ensure_run_dirs(repo_root: Path, run_id: str) -> dict[str, Path]:
    dirs={"auto":repo_root/"auto_output"/"candidate_events"/"runs"/run_id,"validation":repo_root/"validation"/"candidate_events"/"runs"/run_id,"logs":repo_root/"logs"/"candidate_events"/"runs"/run_id,"logs_root":repo_root/"logs"/"candidate_events"}
    for path in dirs.values(): path.mkdir(parents=True,exist_ok=True)
    return dirs


def resolve_repo_path(repo_root: Path, path: Path) -> Path:
    path=path.expanduser(); return path.resolve() if path.is_absolute() else (repo_root/path).resolve()


def run_candidate_event_generation(
    input_dir: Path,
    repo_root: Path,
    workspace_dir: Path,
    chapter_patch_file: Path,
    pagination_patch_file: Path,
    rules_file: Path,
    page_text_run_id: str,
    expected_count: int | None = None,
) -> int:
    run_id = make_run_id()
    started_at = now_iso()
    input_dir = input_dir.expanduser().resolve()
    repo_root = repo_root.expanduser().resolve()
    workspace_dir = workspace_dir.expanduser().resolve()

    chapter_patch_file = resolve_repo_path(
        repo_root,
        chapter_patch_file,
    )
    pagination_patch_file = resolve_repo_path(
        repo_root,
        pagination_patch_file,
    )
    rules_file = resolve_repo_path(
        repo_root,
        rules_file,
    )

    for path, label, is_dir in [
        (input_dir, "PDF输入目录", True),
        (repo_root, "仓库目录", True),
        (chapter_patch_file, "章节Patch", False),
        (pagination_patch_file, "页码Patch", False),
        (rules_file, "候选事件规则", False),
    ]:
        if not (
            path.is_dir() if is_dir else path.is_file()
        ):
            print(
                f"[ERROR] {label}不存在：{path}",
                file=sys.stderr,
            )
            return 2

    rules = load_json(rules_file)
    validate_rule_config(rules)
    frozen_ranges = load_frozen_ranges(
        chapter_patch_file
    )
    pagination_patches = {
        str(item["company_id"]): item
        for item in read_jsonl(
            pagination_patch_file
        )
    }

    range_by_company: dict[
        str,
        dict[str, FrozenRange],
    ] = defaultdict(dict)
    for item in frozen_ranges:
        range_by_company[item.company_id][
            item.chapter_type
        ] = item

    pdf_company_map: dict[str, str] = {}
    for pdf_path in sorted(input_dir.glob("*.pdf")):
        company_id, short_name = parse_filename(
            pdf_path.name
        )
        if company_id and short_name:
            pdf_company_map[company_id] = short_name

    company_ids = sorted(range_by_company)
    if (
        expected_count is not None
        and len(company_ids) != expected_count
    ):
        print(
            f"[ERROR] 冻结Patch公司数为{len(company_ids)}，"
            f"预期为{expected_count}",
            file=sys.stderr,
        )
        return 2

    dirs = ensure_run_dirs(repo_root, run_id)
    external_run_dir = workspace_dir / run_id
    external_run_dir.mkdir(
        parents=True,
        exist_ok=True,
    )

    all_candidates: list[CandidateEvent] = []
    all_evidence: list[CandidateEvidence] = []
    all_shareholder: list[ShareholderEvidence] = []
    all_negatives: list[NegativeDisclosure] = []
    all_summary_disclosures: list[
        SummaryDisclosure
    ] = []
    all_coverage_gaps: list[CoverageGap] = []
    all_reviews: list[dict[str, Any]] = []
    validations: list[dict[str, Any]] = []
    summaries: list[dict[str, Any]] = []
    logs: list[dict[str, Any]] = []
    fragment_manifest: list[dict[str, Any]] = []
    failures = 0
    diagnostics_total: Counter[str] = Counter()

    for index, company_id in enumerate(
        company_ids,
        start=1,
    ):
        ranges = range_by_company[company_id]
        equity_range = ranges.get("equity_history")
        shareholder_range = ranges.get("shareholders")
        print(
            f"[{index}/{len(company_ids)}] "
            f"v0.11候选事件：{company_id}"
        )

        if (
            equity_range is None
            or shareholder_range is None
        ):
            failures += 1
            validations.append({
                "company_id": company_id,
                "validation_status": "FAILED",
                "errors": ["缺少冻结范围"],
                "warnings": [],
            })
            continue

        try:
            page_path = (
                workspace_dir
                / page_text_run_id
                / company_id
                / "page_text.jsonl"
            )
            if not page_path.is_file():
                raise FileNotFoundError(
                    f"逐页文本不存在：{page_path}"
                )

            pages = load_page_text(page_path)
            pagination_patch = pagination_patches.get(
                company_id
            )
            equity_resolver = PrintedPageResolver(
                equity_range,
                pagination_patch,
            )
            shareholder_resolver = PrintedPageResolver(
                shareholder_range,
                pagination_patch,
            )

            equity_units = build_text_units(
                company_id,
                pages,
                equity_range.final_start_pdf_page,
                equity_range.final_end_pdf_page,
            )
            shareholder_units = build_text_units(
                company_id,
                pages,
                shareholder_range.final_start_pdf_page,
                shareholder_range.final_end_pdf_page,
            )

            summary_disclosures = (
                collect_summary_disclosures(
                    equity_range,
                    equity_units,
                    rules,
                )
            )
            heading_sections, negatives = (
                build_heading_sections(
                    equity_range,
                    equity_units,
                    rules,
                )
            )
            (
                fallback_sections,
                company_diagnostics,
            ) = build_fallback_sections(
                equity_range,
                equity_units,
                heading_sections,
                rules,
            )
            recovery_sections = (
                recover_missing_issuer_establishment(
                    equity_range,
                    equity_units,
                    heading_sections + fallback_sections,
                    rules,
                )
            )
            diagnostics_total.update(
                company_diagnostics
            )

            all_event_sections = (
                heading_sections
                + fallback_sections
                + recovery_sections
            )
            other_entity_sections = [
                section
                for section in all_event_sections
                if section.entity_scope_candidate
                == "SUBSIDIARY_OR_OTHER_ENTITY_RISK"
            ]
            candidate_sections = [
                section
                for section in all_event_sections
                if section.entity_scope_candidate
                != "SUBSIDIARY_OR_OTHER_ENTITY_RISK"
            ]
            if other_entity_sections:
                summary_disclosures.extend(
                    other_entity_summary_disclosures_from_sections(
                        equity_range,
                        other_entity_sections,
                        rules,
                        len(summary_disclosures) + 1,
                    )
                )
                company_diagnostics[
                    "other_entity_suppressed_count"
                ] = len(other_entity_sections)
                diagnostics_total.update({
                    "other_entity_suppressed_count":
                    len(other_entity_sections)
                })

            candidates, evidence, reviews = (
                build_candidate_records(
                    equity_range,
                    candidate_sections,
                    equity_resolver,
                    (
                        external_run_dir
                        / company_id
                        / "candidate_fragments"
                    ),
                    rules,
                )
            )
            shareholder = build_shareholder_evidence(
                shareholder_range,
                shareholder_units,
                shareholder_resolver,
                rules,
            )

            for negative in negatives:
                negative.printed_page_start = (
                    equity_resolver.resolve(
                        negative.pdf_page_start
                    )
                )
                negative.printed_page_end = (
                    equity_resolver.resolve(
                        negative.pdf_page_end
                    )
                )
                negative.printed_page_value_type = (
                    equity_resolver.value_type
                )

            for disclosure in summary_disclosures:
                disclosure.printed_page_start = (
                    equity_resolver.resolve(
                        disclosure.pdf_page_start
                    )
                )
                disclosure.printed_page_end = (
                    equity_resolver.resolve(
                        disclosure.pdf_page_end
                    )
                )
                disclosure.printed_page_value_type = (
                    equity_resolver.value_type
                )

            coverage_gaps = build_coverage_gaps(
                equity_range,
                summary_disclosures,
                candidates,
            )
            for gap in coverage_gaps:
                reviews.append({
                    "company_id": company_id,
                    "review_type": "DISCLOSURE_COVERAGE_GAP",
                    "record_id": gap.coverage_gap_id,
                    "pdf_page_start": gap.pdf_page_start,
                    "pdf_page_end": gap.pdf_page_end,
                    "event_type_candidate": "；".join(
                        gap.missing_event_counts
                    ),
                    "event_title": gap.title,
                    "auto_value": asdict(gap),
                    "reason": (
                        gap.gap_reason
                        + " 缺失计数："
                        + json.dumps(
                            gap.missing_event_counts,
                            ensure_ascii=False,
                        )
                    ),
                    "manual_status": "PENDING",
                    "manual_decision": (
                        "ACCEPT_DISCLOSURE_LIMITATION"
                    ),
                    "confirmed_event_type": None,
                    "confirmed_pdf_page_start": None,
                    "confirmed_pdf_page_end": None,
                    "manual_note": None,
                })

            if not candidates:
                reviews.append({
                    "company_id": company_id,
                    "review_type": "NO_CANDIDATE_EVENTS",
                    "record_id": (
                        f"NO-CANDIDATE-{company_id}"
                    ),
                    "pdf_page_start": (
                        equity_range
                        .final_start_pdf_page
                    ),
                    "pdf_page_end": (
                        equity_range
                        .final_end_pdf_page
                    ),
                    "event_type_candidate": None,
                    "event_title": None,
                    "auto_value": {
                        "source_patch_id": (
                            equity_range.patch_id
                        ),
                        "disclosure_scope": (
                            equity_range
                            .disclosure_scope
                        ),
                    },
                    "reason": (
                        "冻结股本历史范围内"
                        "未生成候选事件"
                    ),
                    "manual_status": "PENDING",
                    "manual_decision": None,
                    "confirmed_event_type": None,
                    "confirmed_pdf_page_start": None,
                    "confirmed_pdf_page_end": None,
                    "manual_note": None,
                })

            validation = validate_company_outputs(
                equity_range,
                shareholder_range,
                candidates,
                evidence,
                shareholder,
                negatives,
            )
            if coverage_gaps:
                validation["warnings"].append(
                    f"存在{len(coverage_gaps)}条未逐项展开的历史事件覆盖缺口"
                )
            validations.append(validation)

            for item in evidence:
                fragment_manifest.append({
                    "company_id": company_id,
                    "candidate_event_id": (
                        item.candidate_event_id
                    ),
                    "evidence_id": item.evidence_id,
                    "fragment_file": item.fragment_file,
                    "evidence_sha256": (
                        item.evidence_sha256
                    ),
                    "source_page_text_run_id": (
                        page_text_run_id
                    ),
                    "source_patch_id": (
                        item.source_patch_id
                    ),
                })

            summary = {
                "company_id": company_id,
                "company_short_name": (
                    pdf_company_map.get(company_id)
                    or equity_range.company_short_name
                ),
                "candidate_event_count": len(
                    candidates
                ),
                "candidate_evidence_count": len(
                    evidence
                ),
                "shareholder_evidence_count": len(
                    shareholder
                ),
                "negative_disclosure_count": len(
                    negatives
                ),
                "summary_disclosure_count": len(
                    summary_disclosures
                ),
                "coverage_gap_count": len(
                    coverage_gaps
                ),
                "review_item_count": len(reviews),
                "partial_disclosure_candidate_count": sum(
                    "PARTIAL"
                    in item.disclosure_scope
                    or "REFERENCED"
                    in item.disclosure_scope
                    for item in candidates
                ),
                "wrong_entity_risk_count": sum(
                    item.entity_scope_candidate
                    == "SUBSIDIARY_OR_OTHER_ENTITY_RISK"
                    for item in candidates
                ),
                "printed_page_mapped_candidate_count": sum(
                    item.printed_page_start
                    is not None
                    and item.printed_page_end
                    is not None
                    for item in candidates
                ),
                "company_status": (
                    "FAILED"
                    if validation["errors"]
                    else (
                        "REVIEW_REQUIRED"
                        if reviews
                        else "READY"
                    )
                ),
            }

            all_candidates.extend(candidates)
            all_evidence.extend(evidence)
            all_shareholder.extend(shareholder)
            all_negatives.extend(negatives)
            all_summary_disclosures.extend(
                summary_disclosures
            )
            all_coverage_gaps.extend(
                coverage_gaps
            )
            all_reviews.extend(reviews)
            summaries.append(summary)
            logs.append({
                "timestamp": now_iso(),
                "level": "INFO",
                "run_id": run_id,
                "company_id": company_id,
                "stage": PIPELINE_VERSION,
                "event": "company_completed",
                "message": (
                    f"候选事件{len(candidates)}，"
                    f"复核项{len(reviews)}，"
                    f"否定披露{len(negatives)}，"
                    f"摘要披露{len(summary_disclosures)}，"
                    f"覆盖缺口{len(coverage_gaps)}"
                ),
                "details": {
                    **summary,
                    "diagnostics": (
                        company_diagnostics
                    ),
                },
            })
        except Exception as exc:
            failures += 1
            validations.append({
                "company_id": company_id,
                "validation_status": "FAILED",
                "errors": [
                    f"{type(exc).__name__}: {exc}"
                ],
                "warnings": [],
            })
            summaries.append({
                "company_id": company_id,
                "company_short_name": (
                    pdf_company_map.get(company_id)
                    or equity_range.company_short_name
                ),
                "candidate_event_count": 0,
                "candidate_evidence_count": 0,
                "shareholder_evidence_count": 0,
                "negative_disclosure_count": 0,
                "summary_disclosure_count": 0,
                "coverage_gap_count": 0,
                "review_item_count": 1,
                "partial_disclosure_candidate_count": 0,
                "wrong_entity_risk_count": 0,
                "printed_page_mapped_candidate_count": 0,
                "company_status": "FAILED",
            })
            all_reviews.append({
                "company_id": company_id,
                "review_type": (
                    "COMPANY_PROCESSING_FAILED"
                ),
                "record_id": f"FAILED-{company_id}",
                "pdf_page_start": None,
                "pdf_page_end": None,
                "event_type_candidate": None,
                "event_title": None,
                "auto_value": None,
                "reason": (
                    f"{type(exc).__name__}: {exc}"
                ),
                "manual_status": "PENDING",
                "manual_decision": None,
                "confirmed_event_type": None,
                "confirmed_pdf_page_start": None,
                "confirmed_pdf_page_end": None,
                "manual_note": None,
            })
            logs.append({
                "timestamp": now_iso(),
                "level": "ERROR",
                "run_id": run_id,
                "company_id": company_id,
                "stage": PIPELINE_VERSION,
                "event": "company_failed",
                "message": str(exc),
                "details": {
                    "traceback": traceback.format_exc(),
                },
            })

    validation_errors = sum(
        len(item.get("errors", []))
        for item in validations
    )
    batch_status = (
        "PARTIAL_FAILURE"
        if failures or validation_errors
        else (
            "READY_WITH_REVIEW"
            if all_reviews
            else "READY"
        )
    )

    mapped = sum(
        item.printed_page_start is not None
        and item.printed_page_end is not None
        for item in all_candidates
    )
    metrics = {
        "metrics_version": "0.11",
        "run_id": run_id,
        "batch_status": batch_status,
        "expected_company_count": expected_count,
        "frozen_company_count": len(company_ids),
        "company_success_count": (
            len(company_ids) - failures
        ),
        "company_failure_count": failures,
        "candidate_event_count": len(all_candidates),
        "candidate_evidence_count": len(all_evidence),
        "shareholder_evidence_count": len(
            all_shareholder
        ),
        "negative_disclosure_count": len(
            all_negatives
        ),
        "summary_disclosure_count": len(
            all_summary_disclosures
        ),
        "summary_disclosure_type_counts": dict(
            Counter(
                item.summary_kind
                for item in all_summary_disclosures
            )
        ),
        "coverage_gap_count": len(all_coverage_gaps),
        "coverage_gap_type_counts": dict(
            Counter(
                item.gap_type
                for item in all_coverage_gaps
            )
        ),
        "review_queue_count": len(all_reviews),
        "candidate_printed_page_mapped_count": mapped,
        "candidate_printed_page_coverage_rate": round(
            mapped / max(len(all_candidates), 1),
            4,
        ),
        "candidate_event_type_counts": dict(
            Counter(
                item.event_type_candidate
                for item in all_candidates
            )
        ),
        "partial_or_referenced_candidate_count": sum(
            "PARTIAL" in item.disclosure_scope
            or "REFERENCED" in item.disclosure_scope
            for item in all_candidates
        ),
        "wrong_entity_risk_count": sum(
            item.entity_scope_candidate
            == "SUBSIDIARY_OR_OTHER_ENTITY_RISK"
            for item in all_candidates
        ),
        **dict(diagnostics_total),
        "validation_error_count": validation_errors,
        "chapter_patch_sha256": sha256_file(
            chapter_patch_file
        ),
        "pagination_patch_sha256": sha256_file(
            pagination_patch_file
        ),
        "rules_sha256": sha256_file(rules_file),
        "source_page_text_run_id": page_text_run_id,
        "pdf_access_mode": (
            "FILENAME_DISCOVERY_ONLY_NO_FULL_PDF_SCAN"
        ),
        "llm_called": False,
        "pevc_classification_performed": False,
        "final_event_generated": False,
        "note": (
            "v0.11以标题及开头证据判断详细设立章节，"
            "并支持从“注册资本……由……出资设立”的跨列时间线恢复发行人前身设立。"
        ),
    }

    write_jsonl(
        dirs["auto"] / "candidate_events_auto.jsonl",
        [asdict(item) for item in all_candidates],
    )
    write_jsonl(
        dirs["auto"] / "candidate_evidence_auto.jsonl",
        [asdict(item) for item in all_evidence],
    )
    write_jsonl(
        dirs["auto"] / "shareholder_evidence_auto.jsonl",
        [asdict(item) for item in all_shareholder],
    )
    write_jsonl(
        dirs["auto"] / "negative_disclosures_auto.jsonl",
        [asdict(item) for item in all_negatives],
    )
    write_jsonl(
        dirs["auto"] / "summary_disclosures_auto.jsonl",
        [
            asdict(item)
            for item in all_summary_disclosures
        ],
    )
    write_jsonl(
        dirs["auto"] / "coverage_gaps_auto.jsonl",
        [asdict(item) for item in all_coverage_gaps],
    )
    write_json(
        (
            dirs["auto"]
            / "candidate_fragments_manifest.json"
        ),
        {
            "manifest_version": "0.11",
            "run_id": run_id,
            "generated_at": now_iso(),
            "source_page_text_run_id": (
                page_text_run_id
            ),
            "fragments": fragment_manifest,
        },
    )

    write_json(
        (
            dirs["validation"]
            / "candidate_event_metrics.json"
        ),
        metrics,
    )
    write_json(
        (
            dirs["validation"]
            / "candidate_event_validation.json"
        ),
        {
            "validation_version": "0.11",
            "run_id": run_id,
            "companies": validations,
        },
    )
    write_review_workbook(
        (
            dirs["validation"]
            / "candidate_event_review_queue.xlsx"
        ),
        all_reviews,
        all_candidates,
        all_shareholder,
        all_negatives,
        all_summary_disclosures,
        all_coverage_gaps,
        summaries,
        frozen_ranges,
    )

    write_jsonl(
        (
            dirs["logs"]
            / "candidate_generation.jsonl"
        ),
        logs,
    )
    write_json(
        dirs["logs"] / "output_counts.json",
        {
            "candidate_events": len(all_candidates),
            "candidate_evidence": len(all_evidence),
            "shareholder_evidence": len(
                all_shareholder
            ),
            "negative_disclosures": len(
                all_negatives
            ),
            "summary_disclosures": len(
                all_summary_disclosures
            ),
            "coverage_gaps": len(
                all_coverage_gaps
            ),
            "review_items": len(all_reviews),
        },
    )
    write_json(
        dirs["logs"] / "run_manifest.json",
        {
            "run_id": run_id,
            "pipeline_version": PIPELINE_VERSION,
            "rules_version": rules["rules_version"],
            "started_at": started_at,
            "completed_at": now_iso(),
            "batch_status": batch_status,
            "source_page_text_run_id": (
                page_text_run_id
            ),
            "chapter_patch_file": str(
                chapter_patch_file
            ),
            "pagination_patch_file": str(
                pagination_patch_file
            ),
            "rules_file": str(rules_file),
            "external_fragment_dir": str(
                external_run_dir
            ),
            "auto_freeze_policy": (
                "run_specific_immutable_directory"
            ),
            "llm_called": False,
        },
    )
    write_json(
        dirs["logs_root"] / "latest_run.json",
        {
            "run_id": run_id,
            "pipeline_version": PIPELINE_VERSION,
            "batch_status": batch_status,
            "completed_at": now_iso(),
        },
    )

    print()
    print("候选事件精准截取 v0.11 完成")
    print(f"运行ID：{run_id}")
    print(f"冻结公司：{len(company_ids)}")
    print(f"公司失败：{failures}")
    print(f"候选事件：{len(all_candidates)}")
    print(f"候选证据：{len(all_evidence)}")
    print(
        "候选正文页码覆盖率："
        f"{metrics['candidate_printed_page_coverage_rate']:.2%}"
    )
    print(f"否定披露：{len(all_negatives)}")
    print(
        f"摘要与引用披露："
        f"{len(all_summary_disclosures)}"
    )
    print(f"覆盖缺口：{len(all_coverage_gaps)}")
    print(
        f"股东辅助证据：{len(all_shareholder)}"
    )
    print(f"人工复核项：{len(all_reviews)}")
    print(f"批次状态：{batch_status}")
    return 1 if failures or validation_errors else 0
