from __future__ import annotations

import hashlib
import json
import re
import shutil
import sys
import traceback
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


PIPELINE_VERSION = "chapter_location_v0.5"

DASH_TRANSLATION = str.maketrans({
    "—": "-", "–": "-", "－": "-", "﹣": "-"
})

NUMBERING_PATTERNS: list[tuple[int, re.Pattern[str]]] = [
    (0, re.compile(r"^第[一二三四五六七八九十百\d]+节")),
    (1, re.compile(r"^[一二三四五六七八九十百]+、")),
    (2, re.compile(r"^[（(][一二三四五六七八九十百]+[）)]")),
    (3, re.compile(r"^\d+[\.、]")),
    (4, re.compile(r"^[（(]\d+[）)]")),
]
NUMBERING_ONLY_PATTERNS = [
    re.compile(r"^第[一二三四五六七八九十百\d]+节$"),
    re.compile(r"^[一二三四五六七八九十百]+、$"),
    re.compile(r"^[（(][一二三四五六七八九十百]+[）)]$"),
    re.compile(r"^\d+[\.、]$"),
    re.compile(r"^[（(]\d+[）)]$"),
]

SENTENCE_END_PUNCTUATION = ("。", "；", "！", "？")
CROSS_REFERENCE_WORDS = ("详见", "参见", "请见", "见本招股说明书")
FORBIDDEN_CONFIG_KEY_PARTS = (
    "target_pages", "correct_candidate", "expected_event",
    "known_investor", "answer_mapping", "fixed_page"
)


@dataclass
class PageRecord:
    company_id: str
    short_name: str
    file_name: str
    pdf_page: int
    text_status: str
    text_char_count: int
    page_text_sha256: str
    nonempty_line_count: int
    first_lines: list[str]
    last_lines: list[str]
    extraction_warning: str | None


@dataclass
class PaginationRecord:
    company_id: str
    pdf_page: int
    printed_page_raw: str | None
    printed_page_prefix: str | None
    printed_page_number: int | None
    source_position: str | None
    detection_status: str
    mapping_status: str
    confidence: float
    candidates: list[dict[str, Any]]
    evidence_line: str | None
    review_reason: str | None


@dataclass
class HeadingRecord:
    heading_candidate_id: str
    company_id: str
    pdf_page: int
    line_start_index: int
    line_end_index: int
    line_text: str
    normalized_text: str
    title_core: str
    heading_level: int | None
    heading_class: str
    is_page_top: bool
    repeated_line_count: int
    parent_heading_id: str | None = None
    parent_heading_text: str | None = None
    parent_heading_level: int | None = None
    section_end_pdf_page: int | None = None
    matched_chapter_types: list[str] = field(default_factory=list)
    match_sources: dict[str, list[str]] = field(default_factory=dict)
    matched_aliases: list[str] = field(default_factory=list)
    matched_trigger_groups: list[list[str]] = field(default_factory=list)
    excluded_types: list[str] = field(default_factory=list)
    confidence: float = 0.0
    signals: list[str] = field(default_factory=list)


@dataclass
class ChapterCandidate:
    chapter_candidate_id: str
    company_id: str
    chapter_type: str
    rank_within_type: int
    is_primary: bool
    candidate_source: str
    heading_candidate_id: str | None
    heading_text: str
    heading_level: int | None
    parent_heading_id: str | None
    parent_heading_text: str | None
    supporting_heading_ids: list[str]
    supporting_heading_texts: list[str]
    start_pdf_page: int
    start_printed_page_raw: str | None
    start_printed_page_number: int | None
    end_pdf_page_candidate: int
    end_printed_page_raw_candidate: str | None
    end_printed_page_number_candidate: int | None
    end_inference_status: str
    toc_support_count: int
    context_support_count: int
    match_basis: list[str]
    evidence_excerpt: str
    confidence: float
    confidence_level: str
    signals: list[str]
    selection_status: str
    review_reasons: list[str]



@dataclass
class IssuerMasterSection:
    company_id: str
    status: str
    start_pdf_page: int | None
    end_pdf_page: int | None
    heading_text: str | None
    evidence_line: str | None
    context_support_count: int
    confidence: float
    review_reason: str | None

def now_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def make_run_id() -> str:
    return datetime.now().astimezone().strftime("CHAPTERLOC_V05_%Y%m%d_%H%M%S")


def normalize_line(line: str) -> str:
    return re.sub(r"\s+", "", line.strip()).translate(DASH_TRANSLATION)


def sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8", errors="replace")).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while chunk := handle.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8-sig"))


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )


def write_jsonl(path: Path, rows: Iterable[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8-sig") as handle:
        for line in handle:
            if line.strip():
                rows.append(json.loads(line))
    return rows


def validate_rule_config(payload: Any, location: str = "$") -> None:
    if isinstance(payload, dict):
        for key, value in payload.items():
            lowered = key.lower()
            if any(part in lowered for part in FORBIDDEN_CONFIG_KEY_PARTS):
                raise ValueError(
                    f"配置包含疑似答案映射或固定页码字段：{location}.{key}"
                )
            validate_rule_config(value, f"{location}.{key}")
    elif isinstance(payload, list):
        for index, value in enumerate(payload):
            validate_rule_config(value, f"{location}[{index}]")


def parse_filename(file_name: str) -> tuple[str | None, str | None]:
    match = re.match(
        r"^(?P<code>\d{6})_(?P<name>.+?)_IPO招股说明书\.pdf$",
        file_name,
        flags=re.IGNORECASE
    )
    if not match:
        return None, None
    return match.group("code"), match.group("name").strip()


def strip_numbering(normalized: str) -> str:
    value = normalized
    for _, pattern in NUMBERING_PATTERNS:
        value = pattern.sub("", value, count=1)
    return value.strip("：:")


def clean_toc_suffix(normalized: str) -> str:
    value = re.sub(r"[\.·…]{2,}\d{1,4}$", "", normalized)
    value = re.sub(r"\d{1,4}$", "", value)
    return value.strip("：:")


def detect_heading_level(normalized: str) -> int | None:
    for level, pattern in NUMBERING_PATTERNS:
        if pattern.search(normalized):
            return level
    return None


def is_numbering_only(normalized: str) -> bool:
    return any(pattern.fullmatch(normalized) for pattern in NUMBERING_ONLY_PATTERNS)


def is_title_like(title_core: str, max_chars: int) -> bool:
    if not 3 <= len(title_core) <= max_chars:
        return False
    if any(mark in title_core for mark in SENTENCE_END_PUNCTUATION):
        return False
    if title_core.count("，") >= 3:
        return False
    if title_core.startswith(("注：", "注:", "说明：", "说明:")):
        return False
    return True


def page_is_toc(lines: list[str], rules: dict[str, Any]) -> tuple[bool, list[str]]:
    signals: list[str] = []
    normalized = [normalize_line(line) for line in lines]
    head = "".join(normalized[:20])
    if any(keyword in head for keyword in rules["toc"]["title_keywords"]):
        signals.append("toc_title_keyword")
    dot_pattern = re.compile(rules["toc"]["dot_leader_pattern"])
    toc_like = sum(
        bool(dot_pattern.search(line))
        and bool(re.search(r"\d{1,4}$", normalize_line(line)))
        for line in lines
    )
    if toc_like >= int(rules["toc"]["minimum_toc_like_lines"]):
        signals.append("toc_dense_dot_leaders")
    return bool(signals), signals


def extract_page_text(
    reader: PdfReader,
    company_id: str,
    short_name: str,
    file_name: str,
    page_index: int
) -> tuple[PageRecord, str]:
    warning = None
    try:
        text = reader.pages[page_index].extract_text() or ""
        status = "AVAILABLE" if normalize_line(text) else "EMPTY_OR_NO_TEXT"
    except Exception as exc:
        text = ""
        status = "EXTRACTION_FAILED"
        warning = f"{type(exc).__name__}: {exc}"
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    record = PageRecord(
        company_id=company_id,
        short_name=short_name,
        file_name=file_name,
        pdf_page=page_index + 1,
        text_status=status,
        text_char_count=len(normalize_line(text)),
        page_text_sha256=sha256_text(text),
        nonempty_line_count=len(lines),
        first_lines=lines[:12],
        last_lines=lines[-12:],
        extraction_warning=warning
    )
    return record, text


def load_or_extract_pages(
    pdf_path: Path,
    company_id: str,
    short_name: str,
    reuse_path: Path | None
) -> tuple[list[PageRecord], list[dict[str, Any]], str]:
    if reuse_path and reuse_path.is_file():
        rows = read_jsonl(reuse_path)
        page_records: list[PageRecord] = []
        normalized_rows: list[dict[str, Any]] = []
        for row in rows:
            text = str(row.get("text") or "")
            lines = [line.strip() for line in text.splitlines() if line.strip()]
            pdf_page = int(row["pdf_page"])
            page_records.append(PageRecord(
                company_id=company_id,
                short_name=short_name,
                file_name=pdf_path.name,
                pdf_page=pdf_page,
                text_status=str(row.get("text_status") or (
                    "AVAILABLE" if normalize_line(text) else "EMPTY_OR_NO_TEXT"
                )),
                text_char_count=int(row.get("text_char_count") or len(normalize_line(text))),
                page_text_sha256=str(row.get("page_text_sha256") or sha256_text(text)),
                nonempty_line_count=len(lines),
                first_lines=lines[:12],
                last_lines=lines[-12:],
                extraction_warning=None
            ))
            normalized_rows.append({
                "company_id": company_id,
                "pdf_page": pdf_page,
                "text": text,
                "text_status": page_records[-1].text_status,
                "text_char_count": page_records[-1].text_char_count,
                "page_text_sha256": page_records[-1].page_text_sha256
            })
        return page_records, normalized_rows, "REUSED_PAGE_TEXT"

    reader = PdfReader(pdf_path, strict=False)
    page_records = []
    rows = []
    for page_index in range(len(reader.pages)):
        page_record, text = extract_page_text(
            reader, company_id, short_name, pdf_path.name, page_index
        )
        page_records.append(page_record)
        rows.append({
            "company_id": company_id,
            "pdf_page": page_record.pdf_page,
            "text": text,
            "text_status": page_record.text_status,
            "text_char_count": page_record.text_char_count,
            "page_text_sha256": page_record.page_text_sha256
        })
    return page_records, rows, "PDF_EXTRACTED"


def compile_pagination_patterns(max_digits: int) -> dict[str, re.Pattern[str]]:
    return {
        "exact_prefixed": re.compile(
            rf"^(?P<prefix>\d+(?:-\d+)+)-(?P<number>\d{{1,{max_digits}}})$"
        ),
        "exact_chinese": re.compile(
            rf"^第(?P<number>\d{{1,{max_digits}}})页$"
        ),
        "exact_bare": re.compile(
            rf"^(?P<number>\d{{1,{max_digits}}})$"
        ),
        "header_prefixed": re.compile(
            rf"^(?P<raw>(?P<prefix>\d+(?:-\d+)+)-(?P<number>\d{{1,{max_digits}}}))(?=\D|$)"
        ),
        "header_bare": re.compile(
            rf"^(?P<number>\d{{1,{max_digits}}})(?=\D|$)"
        ),
    }


def pagination_candidates_for_page(
    lines: list[str],
    rules: dict[str, Any],
    page_count: int
) -> list[dict[str, Any]]:
    config = rules["pagination"]
    max_edge = int(config["max_edge_lines"])
    header_scan_lines = int(config.get("prospectus_header_scan_lines", 4))
    bare_footer_lines = int(config["bare_number_footer_lines"])
    max_digits = int(config["bare_number_max_digits"])
    year_min = int(config["exclude_year_min"])
    year_max = int(config["exclude_year_max"])
    max_page_number = page_count + int(config["max_page_number_margin"])
    prospectus_keywords = [
        normalize_line(item)
        for item in config.get("prospectus_header_keywords", ["招股说明书"])
    ]
    patterns = compile_pagination_patterns(max_digits)

    indexed: list[tuple[int, str, str, int]] = []
    for index, line in enumerate(lines[:max_edge]):
        indexed.append((index, line, "HEADER", index))
    footer_start = max(len(lines) - max_edge, 0)
    for index in range(footer_start, len(lines)):
        indexed.append(
            (index, lines[index], "FOOTER", len(lines) - 1 - index)
        )

    candidates: list[dict[str, Any]] = []
    seen: set[tuple[str, int, str, int, str]] = set()

    def append_candidate(
        *,
        raw: str,
        prefix: str | None,
        number: int,
        pattern_name: str,
        position: str,
        line_index: int,
        line: str,
        confidence: float,
        strong: bool,
    ) -> None:
        if number <= 0 or number > max_page_number:
            return
        if prefix is None and year_min <= number <= year_max:
            return
        key = (prefix or "", number, position, line_index, pattern_name)
        if key in seen:
            return
        seen.add(key)
        candidates.append({
            "raw": raw,
            "prefix": prefix,
            "number": number,
            "pattern": pattern_name,
            "position": position,
            "line_index": line_index,
            "evidence_line": line,
            "confidence": confidence,
            "strong": strong,
        })

    for line_index, line, position, edge_distance in indexed:
        normalized = normalize_line(line)
        if not normalized:
            continue

        exact_prefixed = patterns["exact_prefixed"].fullmatch(normalized)
        if exact_prefixed:
            append_candidate(
                raw=exact_prefixed.group(0),
                prefix=exact_prefixed.group("prefix"),
                number=int(exact_prefixed.group("number")),
                pattern_name="EXACT_PREFIXED",
                position=position,
                line_index=line_index,
                line=line,
                confidence=1.00,
                strong=True,
            )

        exact_chinese = patterns["exact_chinese"].fullmatch(normalized)
        if exact_chinese:
            append_candidate(
                raw=exact_chinese.group(0),
                prefix=None,
                number=int(exact_chinese.group("number")),
                pattern_name="EXACT_CHINESE",
                position=position,
                line_index=line_index,
                line=line,
                confidence=0.98,
                strong=True,
            )

        # Only scan the first few header lines and only capture digits
        # immediately after the literal prospectus keyword. Text after
        # the number is allowed, e.g. “招股说明书29第四节”.
        if position == "HEADER" and line_index < header_scan_lines:
            for keyword in prospectus_keywords:
                keyword_pos = normalized.find(keyword)
                if keyword_pos < 0:
                    continue
                tail = normalized[keyword_pos + len(keyword):].lstrip("：:")
                prefixed = patterns["header_prefixed"].match(tail)
                if prefixed:
                    append_candidate(
                        raw=prefixed.group("raw"),
                        prefix=prefixed.group("prefix"),
                        number=int(prefixed.group("number")),
                        pattern_name="PROSPECTUS_HEADER_PREFIXED",
                        position=position,
                        line_index=line_index,
                        line=line,
                        confidence=0.99,
                        strong=True,
                    )
                    break
                bare = patterns["header_bare"].match(tail)
                if bare:
                    append_candidate(
                        raw=bare.group("number"),
                        prefix=None,
                        number=int(bare.group("number")),
                        pattern_name="PROSPECTUS_HEADER_BARE",
                        position=position,
                        line_index=line_index,
                        line=line,
                        confidence=0.97,
                        strong=True,
                    )
                    break

        if position == "FOOTER" and edge_distance < bare_footer_lines:
            bare = patterns["exact_bare"].fullmatch(normalized)
            if bare:
                append_candidate(
                    raw=bare.group("number"),
                    prefix=None,
                    number=int(bare.group("number")),
                    pattern_name="EXACT_BARE",
                    position=position,
                    line_index=line_index,
                    line=line,
                    confidence=0.64,
                    strong=False,
                )

    priority = {
        "EXACT_PREFIXED": 5,
        "PROSPECTUS_HEADER_PREFIXED": 4,
        "PROSPECTUS_HEADER_BARE": 3,
        "EXACT_CHINESE": 2,
        "EXACT_BARE": 1,
    }
    candidates.sort(
        key=lambda item: (
            item["strong"],
            priority.get(item["pattern"], 0),
            item["confidence"],
        ),
        reverse=True,
    )
    return candidates


def select_initial_pagination(
    company_id: str,
    pdf_page: int,
    candidates: list[dict[str, Any]]
) -> PaginationRecord:
    if not candidates:
        return PaginationRecord(
            company_id, pdf_page, None, None, None, None,
            "NOT_DETECTED", "MISSING", 0.0, [], None, None
        )

    strong = [item for item in candidates if item["strong"]]
    if strong:
        top = strong[0]
        same_priority_conflicts = [
            item
            for item in strong[1:]
            if (
                item["number"] != top["number"]
                and item["pattern"] == top["pattern"]
                and abs(item["confidence"] - top["confidence"]) <= 0.01
            )
        ]
        if same_priority_conflicts:
            return PaginationRecord(
                company_id, pdf_page, top["raw"], top["prefix"],
                top["number"], top["position"],
                "MULTIPLE_EQUAL_PRIORITY_CANDIDATES",
                "CONFLICT", top["confidence"], candidates,
                top["evidence_line"],
                "同页存在多个同优先级高置信度页码候选"
            )
        return PaginationRecord(
            company_id, pdf_page, top["raw"], top["prefix"],
            top["number"], top["position"],
            "DIRECT_DETECTION", "CONFIRMED", top["confidence"],
            candidates, top["evidence_line"], None
        )

    top = candidates[0]
    return PaginationRecord(
        company_id, pdf_page, top["raw"], top["prefix"],
        top["number"], top["position"],
        "WEAK_FOOTER_CANDIDATE", "CANDIDATE",
        top["confidence"], candidates, top["evidence_line"],
        "独立裸数字页脚候选，等待相邻页连续性确认"
    )


def confirm_weak_sequences(
    records: list[PaginationRecord],
    minimum_length: int
) -> None:
    weak = [
        item for item in records
        if item.mapping_status == "CANDIDATE"
        and item.printed_page_number is not None
    ]
    by_offset: dict[tuple[str | None, int], list[PaginationRecord]] = defaultdict(list)
    for item in weak:
        offset = item.pdf_page - int(item.printed_page_number)
        by_offset[(item.printed_page_prefix, offset)].append(item)

    for (_, _), items in by_offset.items():
        ordered = sorted(items, key=lambda item: item.pdf_page)
        run: list[PaginationRecord] = []

        def confirm_run() -> None:
            if len(run) >= minimum_length:
                for record in run:
                    record.mapping_status = "CONFIRMED"
                    record.detection_status = "CONFIRMED_BY_NUMERIC_SEQUENCE"
                    record.confidence = max(record.confidence, 0.88)
                    record.review_reason = None

        for item in ordered:
            if not run:
                run = [item]
                continue
            previous = run[-1]
            if (
                item.pdf_page == previous.pdf_page + 1
                and int(item.printed_page_number)
                == int(previous.printed_page_number) + 1
            ):
                run.append(item)
            else:
                confirm_run()
                run = [item]
        confirm_run()

    confirmed_anchors = [
        item for item in records
        if item.mapping_status == "CONFIRMED"
        and item.printed_page_number is not None
    ]
    offsets = Counter(
        (
            item.printed_page_prefix,
            item.pdf_page - int(item.printed_page_number)
        )
        for item in confirmed_anchors
    )
    for item in records:
        if (
            item.mapping_status == "CANDIDATE"
            and item.printed_page_number is not None
        ):
            key = (
                item.printed_page_prefix,
                item.pdf_page - int(item.printed_page_number)
            )
            if offsets[key] >= minimum_length:
                item.mapping_status = "CONFIRMED"
                item.detection_status = "CONFIRMED_BY_ESTABLISHED_OFFSET"
                item.confidence = max(item.confidence, 0.84)
                item.review_reason = None
            else:
                item.mapping_status = "MISSING"
                item.detection_status = "WEAK_CANDIDATE_REJECTED"
                item.printed_page_raw = None
                item.printed_page_prefix = None
                item.printed_page_number = None
                item.source_position = None
                item.confidence = 0.0
                item.review_reason = None


def infer_pagination(
    records: list[PaginationRecord],
    minimum_anchors: int,
    dominant_min_share: float = 0.60
) -> list[dict[str, Any]]:
    anchors = [
        item for item in records
        if (
            item.mapping_status == "CONFIRMED"
            and item.printed_page_number is not None
        )
    ]
    if not anchors:
        return []

    groups: dict[tuple[str | None, int], list[PaginationRecord]] = defaultdict(list)
    for item in anchors:
        offset = item.pdf_page - int(item.printed_page_number)
        groups[(item.printed_page_prefix, offset)].append(item)

    by_prefix: dict[str | None, int] = Counter(
        item.printed_page_prefix for item in anchors
    )
    ranked_groups = sorted(
        groups.items(),
        key=lambda pair: (
            -len(pair[1]),
            min(item.pdf_page for item in pair[1]),
        ),
    )

    by_page = {item.pdf_page: item for item in records}
    segments: list[dict[str, Any]] = []

    for (prefix, offset), group in ranked_groups:
        prefix_total = max(by_prefix[prefix], 1)
        share = len(group) / prefix_total
        if len(group) < minimum_anchors or share < dominant_min_share:
            continue

        ordered = sorted(group, key=lambda item: item.pdf_page)
        start = ordered[0].pdf_page
        end = ordered[-1].pdf_page
        inferred_count = 0

        for pdf_page in range(start, end + 1):
            item = by_page[pdf_page]
            if item.mapping_status != "MISSING":
                continue
            number = pdf_page - offset
            if number <= 0:
                continue
            item.printed_page_raw = (
                f"{prefix}-{number}" if prefix else str(number)
            )
            item.printed_page_prefix = prefix
            item.printed_page_number = number
            item.source_position = None
            item.detection_status = "INFERRED_FROM_DOMINANT_OFFSET"
            item.mapping_status = "INFERRED"
            item.confidence = 0.86
            item.review_reason = None
            inferred_count += 1

        segments.append({
            "segment_id": (
                f"SEG-{ordered[0].company_id}-{len(segments)+1:03d}"
            ),
            "company_id": ordered[0].company_id,
            "pdf_page_start": start,
            "pdf_page_end": end,
            "printed_page_start": start - offset,
            "printed_page_end": end - offset,
            "printed_page_prefix": prefix,
            "offset": offset,
            "anchor_count": len(group),
            "anchor_share_within_prefix": round(share, 4),
            "inferred_count": inferred_count,
            "mapping_type": "dominant_offset_consensus",
        })

    return segments


def repeated_edge_lines(page_lines: dict[int, list[str]]) -> Counter[str]:
    counter: Counter[str] = Counter()
    for lines in page_lines.values():
        candidates = lines[:3] + lines[-3:]
        counter.update({
            normalize_line(line)
            for line in candidates
            if 2 <= len(normalize_line(line)) <= 70
        })
    return counter


def iter_heading_units(
    lines: list[str],
    is_toc: bool,
    rules: dict[str, Any]
) -> Iterable[tuple[int, int, str, list[str]]]:
    config = rules["heading"]
    max_chars = int(config["max_chars"])
    limit = len(lines) if is_toc else min(
        len(lines), int(config["body_scan_lines"])
    )

    for start in range(limit):
        first_raw = lines[start]
        first = normalize_line(first_raw)
        if first and len(first) <= max_chars:
            yield start, start, first_raw, ["single_line"]

        if (
            not first
            or not is_numbering_only(first)
            or start + 1 >= limit
        ):
            continue

        second_raw = lines[start + 1]
        second = normalize_line(second_raw)
        combined = f"{first_raw}{second_raw}"
        normalized = normalize_line(combined)

        if (
            second
            and len(normalized) <= max_chars
            and not any(mark in second for mark in SENTENCE_END_PUNCTUATION)
        ):
            yield start, start + 1, combined, ["merged_numbering_and_title"]


def direct_match_types(
    title_core: str,
    rules: dict[str, Any]
) -> tuple[list[str], dict[str, list[str]], list[str], list[list[str]], list[str]]:
    types: list[str] = []
    sources: dict[str, list[str]] = {}
    aliases_all: list[str] = []
    groups_all: list[list[str]] = []
    excluded: list[str] = []

    for chapter_type, config in rules["chapter_types"].items():
        excluded_terms = [
            normalize_line(term)
            for term in config.get("exclude_title_keywords", [])
            if normalize_line(term) in title_core
        ]
        if excluded_terms:
            excluded.append(chapter_type)
            continue

        aliases = [normalize_line(item) for item in config.get("aliases", [])]
        local_aliases = [
            alias for alias in aliases
            if title_core == alias
            or (alias in title_core and len(title_core) - len(alias) <= 8)
        ]
        local_groups = [
            group for group in config.get("trigger_groups", [])
            if all(normalize_line(term) in title_core for term in group)
        ]
        if local_aliases or local_groups:
            types.append(chapter_type)
            basis: list[str] = []
            if local_aliases:
                basis.append("ALIAS")
                aliases_all.extend(local_aliases)
            if local_groups:
                basis.append("TRIGGER_GROUP")
                groups_all.extend(local_groups)
            sources[chapter_type] = basis

    return sorted(set(types)), sources, sorted(set(aliases_all)), groups_all, sorted(set(excluded))


def base_heading_score(
    heading_class: str,
    match_sources: dict[str, list[str]],
    is_page_top: bool,
    heading_level: int | None,
    merged: bool
) -> tuple[float, list[str]]:
    score = 0.0
    signals: list[str] = []
    if heading_class == "BODY_HEADING":
        score += 0.18
        signals.append("body_heading")
    elif heading_class == "TOC_ENTRY":
        score += 0.06
        signals.append("toc_entry")

    all_sources = {source for values in match_sources.values() for source in values}
    if "ALIAS" in all_sources:
        score += 0.50
        signals.append("alias_match")
    elif "TRIGGER_GROUP" in all_sources:
        score += 0.30
        signals.append("composite_trigger_group_match")

    if is_page_top:
        score += 0.08
        signals.append("page_top")
    if heading_level is not None:
        score += 0.08
        signals.append("numbering_level")
    if merged:
        score += 0.02
        signals.append("merged_title")
    return round(min(score, 1.0), 3), signals


def build_heading_records(
    company_id: str,
    page_lines: dict[int, list[str]],
    rules: dict[str, Any]
) -> tuple[list[HeadingRecord], dict[int, bool]]:
    repeat_counter = repeated_edge_lines(page_lines)
    repeated_min = int(rules["heading"]["repeated_header_min_pages"])
    headings: list[HeadingRecord] = []
    toc_map: dict[int, bool] = {}
    sequence = 0
    seen: set[tuple[int, int, int, str]] = set()

    for pdf_page, lines in page_lines.items():
        is_toc, toc_signals = page_is_toc(lines, rules)
        toc_map[pdf_page] = is_toc
        for start, end, raw_text, unit_signals in iter_heading_units(
            lines, is_toc, rules
        ):
            normalized = normalize_line(raw_text)
            heading_level = detect_heading_level(normalized)
            core = strip_numbering(
                clean_toc_suffix(normalized) if is_toc else normalized
            )
            if not is_title_like(core, int(rules["heading"]["max_chars"])):
                continue
            if any(word in core for word in CROSS_REFERENCE_WORDS):
                continue

            repeat_count = repeat_counter[normalize_line(lines[start])]
            if not is_toc and repeat_count >= repeated_min and start <= 2:
                heading_class = "PAGE_HEADER"
            elif is_toc:
                heading_class = "TOC_ENTRY"
            elif (
                start < int(rules["heading"]["page_top_line_limit"])
                and heading_level is not None
            ):
                heading_class = "BODY_HEADING"
            else:
                heading_class = "UNKNOWN"

            types, sources, aliases, groups, excluded = direct_match_types(
                core, rules
            )
            if (
                heading_class == "UNKNOWN"
                and types
                and start < int(rules["heading"]["page_top_line_limit"])
                and aliases
            ):
                heading_class = "BODY_HEADING"

            generic_heading = (
                heading_class in {"BODY_HEADING", "TOC_ENTRY"}
                and heading_level is not None
            )
            if not generic_heading and not (
                types and heading_class in {"BODY_HEADING", "TOC_ENTRY"}
            ):
                continue

            key = (pdf_page, start, end, core)
            if key in seen:
                continue
            seen.add(key)

            score, score_signals = base_heading_score(
                heading_class,
                sources,
                start < int(rules["heading"]["page_top_line_limit"]),
                heading_level,
                any(signal.startswith("merged_") for signal in unit_signals)
            )
            sequence += 1
            headings.append(HeadingRecord(
                heading_candidate_id=f"HDG-{company_id}-{sequence:05d}",
                company_id=company_id,
                pdf_page=pdf_page,
                line_start_index=start,
                line_end_index=end,
                line_text=raw_text,
                normalized_text=normalized,
                title_core=core,
                heading_level=heading_level,
                heading_class=heading_class,
                is_page_top=start < int(rules["heading"]["page_top_line_limit"]),
                repeated_line_count=repeat_count,
                matched_chapter_types=types,
                match_sources=sources,
                matched_aliases=aliases,
                matched_trigger_groups=groups,
                excluded_types=excluded,
                confidence=score,
                signals=sorted(set(toc_signals + unit_signals + score_signals))
            ))
    return headings, toc_map


def assign_hierarchy(headings: list[HeadingRecord], page_count: int) -> None:
    body = sorted(
        [item for item in headings if item.heading_class == "BODY_HEADING"],
        key=lambda item: (item.pdf_page, item.line_start_index)
    )
    stack: list[HeadingRecord] = []
    for item in body:
        if item.heading_level is None:
            stack = []
        else:
            while (
                stack
                and stack[-1].heading_level is not None
                and int(stack[-1].heading_level) >= int(item.heading_level)
            ):
                stack.pop()
        if stack:
            item.parent_heading_id = stack[-1].heading_candidate_id
            item.parent_heading_text = stack[-1].line_text
            item.parent_heading_level = stack[-1].heading_level
        stack.append(item)

    for index, item in enumerate(body):
        end_page = page_count
        for later in body[index + 1:]:
            if item.heading_level is None:
                end_page = max(item.pdf_page, later.pdf_page - 1)
                break
            if (
                later.heading_level is not None
                and int(later.heading_level) <= int(item.heading_level)
            ):
                end_page = max(item.pdf_page, later.pdf_page - 1)
                break
        item.section_end_pdf_page = end_page


def section_text(
    page_lines: dict[int, list[str]],
    start_page: int,
    end_page: int,
    max_pages: int
) -> str:
    effective_end = min(end_page, start_page + max_pages - 1)
    return "".join(
        normalize_line(line)
        for page in range(start_page, effective_end + 1)
        for line in page_lines.get(page, [])
    )


def enrich_from_section_context(
    headings: list[HeadingRecord],
    page_lines: dict[int, list[str]],
    rules: dict[str, Any]
) -> None:
    max_scan = int(rules["candidate_selection"]["max_section_scan_pages"])
    for heading in headings:
        if heading.heading_class != "BODY_HEADING":
            continue
        end_page = heading.section_end_pdf_page or heading.pdf_page
        text = section_text(page_lines, heading.pdf_page, end_page, max_scan)

        for chapter_type, config in rules["chapter_types"].items():
            if chapter_type in heading.excluded_types:
                continue
            if any(
                normalize_line(term) in heading.title_core
                for term in config.get("exclude_title_keywords", [])
            ):
                if chapter_type not in heading.excluded_types:
                    heading.excluded_types.append(chapter_type)
                if chapter_type in heading.matched_chapter_types:
                    heading.matched_chapter_types.remove(chapter_type)
                    heading.match_sources.pop(chapter_type, None)
                continue

            anchors = [
                normalize_line(term)
                for term in config.get("title_anchor_terms", [])
            ]
            anchor_hits = [term for term in anchors if term and term in heading.title_core]
            context_keywords = [
                normalize_line(term)
                for term in config.get("context_keywords", [])
            ]
            context_count = sum(term in text for term in context_keywords if term)
            minimum = int(config.get("minimum_context_support", 1))

            if anchor_hits and context_count >= minimum:
                if chapter_type not in heading.matched_chapter_types:
                    heading.matched_chapter_types.append(chapter_type)
                heading.match_sources.setdefault(chapter_type, [])
                if "SECTION_CONTEXT" not in heading.match_sources[chapter_type]:
                    heading.match_sources[chapter_type].append("SECTION_CONTEXT")
                heading.signals.append(
                    f"section_context:{chapter_type}:{context_count}"
                )
                heading.confidence = min(1.0, heading.confidence + 0.24)

        heading.matched_chapter_types = sorted(set(heading.matched_chapter_types))
        heading.excluded_types = sorted(set(heading.excluded_types))
        heading.signals = sorted(set(heading.signals))


def count_context_support(
    page_lines: dict[int, list[str]],
    start_page: int,
    end_page: int,
    keywords: list[str],
    max_pages: int
) -> int:
    text = section_text(page_lines, start_page, end_page, max_pages)
    return sum(
        normalize_line(keyword) in text
        for keyword in keywords
        if normalize_line(keyword)
    )


def toc_support_count(
    heading: HeadingRecord,
    chapter_type: str,
    toc_headings: list[HeadingRecord]
) -> int:
    count = 0
    for toc in toc_headings:
        if chapter_type not in toc.matched_chapter_types:
            continue
        if toc.title_core == heading.title_core:
            count += 1
            continue
        shared = set(toc.matched_aliases) & set(heading.matched_aliases)
        if shared:
            count += 1
    return count


def evidence_excerpt(
    page_lines: dict[int, list[str]],
    start_page: int,
    end_page: int,
    max_chars: int = 480
) -> str:
    text = " ".join(
        line.strip()
        for page in range(start_page, min(end_page, start_page + 1) + 1)
        for line in page_lines.get(page, [])
    )
    return text[:max_chars]


def scope_adjustment(
    heading: HeadingRecord,
    chapter_type: str,
    span_pages: int,
    page_count: int,
    rules: dict[str, Any]
) -> tuple[float, list[str]]:
    config = rules["chapter_types"][chapter_type]
    selection = rules["candidate_selection"]
    adjustment = 0.0
    signals: list[str] = []

    if any(
        normalize_line(term) in heading.title_core
        for term in config.get("broad_scope_terms", [])
    ):
        adjustment += float(selection["broad_scope_bonus"])
        signals.append("broad_scope_bonus")

    if any(
        normalize_line(term) in heading.title_core
        for term in config.get("narrow_scope_terms", [])
    ):
        adjustment -= float(selection["narrow_scope_penalty"])
        signals.append("narrow_scope_penalty")

    if chapter_type in {"equity_history", "shareholders"} and span_pages <= 1:
        adjustment -= float(selection["single_page_master_penalty"])
        signals.append("single_page_master_penalty")

    if (
        heading.pdf_page > int(page_count * 0.75)
        and chapter_type in {"equity_history", "shareholders"}
    ):
        adjustment -= float(selection["late_document_penalty"])
        signals.append("late_document_penalty")
    return adjustment, signals


MAJOR_SECTION_PATTERN = re.compile(
    r"^第[一二三四五六七八九十百\d]+节"
)


def major_section_title_from_line(
    line: str,
    aliases: list[str]
) -> str | None:
    normalized = normalize_line(line)
    match = re.search(
        r"第[一二三四五六七八九十百\d]+节",
        normalized
    )
    if not match:
        return None
    tail = normalized[match.start():]
    if any(normalize_line(alias) in tail for alias in aliases):
        return tail
    return None


def locate_issuer_master_section(
    company_id: str,
    page_lines: dict[int, list[str]],
    page_count: int,
    rules: dict[str, Any]
) -> IssuerMasterSection:
    config = rules["issuer_master_section"]
    aliases = config["aliases"]
    scan_lines = int(config["scan_lines"])
    context_keywords = [
        normalize_line(item) for item in config["context_keywords"]
    ]
    minimum_context = int(config["minimum_context_support"])
    max_span = int(config["max_span_pages"])

    candidates: list[tuple[float, int, str, str, int]] = []
    for pdf_page in range(1, page_count + 1):
        lines = page_lines.get(pdf_page, [])
        is_toc, _ = page_is_toc(lines, rules)
        if is_toc:
            continue
        for line in lines[:scan_lines]:
            title = major_section_title_from_line(line, aliases)
            if not title:
                continue
            context = "".join(
                normalize_line(value)
                for page in range(pdf_page, min(page_count, pdf_page + 3) + 1)
                for value in page_lines.get(page, [])
            )
            support = sum(term in context for term in context_keywords)
            score = 0.70 + min(support, 5) * 0.05
            if support >= minimum_context:
                score += 0.05
            # A body section normally appears before the latter half of the document.
            if pdf_page <= int(page_count * 0.50):
                score += 0.05
            candidates.append((
                min(score, 1.0), pdf_page, title, line, support
            ))

    if not candidates:
        return IssuerMasterSection(
            company_id=company_id,
            status="NOT_FOUND",
            start_pdf_page=None,
            end_pdf_page=None,
            heading_text=None,
            evidence_line=None,
            context_support_count=0,
            confidence=0.0,
            review_reason="未找到“第X节 发行人基本情况”正文大章节",
        )

    candidates.sort(key=lambda item: (-item[0], item[1]))
    score, start_page, title, evidence, support = candidates[0]

    next_start: int | None = None
    for pdf_page in range(start_page + 1, page_count + 1):
        lines = page_lines.get(pdf_page, [])
        is_toc, _ = page_is_toc(lines, rules)
        if is_toc:
            continue
        found_next = False
        for line in lines[:scan_lines]:
            normalized = normalize_line(line)
            match = re.search(
                r"第[一二三四五六七八九十百\d]+节",
                normalized
            )
            if not match:
                continue
            tail = normalized[match.start():]
            if tail != title and not any(
                normalize_line(alias) in tail for alias in aliases
            ):
                next_start = pdf_page
                found_next = True
                break
        if found_next:
            break

    if next_start is not None:
        end_page = max(start_page, next_start - 1)
        boundary_status = "CONFIRMED_BY_NEXT_MAJOR_SECTION"
    else:
        end_page = min(page_count, start_page + max_span - 1)
        boundary_status = "CAPPED_BY_MAX_SPAN"

    reason = None
    if support < minimum_context:
        reason = "大章节标题已找到，但相邻页上下文支持较弱"
    elif boundary_status == "CAPPED_BY_MAX_SPAN":
        reason = "未找到下一个大章节标题，结束页按最大跨度截断"

    return IssuerMasterSection(
        company_id=company_id,
        status=boundary_status,
        start_pdf_page=start_page,
        end_pdf_page=end_page,
        heading_text=title,
        evidence_line=evidence,
        context_support_count=support,
        confidence=round(score, 3),
        review_reason=reason,
    )


def within_master_section(
    pdf_page: int,
    master: IssuerMasterSection
) -> bool:
    return bool(
        master.start_pdf_page is not None
        and master.end_pdf_page is not None
        and master.start_pdf_page <= pdf_page <= master.end_pdf_page
    )


def clip_candidate_to_master(
    item: ChapterCandidate,
    master: IssuerMasterSection,
    pagination_by_page: dict[int, PaginationRecord]
) -> ChapterCandidate | None:
    if not within_master_section(item.start_pdf_page, master):
        return None
    assert master.end_pdf_page is not None
    if item.end_pdf_page_candidate > master.end_pdf_page:
        item.end_pdf_page_candidate = master.end_pdf_page
        end_mapping = pagination_by_page.get(master.end_pdf_page)
        item.end_printed_page_raw_candidate = (
            end_mapping.printed_page_raw if end_mapping else None
        )
        item.end_printed_page_number_candidate = (
            end_mapping.printed_page_number if end_mapping else None
        )
        item.end_inference_status = (
            item.end_inference_status + "_CLIPPED_TO_ISSUER_MASTER"
        )
        item.signals = sorted(set(item.signals + ["issuer_master_clip"]))
    return item


def issuer_master_candidate(
    company_id: str,
    master: IssuerMasterSection,
    page_lines: dict[int, list[str]],
    pagination_by_page: dict[int, PaginationRecord]
) -> ChapterCandidate | None:
    if master.start_pdf_page is None or master.end_pdf_page is None:
        return None
    start_mapping = pagination_by_page.get(master.start_pdf_page)
    end_mapping = pagination_by_page.get(master.end_pdf_page)
    reasons = [master.review_reason] if master.review_reason else []
    return ChapterCandidate(
        chapter_candidate_id=f"MASTER-{company_id}-issuer_overview",
        company_id=company_id,
        chapter_type="issuer_overview",
        rank_within_type=0,
        is_primary=False,
        candidate_source="ISSUER_MASTER_SECTION",
        heading_candidate_id=None,
        heading_text=master.heading_text or "发行人基本情况",
        heading_level=0,
        parent_heading_id=None,
        parent_heading_text=None,
        supporting_heading_ids=[],
        supporting_heading_texts=[],
        start_pdf_page=master.start_pdf_page,
        start_printed_page_raw=(
            start_mapping.printed_page_raw if start_mapping else None
        ),
        start_printed_page_number=(
            start_mapping.printed_page_number if start_mapping else None
        ),
        end_pdf_page_candidate=master.end_pdf_page,
        end_printed_page_raw_candidate=(
            end_mapping.printed_page_raw if end_mapping else None
        ),
        end_printed_page_number_candidate=(
            end_mapping.printed_page_number if end_mapping else None
        ),
        end_inference_status=master.status,
        toc_support_count=0,
        context_support_count=master.context_support_count,
        match_basis=["ISSUER_MASTER_SECTION"],
        evidence_excerpt=evidence_excerpt(
            page_lines, master.start_pdf_page, master.end_pdf_page
        ),
        confidence=master.confidence,
        confidence_level="",
        signals=["issuer_master_section"],
        selection_status="UNRANKED",
        review_reasons=reasons,
    )


DATE_PATTERNS = [
    re.compile(r"(?:19|20)\d{2}年(?:\d{1,2}月(?:\d{1,2}日)?)?"),
    re.compile(r"(?:19|20)\d{2}[./-]\d{1,2}(?:[./-]\d{1,2})?"),
]


def date_tokens(text: str) -> set[str]:
    tokens: set[str] = set()
    for pattern in DATE_PATTERNS:
        tokens.update(pattern.findall(text))
    return tokens


def event_category_hits(
    text: str,
    rules: dict[str, Any]
) -> set[str]:
    hits: set[str] = set()
    categories = rules.get("event_window", {}).get("categories", {})
    for category, terms in categories.items():
        if any(normalize_line(term) in text for term in terms):
            hits.add(category)
    return hits


def equity_range_profile(
    page_lines: dict[int, list[str]],
    start_page: int,
    end_page: int,
    rules: dict[str, Any]
) -> dict[str, Any]:
    max_pages = int(
        rules.get("event_window", {}).get("max_window_pages", 24)
    )
    text = section_text(page_lines, start_page, end_page, max_pages)
    categories = event_category_hits(text, rules)
    dates = date_tokens(text)
    excluded_terms = [
        normalize_line(term)
        for term in rules.get("event_window", {}).get(
            "exclude_window_keywords", []
        )
        if normalize_line(term) in text
    ]
    return {
        "text": text,
        "categories": categories,
        "date_tokens": dates,
        "excluded_terms": excluded_terms,
    }


def valid_equity_candidate(
    item: ChapterCandidate,
    page_lines: dict[int, list[str]],
    rules: dict[str, Any]
) -> bool:
    static_terms = [
        normalize_line(term)
        for term in rules.get("event_window", {}).get(
            "static_equity_title_terms", []
        )
    ]
    title = normalize_line(item.heading_text)
    if any(term and term in title for term in static_terms):
        return False

    profile = equity_range_profile(
        page_lines,
        item.start_pdf_page,
        item.end_pdf_page_candidate,
        rules,
    )
    minimum_categories = int(
        rules["chapter_types"]["equity_history"].get(
            "minimum_event_categories", 2
        )
    )
    minimum_dates = int(
        rules["chapter_types"]["equity_history"].get(
            "minimum_date_count", 2
        )
    )
    return (
        len(profile["categories"]) >= minimum_categories
        and len(profile["date_tokens"]) >= minimum_dates
        and not profile["excluded_terms"]
    )


def valid_shareholder_candidate(item: ChapterCandidate) -> bool:
    title = normalize_line(item.heading_text)
    forbidden = (
        "股本和股东变化",
        "股东变化",
        "股本变化",
        "股本演变",
        "股权演变",
        "历史沿革",
        "设立以来",
    )
    return not any(term in title for term in forbidden)


def direct_candidates(
    company_id: str,
    chapter_type: str,
    headings: list[HeadingRecord],
    page_lines: dict[int, list[str]],
    pagination_by_page: dict[int, PaginationRecord],
    page_count: int,
    rules: dict[str, Any]
) -> list[ChapterCandidate]:
    body = [
        item for item in headings
        if item.heading_class == "BODY_HEADING"
        and chapter_type in item.matched_chapter_types
        and chapter_type not in item.excluded_types
    ]
    toc = [
        item for item in headings
        if item.heading_class == "TOC_ENTRY"
        and chapter_type in item.matched_chapter_types
    ]
    candidates: list[ChapterCandidate] = []
    max_scan = int(rules["candidate_selection"]["max_section_scan_pages"])

    for index, heading in enumerate(body, start=1):
        end_page = heading.section_end_pdf_page or heading.pdf_page
        span = end_page - heading.pdf_page + 1
        context_count = count_context_support(
            page_lines,
            heading.pdf_page,
            end_page,
            rules["chapter_types"][chapter_type].get("context_keywords", []),
            max_scan
        )
        toc_count = toc_support_count(heading, chapter_type, toc)
        score = heading.confidence
        signals = list(heading.signals)
        basis = list(heading.match_sources.get(chapter_type, []))

        if toc_count:
            score += float(rules["candidate_selection"]["toc_support_bonus"])
            signals.append("toc_cross_support")
        if context_count:
            score += float(rules["candidate_selection"]["context_support_bonus"])
            signals.append("section_context_support")
        adjustment, scope_signals = scope_adjustment(
            heading, chapter_type, span, page_count, rules
        )
        score += adjustment
        signals.extend(scope_signals)
        score = round(min(max(score, 0.0), 1.0), 3)

        start_mapping = pagination_by_page.get(heading.pdf_page)
        end_mapping = pagination_by_page.get(end_page)
        review_reasons: list[str] = []
        if (
            not start_mapping
            or start_mapping.mapping_status not in {"CONFIRMED", "INFERRED"}
        ):
            review_reasons.append("起始页正文页码未确认")
        if (
            not end_mapping
            or end_mapping.mapping_status not in {"CONFIRMED", "INFERRED"}
        ):
            review_reasons.append("结束页正文页码未确认")

        candidates.append(ChapterCandidate(
            chapter_candidate_id=f"RAW-{company_id}-{chapter_type}-{index:03d}",
            company_id=company_id,
            chapter_type=chapter_type,
            rank_within_type=0,
            is_primary=False,
            candidate_source="HEADING_SECTION",
            heading_candidate_id=heading.heading_candidate_id,
            heading_text=heading.line_text,
            heading_level=heading.heading_level,
            parent_heading_id=heading.parent_heading_id,
            parent_heading_text=heading.parent_heading_text,
            supporting_heading_ids=[heading.heading_candidate_id],
            supporting_heading_texts=[heading.line_text],
            start_pdf_page=heading.pdf_page,
            start_printed_page_raw=(
                start_mapping.printed_page_raw if start_mapping else None
            ),
            start_printed_page_number=(
                start_mapping.printed_page_number if start_mapping else None
            ),
            end_pdf_page_candidate=end_page,
            end_printed_page_raw_candidate=(
                end_mapping.printed_page_raw if end_mapping else None
            ),
            end_printed_page_number_candidate=(
                end_mapping.printed_page_number if end_mapping else None
            ),
            end_inference_status="HEADING_HIERARCHY_BOUNDARY",
            toc_support_count=toc_count,
            context_support_count=context_count,
            match_basis=sorted(set(basis)),
            evidence_excerpt=evidence_excerpt(
                page_lines, heading.pdf_page, end_page
            ),
            confidence=score,
            confidence_level="",
            signals=sorted(set(signals)),
            selection_status="UNRANKED",
            review_reasons=review_reasons
        ))
    return candidates


def build_group_candidate(
    company_id: str,
    chapter_type: str,
    members: list[ChapterCandidate],
    page_lines: dict[int, list[str]],
    pagination_by_page: dict[int, PaginationRecord],
    rules: dict[str, Any],
    source: str
) -> ChapterCandidate:
    ordered = sorted(members, key=lambda item: item.start_pdf_page)
    start_page = ordered[0].start_pdf_page
    end_page = max(item.end_pdf_page_candidate for item in ordered)
    start_mapping = pagination_by_page.get(start_page)
    end_mapping = pagination_by_page.get(end_page)
    score = max(item.confidence for item in ordered)
    if source == "SIBLING_GROUP":
        score += float(rules["candidate_selection"]["sibling_group_bonus"])
    else:
        score += float(rules["candidate_selection"]["event_cluster_bonus"])
    score = round(min(score, 1.0), 3)
    parent_text = next(
        (item.parent_heading_text for item in ordered if item.parent_heading_text),
        None
    )
    review_reasons: list[str] = []
    if not start_mapping or start_mapping.mapping_status not in {"CONFIRMED", "INFERRED"}:
        review_reasons.append("聚合范围起始页正文页码未确认")
    if not end_mapping or end_mapping.mapping_status not in {"CONFIRMED", "INFERRED"}:
        review_reasons.append("聚合范围结束页正文页码未确认")

    return ChapterCandidate(
        chapter_candidate_id=f"GROUP-{company_id}-{chapter_type}-{start_page}-{end_page}",
        company_id=company_id,
        chapter_type=chapter_type,
        rank_within_type=0,
        is_primary=False,
        candidate_source=source,
        heading_candidate_id=None,
        heading_text=parent_text or "；".join(
            item.heading_text for item in ordered[:3]
        ),
        heading_level=min(
            (item.heading_level for item in ordered if item.heading_level is not None),
            default=None
        ),
        parent_heading_id=next(
            (item.parent_heading_id for item in ordered if item.parent_heading_id),
            None
        ),
        parent_heading_text=parent_text,
        supporting_heading_ids=[
            heading_id
            for item in ordered
            for heading_id in item.supporting_heading_ids
        ],
        supporting_heading_texts=[
            text
            for item in ordered
            for text in item.supporting_heading_texts
        ],
        start_pdf_page=start_page,
        start_printed_page_raw=(
            start_mapping.printed_page_raw if start_mapping else None
        ),
        start_printed_page_number=(
            start_mapping.printed_page_number if start_mapping else None
        ),
        end_pdf_page_candidate=end_page,
        end_printed_page_raw_candidate=(
            end_mapping.printed_page_raw if end_mapping else None
        ),
        end_printed_page_number_candidate=(
            end_mapping.printed_page_number if end_mapping else None
        ),
        end_inference_status=f"{source}_BOUNDARY",
        toc_support_count=sum(item.toc_support_count for item in ordered),
        context_support_count=max(item.context_support_count for item in ordered),
        match_basis=sorted(set(
            basis for item in ordered for basis in item.match_basis
        )),
        evidence_excerpt=evidence_excerpt(page_lines, start_page, end_page),
        confidence=score,
        confidence_level="",
        signals=sorted(set(
            [source.lower()]
            + [signal for item in ordered for signal in item.signals]
        )),
        selection_status="UNRANKED",
        review_reasons=review_reasons
    )


def sibling_group_candidates(
    company_id: str,
    chapter_type: str,
    candidates: list[ChapterCandidate],
    page_lines: dict[int, list[str]],
    pagination_by_page: dict[int, PaginationRecord],
    rules: dict[str, Any]
) -> list[ChapterCandidate]:
    max_gap = int(rules["candidate_selection"]["max_sibling_gap_pages"])
    groups: dict[str, list[ChapterCandidate]] = defaultdict(list)
    for item in candidates:
        parent_key = item.parent_heading_id or f"NO_PARENT_{item.heading_level}"
        groups[parent_key].append(item)

    output: list[ChapterCandidate] = []
    for members in groups.values():
        ordered = sorted(members, key=lambda item: item.start_pdf_page)
        cluster: list[ChapterCandidate] = []
        for item in ordered:
            if not cluster:
                cluster = [item]
                continue
            if item.start_pdf_page - cluster[-1].end_pdf_page_candidate <= max_gap:
                cluster.append(item)
            else:
                if len(cluster) >= 2:
                    output.append(build_group_candidate(
                        company_id, chapter_type, cluster,
                        page_lines, pagination_by_page, rules, "SIBLING_GROUP"
                    ))
                cluster = [item]
        if len(cluster) >= 2:
            output.append(build_group_candidate(
                company_id, chapter_type, cluster,
                page_lines, pagination_by_page, rules, "SIBLING_GROUP"
            ))
    return output


def event_cluster_candidates(
    company_id: str,
    headings: list[HeadingRecord],
    page_lines: dict[int, list[str]],
    pagination_by_page: dict[int, PaginationRecord],
    rules: dict[str, Any]
) -> list[ChapterCandidate]:
    keywords = [normalize_line(item) for item in rules["event_heading_keywords"]]
    exclusions = [
        normalize_line(item)
        for item in rules["chapter_types"]["equity_history"]["exclude_title_keywords"]
    ]
    event_headings = [
        item for item in headings
        if item.heading_class == "BODY_HEADING"
        and any(keyword in item.title_core for keyword in keywords)
        and not any(term in item.title_core for term in exclusions)
    ]
    raw_members: list[ChapterCandidate] = []
    for index, heading in enumerate(event_headings, start=1):
        end_page = heading.section_end_pdf_page or heading.pdf_page
        context_count = count_context_support(
            page_lines,
            heading.pdf_page,
            end_page,
            rules["chapter_types"]["equity_history"]["context_keywords"],
            int(rules["candidate_selection"]["max_section_scan_pages"])
        )
        if context_count < 2:
            continue
        start_mapping = pagination_by_page.get(heading.pdf_page)
        end_mapping = pagination_by_page.get(end_page)
        raw_members.append(ChapterCandidate(
            chapter_candidate_id=f"EVENT-{company_id}-{index:03d}",
            company_id=company_id,
            chapter_type="equity_history",
            rank_within_type=0,
            is_primary=False,
            candidate_source="EVENT_HEADING",
            heading_candidate_id=heading.heading_candidate_id,
            heading_text=heading.line_text,
            heading_level=heading.heading_level,
            parent_heading_id=heading.parent_heading_id,
            parent_heading_text=heading.parent_heading_text,
            supporting_heading_ids=[heading.heading_candidate_id],
            supporting_heading_texts=[heading.line_text],
            start_pdf_page=heading.pdf_page,
            start_printed_page_raw=(
                start_mapping.printed_page_raw if start_mapping else None
            ),
            start_printed_page_number=(
                start_mapping.printed_page_number if start_mapping else None
            ),
            end_pdf_page_candidate=end_page,
            end_printed_page_raw_candidate=(
                end_mapping.printed_page_raw if end_mapping else None
            ),
            end_printed_page_number_candidate=(
                end_mapping.printed_page_number if end_mapping else None
            ),
            end_inference_status="EVENT_HEADING_BOUNDARY",
            toc_support_count=0,
            context_support_count=context_count,
            match_basis=["EVENT_HEADING"],
            evidence_excerpt=evidence_excerpt(page_lines, heading.pdf_page, end_page),
            confidence=round(min(0.52 + min(context_count, 6) * 0.04, 0.78), 3),
            confidence_level="",
            signals=["event_heading", f"context_count:{context_count}"],
            selection_status="UNRANKED",
            review_reasons=[]
        ))

    max_gap = int(rules["candidate_selection"]["max_event_cluster_gap_pages"])
    groups: dict[str, list[ChapterCandidate]] = defaultdict(list)
    for item in raw_members:
        parent_key = item.parent_heading_id or f"NO_PARENT_{item.heading_level}"
        groups[parent_key].append(item)

    output: list[ChapterCandidate] = []
    for members in groups.values():
        ordered = sorted(members, key=lambda item: item.start_pdf_page)
        cluster: list[ChapterCandidate] = []
        for item in ordered:
            if not cluster:
                cluster = [item]
                continue
            if item.start_pdf_page - cluster[-1].end_pdf_page_candidate <= max_gap:
                cluster.append(item)
            else:
                if len(cluster) >= 2:
                    output.append(build_group_candidate(
                        company_id, "equity_history", cluster,
                        page_lines, pagination_by_page, rules, "EVENT_CLUSTER"
                    ))
                cluster = [item]
        if len(cluster) >= 2:
            output.append(build_group_candidate(
                company_id, "equity_history", cluster,
                page_lines, pagination_by_page, rules, "EVENT_CLUSTER"
            ))
    return output


def event_window_candidates(
    company_id: str,
    page_lines: dict[int, list[str]],
    pagination_by_page: dict[int, PaginationRecord],
    page_count: int,
    rules: dict[str, Any],
    master: IssuerMasterSection
) -> list[ChapterCandidate]:
    config = rules.get("event_window", {})
    if master.start_pdf_page is None or master.end_pdf_page is None:
        return []

    search_start = master.start_pdf_page
    search_end = master.end_pdf_page
    seed_min_categories = int(config.get("seed_minimum_categories", 1))
    seed_min_dates = int(config.get("seed_minimum_date_count", 1))
    min_categories = int(config.get("minimum_categories", 2))
    min_dates = int(config.get("minimum_date_count", 2))
    max_gap = int(config.get("max_gap_pages", 2))
    pad_before = int(config.get("padding_pages_before", 0))
    pad_after = int(config.get("padding_pages_after", 1))
    max_window = int(config.get("max_window_pages", 20))

    seed_pages: list[int] = []
    for pdf_page in range(search_start, search_end + 1):
        text = "".join(
            normalize_line(line)
            for line in page_lines.get(pdf_page, [])
        )
        if not text:
            continue
        categories = event_category_hits(text, rules)
        dates = date_tokens(text)
        if (
            len(categories) >= seed_min_categories
            and len(dates) >= seed_min_dates
        ):
            seed_pages.append(pdf_page)

    clusters: list[list[int]] = []
    current: list[int] = []
    for pdf_page in seed_pages:
        if not current or pdf_page - current[-1] <= max_gap:
            current.append(pdf_page)
        else:
            clusters.append(current)
            current = [pdf_page]
    if current:
        clusters.append(current)

    output: list[ChapterCandidate] = []
    for index, cluster in enumerate(clusters, start=1):
        start = max(search_start, cluster[0] - pad_before)
        end = min(search_end, cluster[-1] + pad_after)
        if end - start + 1 > max_window:
            end = start + max_window - 1

        profile = equity_range_profile(page_lines, start, end, rules)
        if (
            len(profile["categories"]) < min_categories
            or len(profile["date_tokens"]) < min_dates
            or profile["excluded_terms"]
        ):
            continue

        start_mapping = pagination_by_page.get(start)
        end_mapping = pagination_by_page.get(end)
        # Prefer event windows near the beginning of the issuer section.
        relative_start = (
            (start - search_start)
            / max(search_end - search_start + 1, 1)
        )
        confidence = min(
            0.61
            + 0.05 * len(profile["categories"])
            + 0.012 * min(len(profile["date_tokens"]), 10)
            + max(0.0, 0.06 - relative_start * 0.08),
            0.90,
        )

        review_reasons = [
            "事件窗口位于发行人基本情况大章节内，需人工确认完整边界"
        ]
        if (
            not start_mapping
            or start_mapping.mapping_status not in {"CONFIRMED", "INFERRED"}
        ):
            review_reasons.append("事件窗口起始页正文页码未确认")
        if (
            not end_mapping
            or end_mapping.mapping_status not in {"CONFIRMED", "INFERRED"}
        ):
            review_reasons.append("事件窗口结束页正文页码未确认")

        output.append(ChapterCandidate(
            chapter_candidate_id=f"WINDOW-{company_id}-{index:03d}",
            company_id=company_id,
            chapter_type="equity_history",
            rank_within_type=0,
            is_primary=False,
            candidate_source="EVENT_WINDOW",
            heading_candidate_id=None,
            heading_text="发行人基本情况内正文事件窗口",
            heading_level=None,
            parent_heading_id=None,
            parent_heading_text=master.heading_text,
            supporting_heading_ids=[],
            supporting_heading_texts=[],
            start_pdf_page=start,
            start_printed_page_raw=(
                start_mapping.printed_page_raw if start_mapping else None
            ),
            start_printed_page_number=(
                start_mapping.printed_page_number if start_mapping else None
            ),
            end_pdf_page_candidate=end,
            end_printed_page_raw_candidate=(
                end_mapping.printed_page_raw if end_mapping else None
            ),
            end_printed_page_number_candidate=(
                end_mapping.printed_page_number if end_mapping else None
            ),
            end_inference_status="ISSUER_MASTER_EVENT_WINDOW",
            toc_support_count=0,
            context_support_count=len(profile["categories"]),
            match_basis=[
                "ISSUER_MASTER_EVENT_WINDOW",
                f"event_categories:{len(profile['categories'])}",
                f"date_tokens:{len(profile['date_tokens'])}",
            ],
            evidence_excerpt=evidence_excerpt(page_lines, start, end),
            confidence=round(max(confidence, 0.0), 3),
            confidence_level="",
            signals=[
                "issuer_master_restricted",
                "event_density_window",
                *sorted(profile["categories"]),
            ],
            selection_status="UNRANKED",
            review_reasons=review_reasons,
        ))

    return output


def candidate_key(item: ChapterCandidate) -> tuple[int, int, str]:
    return (
        item.start_pdf_page,
        item.end_pdf_page_candidate,
        item.chapter_type
    )


def rank_candidates(
    company_id: str,
    chapter_type: str,
    candidates: list[ChapterCandidate],
    rules: dict[str, Any]
) -> list[ChapterCandidate]:
    unique: dict[tuple[int, int, str], ChapterCandidate] = {}
    source_priority = rules["candidate_selection"].get("source_priority", {})
    for item in candidates:
        key = candidate_key(item)
        existing = unique.get(key)
        if existing is None:
            unique[key] = item
            continue
        item_key = (
            int(source_priority.get(item.candidate_source, 0)),
            item.confidence,
        )
        existing_key = (
            int(source_priority.get(existing.candidate_source, 0)),
            existing.confidence,
        )
        if item_key > existing_key:
            unique[key] = item

    ordered = sorted(
        unique.values(),
        key=lambda item: (
            -int(source_priority.get(item.candidate_source, 0)),
            -item.confidence,
            -(item.end_pdf_page_candidate - item.start_pdf_page + 1),
            item.start_pdf_page,
        )
    )[:int(rules["candidate_selection"]["top_k_per_type"])]

    for rank, item in enumerate(ordered, start=1):
        item.rank_within_type = rank
        item.is_primary = rank == 1
        item.chapter_candidate_id = (
            f"CHC-{company_id}-{chapter_type}-{rank:02d}"
        )
        item.selection_status = "PRIMARY" if rank == 1 else "ALTERNATIVE"
        item.confidence_level = confidence_level(item.confidence)
    return ordered


def confidence_level(score: float) -> str:
    if score >= 0.85:
        return "HIGH"
    if score >= 0.65:
        return "MEDIUM"
    return "LOW"


def build_all_candidates(
    company_id: str,
    headings: list[HeadingRecord],
    page_lines: dict[int, list[str]],
    pagination: list[PaginationRecord],
    page_count: int,
    rules: dict[str, Any],
    master: IssuerMasterSection
) -> list[ChapterCandidate]:
    by_page = {item.pdf_page: item for item in pagination}
    all_ranked: list[ChapterCandidate] = []

    for chapter_type in rules["chapter_types"]:
        direct = direct_candidates(
            company_id, chapter_type, headings, page_lines,
            by_page, page_count, rules
        )

        if chapter_type in {"equity_history", "shareholders"}:
            clipped: list[ChapterCandidate] = []
            for item in direct:
                clipped_item = clip_candidate_to_master(
                    item, master, by_page
                )
                if clipped_item is not None:
                    clipped.append(clipped_item)
            direct = clipped

        if chapter_type == "equity_history":
            direct = [
                item for item in direct
                if valid_equity_candidate(item, page_lines, rules)
            ]
        elif chapter_type == "shareholders":
            direct = [
                item for item in direct
                if valid_shareholder_candidate(item)
            ]

        pool = list(direct)

        if chapter_type == "issuer_overview":
            master_candidate = issuer_master_candidate(
                company_id, master, page_lines, by_page
            )
            if master_candidate:
                pool.append(master_candidate)

        if chapter_type in {"equity_history", "shareholders"}:
            sibling = sibling_group_candidates(
                company_id, chapter_type, direct,
                page_lines, by_page, rules
            )
            sibling_clipped: list[ChapterCandidate] = []
            for item in sibling:
                clipped_item = clip_candidate_to_master(
                    item, master, by_page
                )
                if clipped_item is not None:
                    sibling_clipped.append(clipped_item)
            sibling = sibling_clipped

            if chapter_type == "equity_history":
                sibling = [
                    item for item in sibling
                    if valid_equity_candidate(item, page_lines, rules)
                ]
            else:
                sibling = [
                    item for item in sibling
                    if valid_shareholder_candidate(item)
                ]
            pool.extend(sibling)

        if chapter_type == "equity_history":
            clusters = []
            for item in event_cluster_candidates(
                company_id, headings, page_lines, by_page, rules
            ):
                clipped_item = clip_candidate_to_master(
                    item, master, by_page
                )
                if (
                    clipped_item is not None
                    and valid_equity_candidate(
                        clipped_item, page_lines, rules
                    )
                ):
                    clusters.append(clipped_item)
            pool.extend(clusters)
            pool.extend(event_window_candidates(
                company_id,
                page_lines,
                by_page,
                page_count,
                rules,
                master,
            ))

        all_ranked.extend(
            rank_candidates(company_id, chapter_type, pool, rules)
        )

    return all_ranked


def build_review_items(
    company_id: str,
    pagination: list[PaginationRecord],
    candidates: list[ChapterCandidate],
    rules: dict[str, Any],
    master: IssuerMasterSection
) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    by_type: dict[str, list[ChapterCandidate]] = defaultdict(list)
    for item in candidates:
        by_type[item.chapter_type].append(item)

    required_primary = [
        sorted(
            by_type.get(chapter_type, []),
            key=lambda item: item.rank_within_type
        )[0]
        for chapter_type in rules["required_chapter_types"]
        if by_type.get(chapter_type)
    ]
    candidate_pages = {
        page
        for item in required_primary
        for page in range(
            item.start_pdf_page,
            item.end_pdf_page_candidate + 1
        )
    }
    pagination_by_page = {item.pdf_page: item for item in pagination}
    unresolved_candidate_pages = [
        page for page in sorted(candidate_pages)
        if (
            page not in pagination_by_page
            or pagination_by_page[page].mapping_status
            not in {"CONFIRMED", "INFERRED"}
        )
    ]
    conflicts = [
        item.pdf_page for item in pagination
        if item.mapping_status == "CONFLICT"
    ]
    if conflicts or unresolved_candidate_pages:
        items.append({
            "company_id": company_id,
            "review_type": "PAGINATION_SYSTEM_REVIEW",
            "record_id": f"PAGINATION-{company_id}",
            "pdf_page": (
                conflicts[0]
                if conflicts
                else unresolved_candidate_pages[0]
            ),
            "chapter_type": None,
            "auto_value": {
                "conflict_pages": conflicts,
                "unresolved_required_candidate_pages": (
                    unresolved_candidate_pages
                ),
                "issuer_master_section": asdict(master),
            },
            "reason": (
                "必需章节候选范围内存在正文页码未确认或冲突；"
                "每家公司合并为一条系统复核记录"
            ),
        })

    alternative_count = int(
        rules["candidate_selection"].get(
            "review_alternative_count", 2
        )
    )
    for chapter_type in rules["required_chapter_types"]:
        type_candidates = sorted(
            by_type.get(chapter_type, []),
            key=lambda item: item.rank_within_type
        )
        if not type_candidates:
            items.append({
                "company_id": company_id,
                "review_type": "REQUIRED_CHAPTER_PRIMARY_REVIEW",
                "record_id": f"MISSING-{company_id}-{chapter_type}",
                "pdf_page": master.start_pdf_page,
                "chapter_type": chapter_type,
                "auto_value": {
                    "primary": None,
                    "alternatives": [],
                    "issuer_master_section": asdict(master),
                },
                "reason": "公共规则未生成该必需章节候选，需人工定位",
            })
            continue

        primary = type_candidates[0]
        alternatives = type_candidates[1:1 + alternative_count]
        reasons: list[str] = []
        if primary.candidate_source in {"EVENT_CLUSTER", "EVENT_WINDOW"}:
            reasons.append("主范围由事件聚合产生")
        if primary.review_reasons:
            reasons.extend(primary.review_reasons)
        if alternatives:
            margin = primary.confidence - alternatives[0].confidence
            if margin <= float(
                rules["candidate_selection"]["ambiguity_margin"]
            ):
                reasons.append(
                    f"第一、第二候选置信度差为{margin:.3f}"
                )
        if master.review_reason:
            reasons.append(master.review_reason)
        if not reasons:
            reasons.append("必需章节主范围需进行一次人工冻结确认")

        items.append({
            "company_id": company_id,
            "review_type": "REQUIRED_CHAPTER_PRIMARY_REVIEW",
            "record_id": f"REQUIRED-{company_id}-{chapter_type}",
            "pdf_page": primary.start_pdf_page,
            "chapter_type": chapter_type,
            "auto_value": {
                "primary": asdict(primary),
                "alternatives": [
                    asdict(item) for item in alternatives
                ],
                "issuer_master_section": asdict(master),
                "review_questions": [
                    "起始PDF页是否准确",
                    "结束PDF页是否准确",
                    "标题或窗口是否属于发行人而非子公司/财务/募集资金",
                    "正文页码是否与页面披露一致",
                ],
            },
            "reason": "；".join(sorted(set(reasons))),
        })
    return items


def autosize_sheet(worksheet, max_width: int = 68) -> None:
    for cells in worksheet.columns:
        letter = get_column_letter(cells[0].column)
        width = max(
            (len(str(cell.value)) for cell in cells if cell.value is not None),
            default=0
        )
        worksheet.column_dimensions[letter].width = min(max(width + 2, 10), max_width)


def write_review_workbook(
    path: Path,
    review_items: list[dict[str, Any]],
    summaries: list[dict[str, Any]],
    primary_candidates: list[ChapterCandidate],
    master_sections: list[IssuerMasterSection]
) -> None:
    workbook = Workbook()
    queue = workbook.active
    queue.title = "人工复核队列"
    queue.append([
        "公司代码", "复核类型", "记录ID", "PDF页码", "章节类型",
        "Auto值", "复核原因", "人工状态", "人工结论",
        "确认起始PDF页", "确认结束PDF页", "证据正文页码",
        "证据原文"
    ])
    for item in review_items:
        value = item.get("auto_value")
        if isinstance(value, (dict, list)):
            value = json.dumps(value, ensure_ascii=False)
        queue.append([
            item.get("company_id"), item.get("review_type"),
            item.get("record_id"), item.get("pdf_page"),
            item.get("chapter_type"), value, item.get("reason"),
            "PENDING", None, None, None, None, None
        ])

    primary = workbook.create_sheet("主章节候选")
    primary.append([
        "公司代码", "章节类型", "候选ID", "候选来源",
        "标题原文", "父级标题", "支持标题", "起始PDF页",
        "起始正文页", "结束PDF页候选", "结束正文页候选",
        "置信度", "置信度等级", "匹配依据", "证据摘录", "信号"
    ])
    for item in primary_candidates:
        primary.append([
            item.company_id, item.chapter_type,
            item.chapter_candidate_id, item.candidate_source,
            item.heading_text, item.parent_heading_text,
            "；".join(item.supporting_heading_texts),
            item.start_pdf_page, item.start_printed_page_raw,
            item.end_pdf_page_candidate,
            item.end_printed_page_raw_candidate,
            item.confidence, item.confidence_level,
            "；".join(item.match_basis),
            item.evidence_excerpt, "；".join(item.signals)
        ])

    master_sheet = workbook.create_sheet("发行人基本情况范围")
    master_sheet.append([
        "公司代码", "状态", "标题原文", "起始PDF页", "结束PDF页",
        "上下文支持数", "置信度", "证据原文", "复核原因"
    ])
    for item in master_sections:
        master_sheet.append([
            item.company_id, item.status, item.heading_text,
            item.start_pdf_page, item.end_pdf_page,
            item.context_support_count, item.confidence,
            item.evidence_line, item.review_reason
        ])

    summary_sheet = workbook.create_sheet("公司摘要")
    summary_sheet.append([
        "公司代码", "PDF页数", "页面文本来源",
        "直接页码数", "推断页码数", "缺失页码数", "冲突页码数",
        "全PDF页码覆盖率", "必需候选范围页码覆盖率",
        "发行人基本情况起始页", "发行人基本情况结束页",
        "全部标题数", "目标正文标题数", "保留章节候选数",
        "必需主候选数", "复核项数", "公司状态"
    ])
    for item in summaries:
        summary_sheet.append([
            item["company_id"], item["pdf_page_count"],
            item["page_text_source"],
            item["pagination_confirmed_count"],
            item["pagination_inferred_count"],
            item["pagination_missing_count"],
            item["pagination_conflict_count"],
            item["pagination_coverage_rate"],
            item["required_candidate_pagination_coverage_rate"],
            item["issuer_master_start_pdf_page"],
            item["issuer_master_end_pdf_page"],
            item["heading_candidate_count"],
            item["target_body_heading_count"],
            item["chapter_candidate_count"],
            item["required_primary_count"],
            item["review_item_count"],
            item["company_status"]
        ])

    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3")
    )
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(
                    vertical="top", wrap_text=True
                )
        autosize_sheet(worksheet)

    queue.column_dimensions["F"].width = 70
    queue.column_dimensions["G"].width = 55
    queue.column_dimensions["M"].width = 65
    primary.column_dimensions["E"].width = 42
    primary.column_dimensions["G"].width = 58
    primary.column_dimensions["O"].width = 68
    primary.column_dimensions["P"].width = 58
    master_sheet.column_dimensions["C"].width = 40
    master_sheet.column_dimensions["H"].width = 65
    workbook.save(path)


def ensure_run_dirs(repo_root: Path, run_id: str) -> dict[str, Path]:
    dirs = {
        "data": repo_root / "data" / "chapter_location" / "runs" / run_id,
        "auto": repo_root / "auto_output" / "chapter_location" / "runs" / run_id,
        "validation": repo_root / "validation" / "chapter_location" / "runs" / run_id,
        "logs": repo_root / "logs" / "chapter_location" / "runs" / run_id,
        "logs_root": repo_root / "logs" / "chapter_location"
    }
    for path in dirs.values():
        path.mkdir(parents=True, exist_ok=True)
    return dirs


def run_chapter_location(
    input_dir: Path,
    repo_root: Path,
    workspace_dir: Path,
    rules_file: Path,
    expected_count: int | None,
    reuse_run_id: str | None = None
) -> int:
    run_id = make_run_id()
    started_at = now_iso()
    input_dir = input_dir.expanduser().resolve()
    repo_root = repo_root.expanduser().resolve()
    workspace_dir = workspace_dir.expanduser().resolve()
    rules_file = rules_file.expanduser().resolve()

    if not input_dir.is_dir():
        print(f"[ERROR] PDF输入目录不存在：{input_dir}", file=sys.stderr)
        return 2
    if not repo_root.is_dir():
        print(f"[ERROR] 仓库目录不存在：{repo_root}", file=sys.stderr)
        return 2
    if not rules_file.is_file():
        print(f"[ERROR] 规则文件不存在：{rules_file}", file=sys.stderr)
        return 2

    rules = load_json(rules_file)
    validate_rule_config(rules)
    dirs = ensure_run_dirs(repo_root, run_id)
    external_run_dir = workspace_dir / run_id
    external_run_dir.mkdir(parents=True, exist_ok=True)

    pdf_paths = sorted(input_dir.glob("*.pdf"), key=lambda item: item.name.casefold())
    all_page_manifest: list[dict[str, Any]] = []
    all_pagination: list[PaginationRecord] = []
    all_headings: list[HeadingRecord] = []
    all_candidates: list[ChapterCandidate] = []
    all_review: list[dict[str, Any]] = []
    all_segments: list[dict[str, Any]] = []
    all_master_sections: list[IssuerMasterSection] = []
    summaries: list[dict[str, Any]] = []
    logs: list[dict[str, Any]] = []
    failures = 0

    for index, pdf_path in enumerate(pdf_paths, start=1):
        company_id, short_name = parse_filename(pdf_path.name)
        if not company_id or not short_name:
            company_id = f"UNMATCHED-{index:03d}"
            short_name = pdf_path.stem
        print(f"[{index}/{len(pdf_paths)}] v0.5定位：{pdf_path.name}")

        company_output = external_run_dir / company_id
        company_output.mkdir(parents=True, exist_ok=True)
        reuse_path = (
            workspace_dir / reuse_run_id / company_id / "page_text.jsonl"
            if reuse_run_id else None
        )

        try:
            page_records, page_rows, text_source = load_or_extract_pages(
                pdf_path, company_id, short_name, reuse_path
            )
            page_count = len(page_rows)
            page_lines = {
                int(row["pdf_page"]): [
                    line.strip()
                    for line in str(row.get("text") or "").splitlines()
                    if line.strip()
                ]
                for row in page_rows
            }

            pagination: list[PaginationRecord] = []
            for pdf_page in range(1, page_count + 1):
                candidates = pagination_candidates_for_page(
                    page_lines.get(pdf_page, []), rules, page_count
                )
                pagination.append(select_initial_pagination(
                    company_id, pdf_page, candidates
                ))
            confirm_weak_sequences(
                pagination,
                int(rules["pagination"]["minimum_sequence_length"])
            )
            segments = infer_pagination(
                pagination,
                int(rules["pagination"]["minimum_inference_anchors"]),
                float(rules["pagination"].get("dominant_offset_min_share", 0.60))
            )

            master = locate_issuer_master_section(
                company_id, page_lines, page_count, rules
            )

            headings, _ = build_heading_records(company_id, page_lines, rules)
            assign_hierarchy(headings, page_count)
            enrich_from_section_context(headings, page_lines, rules)
            candidates = build_all_candidates(
                company_id, headings, page_lines, pagination, page_count, rules, master
            )
            review = build_review_items(
                company_id, pagination, candidates, rules, master
            )

            write_jsonl(company_output / "page_text.jsonl", page_rows)
            write_jsonl(
                company_output / "pagination_candidates_v05.jsonl",
                [asdict(item) for item in pagination]
            )
            write_jsonl(
                company_output / "heading_candidates_v05.jsonl",
                [asdict(item) for item in headings]
            )

            confirmed = sum(item.mapping_status == "CONFIRMED" for item in pagination)
            inferred = sum(item.mapping_status == "INFERRED" for item in pagination)
            missing = sum(item.mapping_status == "MISSING" for item in pagination)
            conflicts = sum(item.mapping_status == "CONFLICT" for item in pagination)
            required_primary = sum(
                item.is_primary and item.chapter_type in rules["required_chapter_types"]
                for item in candidates
            )
            target_body = sum(
                item.heading_class == "BODY_HEADING"
                and bool(item.matched_chapter_types)
                for item in headings
            )
            required_primary_items = [
                item for item in candidates
                if item.is_primary
                and item.chapter_type in rules["required_chapter_types"]
            ]
            required_candidate_pages = {
                page
                for item in required_primary_items
                for page in range(
                    item.start_pdf_page,
                    item.end_pdf_page_candidate + 1
                )
            }
            pagination_by_page = {
                item.pdf_page: item for item in pagination
            }
            mapped_required_pages = sum(
                pagination_by_page[page].mapping_status
                in {"CONFIRMED", "INFERRED"}
                for page in required_candidate_pages
                if page in pagination_by_page
            )
            required_candidate_pagination_coverage = round(
                mapped_required_pages
                / max(len(required_candidate_pages), 1),
                4
            )

            status = "READY" if not review else "REVIEW_REQUIRED"
            summary = {
                "company_id": company_id,
                "short_name": short_name,
                "pdf_page_count": page_count,
                "page_text_source": text_source,
                "pagination_confirmed_count": confirmed,
                "pagination_inferred_count": inferred,
                "pagination_missing_count": missing,
                "pagination_conflict_count": conflicts,
                "pagination_coverage_rate": round(
                    (confirmed + inferred) / max(page_count, 1), 4
                ),
                "mapping_segment_count": len(segments),
                "required_candidate_page_count": len(required_candidate_pages),
                "required_candidate_mapped_page_count": mapped_required_pages,
                "required_candidate_pagination_coverage_rate": required_candidate_pagination_coverage,
                "issuer_master_status": master.status,
                "issuer_master_start_pdf_page": master.start_pdf_page,
                "issuer_master_end_pdf_page": master.end_pdf_page,
                "heading_candidate_count": len(headings),
                "target_body_heading_count": target_body,
                "chapter_candidate_count": len(candidates),
                "required_primary_count": required_primary,
                "review_item_count": len(review),
                "company_status": status
            }

            all_page_manifest.extend(asdict(item) for item in page_records)
            all_pagination.extend(pagination)
            all_headings.extend(headings)
            all_candidates.extend(candidates)
            all_review.extend(review)
            all_segments.extend(segments)
            all_master_sections.append(master)
            summaries.append(summary)
            logs.append({
                "timestamp": now_iso(),
                "level": "INFO",
                "run_id": run_id,
                "company_id": company_id,
                "stage": "chapter_location_v0.5",
                "event": "company_completed",
                "message": pdf_path.name,
                "details": summary
            })
        except Exception as exc:
            failures += 1
            summary = {
                "company_id": company_id,
                "short_name": short_name,
                "pdf_page_count": None,
                "page_text_source": None,
                "pagination_confirmed_count": 0,
                "pagination_inferred_count": 0,
                "pagination_missing_count": 0,
                "pagination_conflict_count": 0,
                "pagination_coverage_rate": 0.0,
                "mapping_segment_count": 0,
                "required_candidate_page_count": 0,
                "required_candidate_mapped_page_count": 0,
                "required_candidate_pagination_coverage_rate": 0.0,
                "issuer_master_status": "FAILED",
                "issuer_master_start_pdf_page": None,
                "issuer_master_end_pdf_page": None,
                "heading_candidate_count": 0,
                "target_body_heading_count": 0,
                "chapter_candidate_count": 0,
                "required_primary_count": 0,
                "review_item_count": 1,
                "company_status": "FAILED"
            }
            summaries.append(summary)
            all_review.append({
                "company_id": company_id,
                "review_type": "COMPANY_PROCESSING_FAILED",
                "record_id": f"FAILED-{company_id}",
                "pdf_page": None,
                "chapter_type": None,
                "auto_value": None,
                "reason": f"{type(exc).__name__}: {exc}"
            })
            logs.append({
                "timestamp": now_iso(),
                "level": "ERROR",
                "run_id": run_id,
                "company_id": company_id,
                "stage": "chapter_location_v0.5",
                "event": "company_failed",
                "message": pdf_path.name,
                "details": {
                    "error_type": type(exc).__name__,
                    "error_message": str(exc),
                    "traceback": traceback.format_exc()
                }
            })

    count_mismatch = expected_count is not None and len(pdf_paths) != expected_count
    if failures:
        batch_status = "PARTIAL_FAILURE"
    elif count_mismatch:
        batch_status = "BLOCKED"
    elif all_review:
        batch_status = "READY_WITH_REVIEW"
    else:
        batch_status = "READY"

    primary = [item for item in all_candidates if item.is_primary]
    required_primary = [
        item for item in primary
        if item.chapter_type in rules["required_chapter_types"]
    ]
    required_expected = len(pdf_paths) * len(rules["required_chapter_types"])
    required_coverage = round(
        len(required_primary) / max(required_expected, 1), 4
    )

    metrics = {
        "metrics_version": "0.5",
        "run_id": run_id,
        "batch_status": batch_status,
        "expected_pdf_count": expected_count,
        "discovered_pdf_count": len(pdf_paths),
        "company_success_count": sum(
            item["company_status"] != "FAILED" for item in summaries
        ),
        "company_failure_count": failures,
        "company_review_required_count": sum(
            item["company_status"] == "REVIEW_REQUIRED" for item in summaries
        ),
        "total_pdf_pages": sum(item["pdf_page_count"] or 0 for item in summaries),
        "pagination_confirmed_count": sum(
            item["pagination_confirmed_count"] for item in summaries
        ),
        "pagination_inferred_count": sum(
            item["pagination_inferred_count"] for item in summaries
        ),
        "pagination_missing_count": sum(
            item["pagination_missing_count"] for item in summaries
        ),
        "pagination_conflict_count": sum(
            item["pagination_conflict_count"] for item in summaries
        ),
        "pagination_coverage_rate": round(
            (
                sum(item["pagination_confirmed_count"] for item in summaries)
                + sum(item["pagination_inferred_count"] for item in summaries)
            ) / max(sum(item["pdf_page_count"] or 0 for item in summaries), 1),
            4
        ),
        "required_candidate_page_count": sum(
            item["required_candidate_page_count"] for item in summaries
        ),
        "required_candidate_mapped_page_count": sum(
            item["required_candidate_mapped_page_count"] for item in summaries
        ),
        "required_candidate_pagination_coverage_rate": round(
            sum(
                item["required_candidate_mapped_page_count"]
                for item in summaries
            )
            / max(
                sum(
                    item["required_candidate_page_count"]
                    for item in summaries
                ),
                1
            ),
            4
        ),
        "issuer_master_section_found_count": sum(
            item["issuer_master_start_pdf_page"] is not None
            for item in summaries
        ),
        "heading_candidate_count": len(all_headings),
        "target_body_heading_count": sum(
            item["target_body_heading_count"] for item in summaries
        ),
        "retained_chapter_candidate_count": len(all_candidates),
        "primary_chapter_candidate_count": len(primary),
        "required_primary_candidate_count": len(required_primary),
        "required_primary_expected_count": required_expected,
        "required_primary_coverage_rate": required_coverage,
        "review_queue_count": len(all_review),
        "reuse_run_id": reuse_run_id,
        "note": (
            "指标仅评价页码映射和章节定位，不代表融资事件或PE/VC识别准确率。"
        )
    }

    write_json(
        dirs["data"] / "page_text_manifest.json",
        {
            "manifest_version": "0.5",
            "run_id": run_id,
            "generated_at": now_iso(),
            "reuse_run_id": reuse_run_id,
            "pages": all_page_manifest
        }
    )
    write_jsonl(
        dirs["auto"] / "pagination_mapping_auto.jsonl",
        [asdict(item) for item in all_pagination]
    )
    write_jsonl(
        dirs["auto"] / "issuer_master_sections_auto.jsonl",
        [asdict(item) for item in all_master_sections]
    )
    write_jsonl(
        dirs["auto"] / "heading_candidates_auto.jsonl",
        [asdict(item) for item in all_headings]
    )
    write_jsonl(
        dirs["auto"] / "chapter_candidates_topk_auto.jsonl",
        [asdict(item) for item in all_candidates]
    )
    write_jsonl(
        dirs["auto"] / "chapter_ranges_primary_auto.jsonl",
        [asdict(item) for item in primary]
    )
    write_json(
        dirs["validation"] / "pagination_validation.json",
        {
            "validation_version": "0.5",
            "run_id": run_id,
            "segments": all_segments,
            "issuer_master_sections": [asdict(item) for item in all_master_sections],
            "companies": summaries
        }
    )
    write_json(
        dirs["validation"] / "chapter_candidate_validation.json",
        {
            "validation_version": "0.5",
            "run_id": run_id,
            "required_chapter_types": rules["required_chapter_types"],
            "companies": summaries
        }
    )
    write_json(
        dirs["validation"] / "chapter_location_metrics.json",
        metrics
    )
    write_review_workbook(
        dirs["validation"] / "chapter_location_review_queue.xlsx",
        all_review, summaries, primary, all_master_sections
    )
    write_jsonl(dirs["logs"] / "chapter_location.jsonl", logs)
    write_json(
        dirs["logs"] / "run_manifest.json",
        {
            "run_id": run_id,
            "pipeline_version": PIPELINE_VERSION,
            "rules_version": rules["rules_version"],
            "rules_sha256": sha256_file(rules_file),
            "started_at": started_at,
            "completed_at": now_iso(),
            "batch_status": batch_status,
            "reuse_run_id": reuse_run_id,
            "auto_freeze_policy": "run_specific_immutable_directory"
        }
    )
    write_json(
        dirs["logs"] / "output_counts.json",
        {
            "run_id": run_id,
            "page_manifest_count": len(all_page_manifest),
            "pagination_record_count": len(all_pagination),
            "issuer_master_section_count": len(all_master_sections),
            "heading_candidate_count": len(all_headings),
            "retained_chapter_candidate_count": len(all_candidates),
            "primary_chapter_candidate_count": len(primary),
            "required_primary_candidate_count": len(required_primary),
            "required_candidate_page_count": metrics[
                "required_candidate_page_count"
            ],
            "required_candidate_mapped_page_count": metrics[
                "required_candidate_mapped_page_count"
            ],
            "required_candidate_pagination_coverage_rate": metrics[
                "required_candidate_pagination_coverage_rate"
            ],
            "review_queue_count": len(all_review)
        }
    )
    write_json(
        dirs["logs_root"] / "latest_run.json",
        {
            "run_id": run_id,
            "pipeline_version": PIPELINE_VERSION,
            "batch_status": batch_status,
            "validation_relative_path": (
                f"validation/chapter_location/runs/{run_id}/"
                "chapter_location_review_queue.xlsx"
            )
        }
    )

    print()
    print("页码映射与章节定位 v0.5 完成")
    print(f"运行ID：{run_id}")
    print(f"发现PDF：{len(pdf_paths)}")
    print(f"批次状态：{batch_status}")
    print(f"公司失败：{failures}")
    print(f"全PDF页码覆盖率：{metrics['pagination_coverage_rate']}")
    print(
        "必需候选范围页码覆盖率："
        f"{metrics['required_candidate_pagination_coverage_rate']}"
    )
    print(
        "发行人基本情况大章节："
        f"{metrics['issuer_master_section_found_count']}/{len(pdf_paths)}"
    )
    print(
        "必需主章节覆盖："
        f"{len(required_primary)}/{required_expected}"
    )
    print(f"保留章节候选：{len(all_candidates)}")
    print(f"人工复核项：{len(all_review)}")
    print(
        "复核队列："
        + str(dirs["validation"] / "chapter_location_review_queue.xlsx")
    )
    return 0 if batch_status not in {"BLOCKED", "PARTIAL_FAILURE"} else 1
