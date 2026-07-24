\
from __future__ import annotations

import csv
import hashlib
import json
import re
import sys
import traceback
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


PIPELINE_VERSION = "batch_check_v0.1"
FILENAME_PATTERN = re.compile(
    r"^(?P<security_code>\d{6})_(?P<short_name>.+?)_IPO招股说明书\.pdf$",
    flags=re.IGNORECASE,
)
MIN_TEXT_CHARS = 40


@dataclass
class PdfCheckRecord:
    sequence: int
    security_code: str | None
    short_name: str | None
    file_name: str
    relative_path: str
    size_bytes: int
    size_mb: float
    sha256: str
    filename_status: str
    pdf_read_status: str
    page_count: int | None
    encrypted: bool | None
    sampled_pages: list[int]
    sampled_page_count: int
    sampled_text_page_count: int
    text_layer_status: str
    identity_source: str
    identity_status: str
    readiness_status: str
    warnings: list[str]
    error_type: str | None
    error_message: str | None


def now_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def make_run_id() -> str:
    return datetime.now().astimezone().strftime("BATCHCHECK_%Y%m%d_%H%M%S")


def sha256_file(path: Path, chunk_size: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while chunk := handle.read(chunk_size):
            digest.update(chunk)
    return digest.hexdigest()


def parse_pdf_filename(file_name: str) -> tuple[str | None, str | None, str]:
    match = FILENAME_PATTERN.fullmatch(file_name.strip())
    if not match:
        return None, None, "UNMATCHED"

    return (
        match.group("security_code"),
        match.group("short_name").strip(),
        "MATCHED",
    )


def build_sample_indices(page_count: int) -> list[int]:
    if page_count <= 0:
        return []

    candidates = {
        0,
        1,
        2,
        page_count // 4,
        page_count // 2,
        (page_count * 3) // 4,
        page_count - 3,
        page_count - 2,
        page_count - 1,
    }
    return sorted(index for index in candidates if 0 <= index < page_count)


def normalize_text(text: str) -> str:
    return "".join(text.split())


def inspect_text_layer(
    reader: PdfReader,
    sampled_indices: list[int],
) -> tuple[int, str, list[str]]:
    text_page_count = 0
    warnings: list[str] = []

    for page_index in sampled_indices:
        try:
            text = reader.pages[page_index].extract_text() or ""
            if len(normalize_text(text)) >= MIN_TEXT_CHARS:
                text_page_count += 1
        except Exception as exc:  # 页面级失败不能中断整份PDF
            warnings.append(
                f"PDF第{page_index + 1}页抽样文本提取失败："
                f"{type(exc).__name__}: {exc}"
            )

    sampled_count = len(sampled_indices)
    if sampled_count == 0:
        status = "NOT_CHECKED"
    elif text_page_count == sampled_count:
        status = "LIKELY_TEXT"
    elif text_page_count == 0:
        status = "LIKELY_SCANNED_OR_NO_TEXT"
    else:
        status = "MIXED_OR_UNCERTAIN"

    return text_page_count, status, warnings


def inspect_pdf(path: Path, input_dir: Path, sequence: int) -> PdfCheckRecord:
    warnings: list[str] = []
    security_code, short_name, filename_status = parse_pdf_filename(path.name)

    if filename_status != "MATCHED":
        warnings.append(
            "文件名未匹配“6位证券代码_公司简称_IPO招股说明书.pdf”规范"
        )

    size_bytes = path.stat().st_size
    file_sha256 = sha256_file(path)

    try:
        reader = PdfReader(path, strict=False)
        encrypted = bool(reader.is_encrypted)

        if encrypted:
            warnings.append("PDF已加密，当前基础检查不尝试破解或绕过加密")
            return PdfCheckRecord(
                sequence=sequence,
                security_code=security_code,
                short_name=short_name,
                file_name=path.name,
                relative_path=path.relative_to(input_dir).as_posix(),
                size_bytes=size_bytes,
                size_mb=round(size_bytes / 1024 / 1024, 3),
                sha256=file_sha256,
                filename_status=filename_status,
                pdf_read_status="ENCRYPTED",
                page_count=None,
                encrypted=True,
                sampled_pages=[],
                sampled_page_count=0,
                sampled_text_page_count=0,
                text_layer_status="NOT_CHECKED",
                identity_source="filename_only",
                identity_status="pending_cover_verification",
                readiness_status="REVIEW_REQUIRED",
                warnings=warnings,
                error_type=None,
                error_message=None,
            )

        page_count = len(reader.pages)
        sampled_indices = build_sample_indices(page_count)
        text_page_count, text_layer_status, text_warnings = inspect_text_layer(
            reader,
            sampled_indices,
        )
        warnings.extend(text_warnings)

        readiness_status = "READY"
        if filename_status != "MATCHED":
            readiness_status = "REVIEW_REQUIRED"
        if page_count <= 0:
            readiness_status = "FAILED"
            warnings.append("PDF页数为0")
        if text_layer_status in {"LIKELY_SCANNED_OR_NO_TEXT", "MIXED_OR_UNCERTAIN"}:
            readiness_status = "REVIEW_REQUIRED"
        if text_warnings:
            readiness_status = "REVIEW_REQUIRED"

        return PdfCheckRecord(
            sequence=sequence,
            security_code=security_code,
            short_name=short_name,
            file_name=path.name,
            relative_path=path.relative_to(input_dir).as_posix(),
            size_bytes=size_bytes,
            size_mb=round(size_bytes / 1024 / 1024, 3),
            sha256=file_sha256,
            filename_status=filename_status,
            pdf_read_status="SUCCESS",
            page_count=page_count,
            encrypted=False,
            sampled_pages=[index + 1 for index in sampled_indices],
            sampled_page_count=len(sampled_indices),
            sampled_text_page_count=text_page_count,
            text_layer_status=text_layer_status,
            identity_source="filename_only",
            identity_status="pending_cover_verification",
            readiness_status=readiness_status,
            warnings=warnings,
            error_type=None,
            error_message=None,
        )

    except Exception as exc:
        warnings.append("PDF读取失败")
        return PdfCheckRecord(
            sequence=sequence,
            security_code=security_code,
            short_name=short_name,
            file_name=path.name,
            relative_path=path.relative_to(input_dir).as_posix(),
            size_bytes=size_bytes,
            size_mb=round(size_bytes / 1024 / 1024, 3),
            sha256=file_sha256,
            filename_status=filename_status,
            pdf_read_status="FAILED",
            page_count=None,
            encrypted=None,
            sampled_pages=[],
            sampled_page_count=0,
            sampled_text_page_count=0,
            text_layer_status="NOT_CHECKED",
            identity_source="filename_only",
            identity_status="pending_cover_verification",
            readiness_status="FAILED",
            warnings=warnings,
            error_type=type(exc).__name__,
            error_message=str(exc),
        )


def find_duplicate_values(
    records: list[PdfCheckRecord],
    field_name: str,
) -> dict[str, list[str]]:
    values: dict[str, list[str]] = {}
    for record in records:
        value = getattr(record, field_name)
        if not value:
            continue
        values.setdefault(str(value), []).append(record.file_name)
    return {
        value: file_names
        for value, file_names in values.items()
        if len(file_names) > 1
    }


def ensure_output_dirs(repo_root: Path) -> dict[str, Path]:
    dirs = {
        "data": repo_root / "data",
        "validation": repo_root / "validation",
        "logs": repo_root / "logs",
    }
    for directory in dirs.values():
        directory.mkdir(parents=True, exist_ok=True)
    return dirs


def write_json(path: Path, payload: Any) -> None:
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")


def write_manifest_csv(path: Path, records: list[PdfCheckRecord]) -> None:
    headers = [
        "sequence",
        "security_code",
        "short_name",
        "file_name",
        "relative_path",
        "size_bytes",
        "size_mb",
        "sha256",
        "filename_status",
        "pdf_read_status",
        "page_count",
        "encrypted",
        "sampled_pages",
        "sampled_page_count",
        "sampled_text_page_count",
        "text_layer_status",
        "identity_source",
        "identity_status",
        "readiness_status",
        "warnings",
        "error_type",
        "error_message",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for record in records:
            row = asdict(record)
            row["sampled_pages"] = ",".join(map(str, record.sampled_pages))
            row["warnings"] = "；".join(record.warnings)
            writer.writerow(row)


def build_company_registry(
    records: list[PdfCheckRecord],
    run_id: str,
) -> dict[str, Any]:
    companies = []
    unmatched_files = []

    for record in records:
        if record.security_code and record.short_name:
            companies.append(
                {
                    "company_id": record.security_code,
                    "security_code": record.security_code,
                    "company_name": None,
                    "short_name": record.short_name,
                    "aliases": [record.short_name],
                    "pdf_file_name": record.file_name,
                    "pdf_relative_path": record.relative_path,
                    "pdf_sha256": record.sha256,
                    "enabled": record.readiness_status != "FAILED",
                    "identity_source": "filename_only",
                    "identity_status": "pending_cover_verification",
                    "config_status": "not_created",
                }
            )
        else:
            unmatched_files.append(record.file_name)

    return {
        "registry_version": "0.1-draft",
        "generated_by": PIPELINE_VERSION,
        "run_id": run_id,
        "generated_at": now_iso(),
        "status": "DRAFT_PENDING_PDF_COVER_VERIFICATION",
        "companies": companies,
        "unmatched_files": unmatched_files,
    }


def autosize_sheet(worksheet, max_width: int = 55) -> None:
    for column_cells in worksheet.columns:
        letter = get_column_letter(column_cells[0].column)
        length = 0
        for cell in column_cells:
            if cell.value is not None:
                length = max(length, len(str(cell.value)))
        worksheet.column_dimensions[letter].width = min(max(length + 2, 10), max_width)


def write_excel_report(
    path: Path,
    records: list[PdfCheckRecord],
    summary: dict[str, Any],
) -> None:
    workbook = Workbook()
    inventory = workbook.active
    inventory.title = "PDF清单"

    headers = [
        "序号",
        "证券代码",
        "公司简称",
        "文件名",
        "相对路径",
        "大小(MB)",
        "SHA-256",
        "PDF页数",
        "是否加密",
        "文件名状态",
        "PDF读取状态",
        "抽样页",
        "抽样页数",
        "含文本抽样页数",
        "文本层初判",
        "身份来源",
        "身份确认状态",
        "就绪状态",
        "警告",
        "错误类型",
        "错误信息",
    ]
    inventory.append(headers)

    for record in records:
        inventory.append(
            [
                record.sequence,
                record.security_code,
                record.short_name,
                record.file_name,
                record.relative_path,
                record.size_mb,
                record.sha256,
                record.page_count,
                record.encrypted,
                record.filename_status,
                record.pdf_read_status,
                ",".join(map(str, record.sampled_pages)),
                record.sampled_page_count,
                record.sampled_text_page_count,
                record.text_layer_status,
                record.identity_source,
                record.identity_status,
                record.readiness_status,
                "；".join(record.warnings),
                record.error_type,
                record.error_message,
            ]
        )

    summary_sheet = workbook.create_sheet("批次摘要")
    summary_sheet.append(["指标", "结果"])
    for key, value in summary.items():
        if isinstance(value, (dict, list)):
            value = json.dumps(value, ensure_ascii=False)
        summary_sheet.append([key, value])

    review_sheet = workbook.create_sheet("待人工复核")
    review_sheet.append(
        ["证券代码", "公司简称", "文件名", "就绪状态", "复核原因"]
    )
    for record in records:
        if record.readiness_status != "READY":
            reasons = list(record.warnings)
            if record.error_message:
                reasons.append(record.error_message)
            review_sheet.append(
                [
                    record.security_code,
                    record.short_name,
                    record.file_name,
                    record.readiness_status,
                    "；".join(reasons) or "需要人工确认",
                ]
            )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        autosize_sheet(worksheet)

    inventory.column_dimensions["D"].width = 42
    inventory.column_dimensions["G"].width = 68
    inventory.column_dimensions["S"].width = 45
    inventory.column_dimensions["U"].width = 40
    review_sheet.column_dimensions["C"].width = 42
    review_sheet.column_dimensions["E"].width = 70

    for row_index in range(2, inventory.max_row + 1):
        inventory.cell(row=row_index, column=6).number_format = "0.000"

    workbook.save(path)


def make_summary(
    records: list[PdfCheckRecord],
    expected_count: int | None,
    duplicate_codes: dict[str, list[str]],
    duplicate_hashes: dict[str, list[str]],
) -> dict[str, Any]:
    discovered_count = len(records)
    failed_count = sum(r.readiness_status == "FAILED" for r in records)
    review_count = sum(r.readiness_status == "REVIEW_REQUIRED" for r in records)
    ready_count = sum(r.readiness_status == "READY" for r in records)
    filename_unmatched_count = sum(
        r.filename_status != "MATCHED" for r in records
    )

    batch_status = "READY"
    issues: list[str] = []

    if expected_count is not None and discovered_count != expected_count:
        batch_status = "BLOCKED"
        issues.append(
            f"发现PDF数量{discovered_count}与预期数量{expected_count}不一致"
        )
    if failed_count > 0:
        batch_status = "PARTIAL_FAILURE"
        issues.append(f"{failed_count}份PDF基础检查失败")
    elif review_count > 0 and batch_status == "READY":
        batch_status = "READY_WITH_REVIEW"
        issues.append(f"{review_count}份PDF需要人工复核")
    if duplicate_codes:
        batch_status = "BLOCKED"
        issues.append("检测到重复证券代码")
    if duplicate_hashes:
        batch_status = "BLOCKED"
        issues.append("检测到重复PDF文件哈希")
    if filename_unmatched_count > 0:
        if batch_status == "READY":
            batch_status = "READY_WITH_REVIEW"
        issues.append(f"{filename_unmatched_count}份PDF文件名不符合规范")

    return {
        "pipeline_version": PIPELINE_VERSION,
        "batch_status": batch_status,
        "expected_pdf_count": expected_count,
        "discovered_pdf_count": discovered_count,
        "ready_count": ready_count,
        "review_required_count": review_count,
        "failed_count": failed_count,
        "filename_unmatched_count": filename_unmatched_count,
        "duplicate_security_codes": duplicate_codes,
        "duplicate_pdf_hashes": duplicate_hashes,
        "issues": issues,
        "note": (
            "公司身份目前仅由文件名生成，必须在后续通过PDF封面或正文证据确认；"
            "文本层状态仅为抽样初判，不是全页OCR结论。"
        ),
    }


def run_batch_check(
    input_dir: Path,
    repo_root: Path,
    expected_count: int | None,
) -> int:
    started_at = now_iso()
    run_id = make_run_id()

    input_dir = input_dir.expanduser().resolve()
    repo_root = repo_root.expanduser().resolve()

    if not input_dir.exists() or not input_dir.is_dir():
        print(f"[ERROR] PDF输入目录不存在或不是目录：{input_dir}", file=sys.stderr)
        return 2

    if not repo_root.exists() or not repo_root.is_dir():
        print(f"[ERROR] 仓库根目录不存在或不是目录：{repo_root}", file=sys.stderr)
        return 2

    output_dirs = ensure_output_dirs(repo_root)
    pdf_paths = sorted(
        (path for path in input_dir.rglob("*.pdf") if path.is_file()),
        key=lambda path: path.name.casefold(),
    )

    records: list[PdfCheckRecord] = []
    log_rows: list[dict[str, Any]] = [
        {
            "timestamp": started_at,
            "level": "INFO",
            "run_id": run_id,
            "stage": "batch_check",
            "event": "run_started",
            "message": "开始批量PDF基础检查",
            "details": {
                "pipeline_version": PIPELINE_VERSION,
                "expected_count": expected_count,
            },
        }
    ]

    for sequence, pdf_path in enumerate(pdf_paths, start=1):
        print(f"[{sequence}/{len(pdf_paths)}] 检查：{pdf_path.name}")
        try:
            record = inspect_pdf(pdf_path, input_dir, sequence)
        except Exception as exc:  # 最外层保护，单公司失败不影响其余公司
            record = PdfCheckRecord(
                sequence=sequence,
                security_code=None,
                short_name=None,
                file_name=pdf_path.name,
                relative_path=pdf_path.relative_to(input_dir).as_posix(),
                size_bytes=pdf_path.stat().st_size,
                size_mb=round(pdf_path.stat().st_size / 1024 / 1024, 3),
                sha256="",
                filename_status="UNKNOWN",
                pdf_read_status="FAILED",
                page_count=None,
                encrypted=None,
                sampled_pages=[],
                sampled_page_count=0,
                sampled_text_page_count=0,
                text_layer_status="NOT_CHECKED",
                identity_source="filename_only",
                identity_status="pending_cover_verification",
                readiness_status="FAILED",
                warnings=["未捕获异常，已隔离当前PDF"],
                error_type=type(exc).__name__,
                error_message=str(exc),
            )
            log_rows.append(
                {
                    "timestamp": now_iso(),
                    "level": "ERROR",
                    "run_id": run_id,
                    "stage": "batch_check",
                    "event": "document_unhandled_error",
                    "message": pdf_path.name,
                    "details": {
                        "error_type": type(exc).__name__,
                        "error_message": str(exc),
                        "traceback": traceback.format_exc(),
                    },
                }
            )

        records.append(record)
        log_rows.append(
            {
                "timestamp": now_iso(),
                "level": (
                    "ERROR"
                    if record.readiness_status == "FAILED"
                    else "WARNING"
                    if record.readiness_status == "REVIEW_REQUIRED"
                    else "INFO"
                ),
                "run_id": run_id,
                "company_id": record.security_code,
                "stage": "batch_check",
                "event": "document_checked",
                "message": record.file_name,
                "details": {
                    "readiness_status": record.readiness_status,
                    "page_count": record.page_count,
                    "text_layer_status": record.text_layer_status,
                    "warning_count": len(record.warnings),
                },
            }
        )

    duplicate_codes = find_duplicate_values(records, "security_code")
    duplicate_hashes = find_duplicate_values(records, "sha256")
    summary = make_summary(
        records,
        expected_count,
        duplicate_codes,
        duplicate_hashes,
    )
    registry = build_company_registry(records, run_id)

    inventory_payload = {
        "inventory_version": "0.1",
        "generated_by": PIPELINE_VERSION,
        "run_id": run_id,
        "generated_at": now_iso(),
        "input_policy": {
            "path_storage": "relative_to_input_dir",
            "company_identity": "filename_only_pending_cover_verification",
            "text_layer_check": "sampled_pages_only",
        },
        "records": [asdict(record) for record in records],
    }

    validation_payload = {
        "validation_version": "0.1",
        "run_id": run_id,
        "generated_at": now_iso(),
        "summary": summary,
        "document_results": [
            {
                "security_code": record.security_code,
                "file_name": record.file_name,
                "readiness_status": record.readiness_status,
                "warnings": record.warnings,
                "error_type": record.error_type,
                "error_message": record.error_message,
            }
            for record in records
        ],
    }

    run_manifest = {
        "run_id": run_id,
        "pipeline_version": PIPELINE_VERSION,
        "started_at": started_at,
        "completed_at": now_iso(),
        "command": "batch-check",
        "input": {
            "pdf_count": len(pdf_paths),
            "expected_count": expected_count,
            "paths_persisted_as": "relative_paths_only",
        },
        "output_files": [
            "data/pdf_manifest.csv",
            "data/document_inventory.json",
            "data/company_registry.json",
            "validation/pdf_basic_check.json",
            "validation/batch_readiness_report.xlsx",
            "logs/file_hashes.json",
            "logs/batch_check.jsonl",
            "logs/run_manifest.json",
        ],
        "summary": summary,
    }

    file_hashes_payload = {
        "run_id": run_id,
        "hash_algorithm": "SHA-256",
        "files": [
            {
                "file_name": record.file_name,
                "relative_path": record.relative_path,
                "sha256": record.sha256,
                "size_bytes": record.size_bytes,
            }
            for record in records
        ],
    }

    write_manifest_csv(output_dirs["data"] / "pdf_manifest.csv", records)
    write_json(
        output_dirs["data"] / "document_inventory.json",
        inventory_payload,
    )
    write_json(
        output_dirs["data"] / "company_registry.json",
        registry,
    )
    write_json(
        output_dirs["validation"] / "pdf_basic_check.json",
        validation_payload,
    )
    write_excel_report(
        output_dirs["validation"] / "batch_readiness_report.xlsx",
        records,
        summary,
    )
    write_json(
        output_dirs["logs"] / "file_hashes.json",
        file_hashes_payload,
    )

    log_rows.append(
        {
            "timestamp": now_iso(),
            "level": (
                "ERROR"
                if summary["batch_status"] in {"BLOCKED", "PARTIAL_FAILURE"}
                else "INFO"
            ),
            "run_id": run_id,
            "stage": "batch_check",
            "event": "run_completed",
            "message": "批量PDF基础检查完成",
            "details": summary,
        }
    )
    write_jsonl(output_dirs["logs"] / "batch_check.jsonl", log_rows)
    write_json(output_dirs["logs"] / "run_manifest.json", run_manifest)

    print()
    print("批量PDF基础检查完成")
    print(f"运行ID：{run_id}")
    print(f"发现PDF：{len(records)}")
    print(f"批次状态：{summary['batch_status']}")
    print(f"可直接继续：{summary['ready_count']}")
    print(f"待人工复核：{summary['review_required_count']}")
    print(f"失败：{summary['failed_count']}")
    print(
        "报告："
        + str(
            output_dirs["validation"] / "batch_readiness_report.xlsx"
        )
    )

    return 0 if summary["batch_status"] not in {"BLOCKED", "PARTIAL_FAILURE"} else 1
